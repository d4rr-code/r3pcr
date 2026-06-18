import logging
import tempfile
from decimal import Decimal

from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from apps.shipments.models import Shipment, HSCode
from ..models import DutyComputation

logger = logging.getLogger('r3pcr.computation')

def _can_download_computation(user, shipment):
    is_assigned_declarant = user.role == 'declarant' and shipment.declarant == user
    is_consignee = user.role == 'consignee' and shipment.consignee == user
    is_supervisor = user.role == 'supervisor'
    return is_assigned_declarant or is_consignee or is_supervisor


def _num(value):
    return float(value or 0)


def _resolve_report_items(computation):
    items = computation.get_items()
    hs_ids = [item.get('hs_code_id') for item in items if item.get('hs_code_id')]
    hs_map = {str(hs.id): hs for hs in HSCode.objects.filter(id__in=hs_ids)}
    report_items = []
    for item in items:
        hs = hs_map.get(str(item.get('hs_code_id'))) or computation.hs_code
        report_items.append({
            'description': item.get('description', ''),
            'quantity': item.get('quantity', ''),
            'unit': item.get('unit', ''),
            'hs_code': hs.code if hs else '',
            'duty_rate': _num(item.get('duty_rate', computation.duty_rate)),
            'dutiable_value': _num(item.get('dv_php')),
            'cud': _num(item.get('cud')),
            'unit_price': item.get('unit_price', ''),
        })
    return report_items


def _summary_rows(computation):
    csf_php = (computation.csf_usd or 0) * (computation.exchange_rate or 0)
    boc_total = (computation.customs_duty or 0) + (computation.vat_amount or 0) + Decimal('130') + (computation.ipf or 0)
    return [
        ('Dutiable Value', computation.dutiable_value or 0),
        ('Bank Charges', computation.bank_charges or 0),
        ('Customs Duties', computation.customs_duty or 0),
        ('Brokerage Fee', computation.brokerage_fee or 0),
        ('Arrastre', computation.arrastre or 0),
        ('Wharfage', computation.wharfage or 0),
        ('Container Service Fee', csf_php),
        ('Customs Documentary Stamp', Decimal('130')),
        ('Import Processing Fee', computation.ipf or 0),
        ('Total Landed Cost', computation.total_landed_cost or 0),
        ('VAT', computation.vat_amount or 0),
        ('BOC Payable', boc_total),
    ]


def _download_excel_report(shipment, computation):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_details = wb.active
    ws_details.title = 'Shipment Details'
    ws_items = wb.create_sheet('Line Items')
    ws_summary = wb.create_sheet('ECDT Summary')

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(bold=True, color='FFFFFF')
    title_font = Font(bold=True, size=14, color='1E3A5F')
    bold = Font(bold=True)
    border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )
    money_format = '#,##0.00'

    consignee = shipment.consignee.company_name or shipment.consignee.get_full_name() or shipment.consignee.username
    declarant = shipment.declarant.get_full_name() or shipment.declarant.username if shipment.declarant else ''
    prepared = computation.computed_by.get_full_name() or computation.computed_by.username if computation.computed_by else ''

    ws_details['A1'] = 'RTripleJ Customs Brokerage'
    ws_details['A1'].font = title_font
    ws_details['A2'] = 'ECDT Computation Sheet'
    detail_rows = [
        ('HAWB / BOL', shipment.hawb_number),
        ('Consignee', consignee),
        ('Declarant', declarant),
        ('Date', computation.computed_at.strftime('%Y-%m-%d') if computation.computed_at else ''),
        ('Shipment Mode', computation.container_type or shipment.shipment_type or ''),
        ('Container No.', shipment.container_number or ''),
        ('Job Number', shipment.job_order_reference or ''),
        ('Import Type', shipment.get_import_type_display()),
        ('Invoice Currency', shipment.invoice_currency or 'USD'),
        ('Exchange Rate (to PHP)', _num(computation.exchange_rate)),
        ('Prepared By', prepared),
    ]
    for row, (label, value) in enumerate(detail_rows, 4):
        ws_details.cell(row=row, column=1, value=label).font = bold
        ws_details.cell(row=row, column=2, value=value)
    ws_details.column_dimensions['A'].width = 22
    ws_details.column_dimensions['B'].width = 38

    item_headers = ['Description', 'Quantity', 'Unit', 'HS Code', 'Duty Rate', 'Dutiable Value', 'CUD per Item']
    for col, label in enumerate(item_headers, 1):
        cell = ws_items.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for row, item in enumerate(_resolve_report_items(computation), 2):
        values = [
            item['description'], item['quantity'], item['unit'], item['hs_code'],
            item['duty_rate'], item['dutiable_value'], item['cud'],
        ]
        for col, value in enumerate(values, 1):
            cell = ws_items.cell(row=row, column=col, value=value)
            cell.border = border
            if col in (6, 7):
                cell.number_format = money_format
            if col == 5:
                cell.number_format = '0.00"%"'
    for col, width in enumerate([42, 12, 12, 18, 12, 18, 18], 1):
        ws_items.column_dimensions[get_column_letter(col)].width = width

    ws_summary['A1'] = 'ECDT Summary'
    ws_summary['A1'].font = title_font
    ws_summary['A3'] = 'Charge'
    ws_summary['B3'] = 'Amount'
    for cell in ws_summary[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    for row, (label, value) in enumerate(_summary_rows(computation), 4):
        ws_summary.cell(row=row, column=1, value=label).border = border
        amount_cell = ws_summary.cell(row=row, column=2, value=_num(value))
        amount_cell.border = border
        amount_cell.number_format = money_format
        if label in {'Total Landed Cost', 'BOC Payable'}:
            ws_summary.cell(row=row, column=1).font = bold
            amount_cell.font = bold
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ECDT_{shipment.hawb_number}.xlsx'
    wb.save(response)
    return response


def _download_pdf_report(request, shipment, computation):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError:
        messages.error(request, 'PDF export requires reportlab. Install project requirements, then try again.')
        return redirect('declarant:process', shipment_id=shipment.id)

    buffer = tempfile.SpooledTemporaryFile()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.45 * inch, leftMargin=0.45 * inch)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('RTripleJ Customs Brokerage', styles['Title']),
        Paragraph('ECDT Computation Sheet', styles['Heading2']),
        Spacer(1, 10),
    ]

    consignee = shipment.consignee.company_name or shipment.consignee.get_full_name() or shipment.consignee.username
    declarant = shipment.declarant.get_full_name() or shipment.declarant.username if shipment.declarant else ''
    details = [
        ['HAWB / BOL', shipment.hawb_number, 'Consignee', consignee],
        ['Declarant', declarant, 'Date', computation.computed_at.strftime('%Y-%m-%d') if computation.computed_at else ''],
        ['Shipment Mode', computation.container_type or shipment.shipment_type or '', 'Import Currency', shipment.invoice_currency or 'USD'],
        ['Container No.', shipment.container_number or '', 'Job Number', shipment.job_order_reference or ''],
        ['Exchange Rate (to PHP)', f'{_num(computation.exchange_rate):,.4f}', '', ''],
    ]
    detail_table = Table(details, colWidths=[1.1 * inch, 2.0 * inch, 1.1 * inch, 2.3 * inch])
    detail_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
        ('BACKGROUND', (2, 0), (2, -1), colors.whitesmoke),
    ]))
    story.extend([detail_table, Spacer(1, 12)])

    item_data = [['Description', 'Qty', 'Unit', 'HS Code', 'Duty %', 'D/V PHP', 'CUD']]
    for item in _resolve_report_items(computation):
        item_data.append([
            Paragraph(item['description'] or '', styles['BodyText']),
            item['quantity'], item['unit'], item['hs_code'],
            f"{item['duty_rate']:,.2f}", f"{item['dutiable_value']:,.2f}", f"{item['cud']:,.2f}",
        ])
    item_table = Table(item_data, colWidths=[2.25 * inch, 0.5 * inch, 0.5 * inch, 0.9 * inch, 0.6 * inch, 0.9 * inch, 0.9 * inch])
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
    ]))
    story.extend([Paragraph('Line Items', styles['Heading3']), item_table, Spacer(1, 12)])

    summary_data = [['Charge', 'Amount']]
    for label, value in _summary_rows(computation):
        summary_data.append([label, f"{_num(value):,.2f}"])
    summary_table = Table(summary_data, colWidths=[3.0 * inch, 1.5 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    story.extend([Paragraph('ECDT Summary', styles['Heading3']), summary_table])

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=ECDT_{shipment.hawb_number}.pdf'
    return response


@login_required
def download_computation(request, shipment_id):
    shipment = get_object_or_404(Shipment, id=shipment_id)
    if not _can_download_computation(request.user, shipment):
        messages.error(request, 'Access denied.')
        return redirect('accounts:login')

    computation = get_object_or_404(DutyComputation, shipment=shipment)
    if request.GET.get('format') == 'pdf':
        return _download_pdf_report(request, shipment, computation)
    return _download_excel_report(shipment, computation)
