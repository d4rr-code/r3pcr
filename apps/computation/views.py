import json
import os
import re
import tempfile
import threading
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count
from django.utils import timezone
from django.utils import timezone
from apps.shipments.models import Shipment, ShipmentDocument, HSCode, ShipmentHSCode, StatusLog
from apps.supervisor.models import SystemConfig
from apps.notifications.utils import notify_shipment_status_change
from .models import DutyComputation, ShipmentLineItem, ShippingAdvisory
from .ocr import process_document


# â”€â”€â”€ Lookup Tables â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def get_brokerage_fee(taxable_value):
    tv = float(taxable_value)
    if tv <= 10000:   return Decimal('1300')
    if tv <= 20000:   return Decimal('2000')
    if tv <= 30000:   return Decimal('2700')
    if tv <= 40000:   return Decimal('3300')
    if tv <= 50000:   return Decimal('3600')
    if tv <= 60000:   return Decimal('4000')
    if tv <= 100000:  return Decimal('4700')
    if tv <= 200000:  return Decimal('5300')
    # â‚±5,300 + 0.125% of excess above â‚±200,000
    excess = Decimal(str(round(tv - 200000, 2)))
    return Decimal('5300') + round(excess * Decimal('0.00125'), 2)


def get_ipf(taxable_value):
    tv = float(taxable_value)
    if tv <= 25000:  return Decimal('250')
    if tv <= 50000:  return Decimal('500')
    if tv <= 250000: return Decimal('750')
    if tv <= 500000: return Decimal('1000')
    if tv <= 750000: return Decimal('1500')
    return Decimal('2000')


def normalize_charge_mode(value, shipment_type=''):
    value = (value or shipment_type or '').strip().lower()
    if value in {'fcl', 'lcl', 'air'}:
        return value
    if value == 'sea':
        return 'lcl'
    return 'air' if shipment_type == 'air' else 'lcl'


def apply_transport_charges(charge_mode, arrastre, wharfage, gross_weight=0, volume_cbm=0):
    arrastre     = Decimal(str(arrastre     or 0))
    wharfage     = Decimal(str(wharfage     or 0))
    gross_weight = Decimal(str(gross_weight or 0))
    volume_cbm   = Decimal(str(volume_cbm   or 0))
    revenue_ton  = max(volume_cbm, gross_weight / Decimal('1000'))

    # Arrastre and wharfage are FLAT total amounts entered by the declarant.
    # Verified from RTripleJ CDT Excel: the declarant enters the actual
    # terminal charge for the shipment — NOT a per-ton rate to be multiplied.
    # (The ₱5,496 and ₱519.35 references are starting hints, not multipliers.)
    return arrastre, wharfage, revenue_ton


def _store_document_ocr(doc, fields, raw_text, quality):
    doc.ocr_text = raw_text or ''
    doc.ocr_fields_json = json.dumps(fields or {}, default=str)
    doc.ocr_quality = quality
    doc.ocr_ran_at = timezone.now()
    doc.save(update_fields=['ocr_text', 'ocr_fields_json', 'ocr_quality', 'ocr_ran_at'])


# â”€â”€â”€ Per-Item ECDT Formula â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def compute_ecdt(items_data, exchange_rate,
                 arrastre=0, wharfage=0, csf_php=0, bank_charges=0):
    """
    items_data keys: exw_usd, freight_usd, insurance_usd, duty_rate,
                     description, quantity, hs_code_id, gw, nw, pkgs
    D/V = EXW + Freight + Insurance  (no auto-3% O/C â€” matches client CDT tool)
    Total Landed Cost excludes VAT; VAT = 12% of Total Landed Cost
    Brokerage Fee: tiered table up to â‚±200,000, then +0.125% of excess
    """
    computed_items = []
    total_dv_php   = Decimal('0')
    total_cud      = Decimal('0')

    for i, item in enumerate(items_data):
        exw            = Decimal(str(item['exw_usd']))
        item_freight   = Decimal(str(item.get('freight_usd',   0) or 0))
        item_insurance = Decimal(str(item.get('insurance_usd', 0) or 0))
        duty_rate      = Decimal(str(item.get('duty_rate',     0) or 0))

        # D/V = EXW + Freight + Insurance
        dv_usd  = exw + item_freight + item_insurance
        dv_php  = dv_usd * exchange_rate
        cud     = dv_php * (duty_rate / Decimal('100'))
        total_dv_php += dv_php
        total_cud    += cud

        computed_items.append({
            'no':             i + 1,
            'description':    item.get('description', ''),
            'quantity':       item.get('quantity', ''),
            'unit':           item.get('unit', ''),
            'unit_price':     item.get('unit_price', ''),
            'hs_code_id':     item.get('hs_code_id', ''),
            'hs_code':        item.get('hs_code', ''),
            'duty_rate':      float(duty_rate),
            'exw':            float(round(exw, 2)),
            'item_freight':   float(round(item_freight, 2)),
            'item_insurance': float(round(item_insurance, 2)),
            'dv_usd':         float(round(dv_usd, 2)),
            'dv_php':         float(round(dv_php, 2)),
            'cud':            float(round(cud, 2)),
            'gw':             item.get('gw', ''),
            'nw':             item.get('nw', ''),
            'pkgs':           item.get('pkgs', ''),
        })

    taxable_value   = round(total_dv_php, 2)
    customs_duties  = round(total_cud, 2)
    brokerage_fee   = get_brokerage_fee(taxable_value)
    cds             = Decimal('130')
    ipf             = get_ipf(taxable_value)

    arrastre_d      = Decimal(str(arrastre     or 0))
    wharfage_d      = Decimal(str(wharfage     or 0))
    csf_d           = Decimal(str(csf_php      or 0))
    bank_charges_d  = Decimal(str(bank_charges or 0))

    # Total Landed Cost = DV + Bank Charges + CUD + BF + Arrastre + Wharfage + CDS + IPF
    # NOTE: CSF is NOT included in TLC â€” it appears only in the BOC fees total (FCL)
    total_landed_cost = round(
        taxable_value + bank_charges_d + customs_duties + brokerage_fee
        + cds + ipf + arrastre_d + wharfage_d, 2
    )

    # VAT = 12% of Total Landed Cost (matches client CDT Excel convention)
    vat = round(total_landed_cost * Decimal('0.12'), 2)

    # BOC total = CUD + VAT + CDS + IPF + CSF (for FCL).
    # Verified from RTripleJ ECDT_FCL.xlsx: CSF appears in the SUMMARY/TOTAL column.
    # For LCL/Air/Land, csf_d = 0 so this formula is safe across all modes.
    boc_total = round(customs_duties + vat + cds + ipf + csf_d, 2)

    summary = {
        'taxable_value':    taxable_value,
        'bank_charges':     bank_charges_d,
        'customs_duties':   customs_duties,
        'brokerage_fee':    brokerage_fee,
        'cds':              cds,
        'ipf':              ipf,
        'arrastre':         arrastre_d,
        'wharfage':         wharfage_d,
        'csf_php':          csf_d,
        'total_landed_cost': total_landed_cost,
        'vat_base':         total_landed_cost,   # stored as vat_base in model
        'vat':              vat,
        'boc_total':        boc_total,
    }
    return computed_items, summary


# â”€â”€â”€ OCR Merge Helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# Priority order per field: which document type to prefer when the same field
# appears in more than one document.
_OCR_FIELD_PRIORITY = {
    'declared_value':    ['invoice', 'packing_list'],
    'description':       ['invoice', 'packing_list', 'airway_bill'],
    'total_quantity':    ['packing_list', 'invoice'],
    'gross_weight':      ['airway_bill', 'packing_list'],
    'volume_cbm':        ['airway_bill', 'packing_list'],
    'hawb_number':       ['airway_bill'],
    'invoice_number':    ['invoice'],
    'invoice_date':      ['invoice'],
    'shipper_name':      ['invoice', 'airway_bill'],
    'country_of_origin': ['invoice'],
    'hs_code':           ['invoice', 'airway_bill'],
    'flight_number':     ['airway_bill'],
    'flight_date':       ['airway_bill'],
    'port_loading':      ['airway_bill'],
    'port_discharge':    ['airway_bill'],
    'port_origin':       ['airway_bill'],
    'port_destination':  ['airway_bill'],
    'origin':            ['airway_bill', 'invoice'],
    'destination':       ['airway_bill', 'invoice'],
    'consignee_name':    ['invoice'],
    'consignee_address': ['invoice'],
    'currency':          ['invoice'],
    'net_weight':        ['packing_list'],
    'num_packages':      ['packing_list'],
    'total_gross_weight':['airway_bill', 'packing_list'],
    'number_of_pieces':  ['airway_bill', 'packing_list'],
    'bol_number':        ['airway_bill'],
}

_DOC_LABEL = {
    'invoice':      'Invoice',
    'airway_bill':  'Airway Bill',
    'packing_list': 'Packing List',
}


def merge_ocr_results(results):
    """
    Merge OCR results from multiple documents into one dict.
    Each merged field: {'value': ..., 'confidence': ..., 'source': doc_type}
    Priority per field is defined in _OCR_FIELD_PRIORITY.
    """
    merged = {}
    all_fields = set()
    for doc_data in results.values():
        all_fields.update(doc_data.get('fields', {}).keys())

    for field in all_fields:
        priority = _OCR_FIELD_PRIORITY.get(field, list(results.keys()))
        # Try priority order first, then any remaining doc
        search_order = priority + [d for d in results if d not in priority]
        for doc_type in search_order:
            if doc_type not in results:
                continue
            fdata = results[doc_type].get('fields', {}).get(field)
            if fdata and isinstance(fdata, dict) and fdata.get('value'):
                merged[field] = {
                    'value':      fdata['value'],
                    'confidence': fdata.get('confidence', 0.0),
                    'source':     doc_type,
                }
                break

    return merged


# â”€â”€â”€ OCR Extract (single document â€” kept for fallback) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@login_required
def ocr_extract(request, shipment_id, doc_id):
    shipment = get_object_or_404(Shipment, id=shipment_id)

    # Only the assigned declarant may run OCR on a shipment's documents
    if request.user.role != 'declarant' or shipment.declarant != request.user:
        messages.error(request, 'Access denied.')
        return redirect('declarant:queue')

    doc = get_object_or_404(ShipmentDocument, id=doc_id, shipment=shipment)
    try:
        # Download file to a temp path (works for both local and S3/Supabase storage)
        ext = os.path.splitext(doc.file.name)[1] or '.pdf'
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            doc.file.open('rb')
            tmp.write(doc.file.read())
            doc.file.close()
            tmp_path = tmp.name
        try:
            print(f'[OCR] Starting: {doc.file.name} | type={doc.document_type}')
            fields, raw_text, quality = process_document(tmp_path, doc.document_type)
            _store_document_ocr(doc, fields, raw_text, quality)
            print(f'[OCR] Raw text length: {len(raw_text) if raw_text else 0} chars')
            print(f'[OCR] Fields returned: {list(fields.keys()) if fields else None}')

            if fields:
                line_items = fields.pop('__items__', [])
                request.session['ocr_fields']      = fields
                request.session['ocr_items']       = line_items
                request.session['ocr_shipment_id'] = shipment_id
                found    = sum(1 for v in fields.values() if isinstance(v, dict) and v.get('value'))
                item_msg = f', {len(line_items)} line items detected' if line_items else ''
                request.session['ocr_toast'] = ('success', f'OCR complete â€” {found} fields extracted{item_msg}.')
                print(f'[OCR] Success: {found} fields, {len(line_items)} items')
            else:
                request.session['ocr_toast'] = ('warning', 'OCR ran but found no structured fields. Fill in manually.')
                print(f'[OCR] No fields extracted. Raw text snippet: {repr(raw_text[:200]) if raw_text else "EMPTY"}')
        finally:
            os.unlink(tmp_path)
    except Exception as e:
        import traceback
        print(f'[OCR] Exception: {e}')
        traceback.print_exc()
        request.session['ocr_toast'] = ('error', f'OCR failed: {e}')
    return redirect('declarant:process', shipment_id=shipment_id)


# â”€â”€â”€ OCR Extract All (single button â€” merges all documents) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@login_required
def ocr_extract_all(request, shipment_id):
    """Run OCR on every invoice/airway_bill/packing_list document at once.
    Starts in a background thread and redirects immediately so the page
    doesn't block. The process page auto-refreshes until results appear."""
    shipment = get_object_or_404(Shipment, id=shipment_id)

    if request.user.role != 'declarant' or shipment.declarant != request.user:
        messages.error(request, 'Access denied.')
        return redirect('declarant:queue')

    documents = list(shipment.documents.filter(
        document_type__in=['invoice', 'airway_bill', 'packing_list']
    ))
    if not documents:
        request.session['ocr_toast'] = ('warning', 'No supported documents uploaded yet.')
        return redirect('declarant:process', shipment_id=shipment_id)

    def _run_all(docs):
        for doc in docs:
            doc_type = doc.document_type
            try:
                ext = os.path.splitext(doc.file.name)[1] or '.pdf'
                with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                    doc.file.open('rb')
                    tmp.write(doc.file.read())
                    doc.file.close()
                    tmp_path = tmp.name
                try:
                    print(f'[OCR-ALL] Processing {doc_type}: {doc.file.name}')
                    fields, raw_text, quality = process_document(tmp_path, doc_type)
                    _store_document_ocr(doc, fields, raw_text, quality)
                    found = sum(1 for v in (fields or {}).values()
                                if isinstance(v, dict) and v.get('value'))
                    print(f'[OCR-ALL] {doc_type}: quality={quality}, {found} fields, '
                          f'{len(raw_text or "")} chars')
                finally:
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass
            except Exception as e:
                import traceback
                traceback.print_exc()
                print(f'[OCR-ALL] Failed on {doc_type}: {e}')

    t = threading.Thread(target=_run_all, args=(documents,), daemon=True)
    t.start()

    request.session['ocr_toast'] = (
        'info',
        f'Scanning {len(documents)} document{"s" if len(documents) != 1 else ""}â€¦ '
        'Results will appear automatically in a few seconds.'
    )
    return redirect('declarant:process', shipment_id=shipment_id)


# â”€â”€â”€ Computation â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@login_required
def compute_shipment(request, shipment_id):
    shipment = get_object_or_404(Shipment, id=shipment_id)

    # Only the assigned declarant may compute duties for a shipment
    if request.user.role != 'declarant' or shipment.declarant != request.user:
        messages.error(request, 'Access denied.')
        return redirect('declarant:queue')

    hs_codes    = HSCode.objects.filter(is_active=True)
    existing    = DutyComputation.objects.filter(shipment=shipment).first()
    advisory_ex = getattr(shipment, 'shipping_advisory', None)
    documents   = shipment.documents.all()

    result      = None
    items       = None
    wmcda_scores      = None
    wmcda_recommended = None
    wmcda_breakdown   = None
    wmcda_explanation = None
    wmcda_history     = None

    # â”€â”€ Pull default exchange rate from SystemConfig â”€â”€
    try:
        default_rate = SystemConfig.objects.get(key='exchange_rate').value
    except SystemConfig.DoesNotExist:
        default_rate = '59.1480'

    if request.method == 'POST':
        try:
            exchange_rate   = Decimal(request.POST.get('exchange_rate', default_rate) or default_rate)
            arrastre        = Decimal(request.POST.get('arrastre',   '0') or '0')
            wharfage        = Decimal(request.POST.get('wharfage',   '0') or '0')
            csf_usd_val     = Decimal(request.POST.get('csf_usd',   '0') or '0')
            bank_charges    = Decimal(request.POST.get('bank_charges', '0') or '0')
            container_type  = (request.POST.get('container_type', '') or '').strip()
            charge_mode     = normalize_charge_mode(request.POST.get('charge_mode'), shipment.shipment_type)
            cargo_volume    = Decimal(request.POST.get('cargo_volume', '0') or '0')
            gross_weight    = Decimal(str(shipment.gross_weight or 0))
            arrastre, wharfage, revenue_ton = apply_transport_charges(
                charge_mode, arrastre, wharfage,
                gross_weight=gross_weight,
                volume_cbm=cargo_volume,
            )
            csf_php_val     = csf_usd_val * exchange_rate

            # â”€â”€ Server-side port fee defaults (only when declarant left all at 0) â”€â”€
            # This mirrors the JS auto-fill so submissions without JS still get
            # the correct defaults applied.
            _stype = (shipment.shipment_type or '').lower()
            _csize = container_type.lower() if container_type else ''
            if arrastre == Decimal('0') and wharfage == Decimal('0'):
                if _stype in ('lcl', 'sea'):
                    arrastre    = Decimal('5496.00')
                    wharfage    = Decimal('519.35')
                    csf_usd_val = Decimal('0.00')
                    csf_php_val = Decimal('0.00')
                elif _stype == 'fcl':
                    if '40' in _csize:
                        arrastre    = Decimal('12608.00')
                        wharfage    = Decimal('779.05')
                        csf_usd_val = Decimal('10.00')
                    else:                              # default: 20FT
                        arrastre    = Decimal('5496.00')
                        wharfage    = Decimal('519.35')
                        csf_usd_val = Decimal('5.00')
                    csf_php_val = csf_usd_val * exchange_rate
                # AIR / LAND: leave at 0

            descriptions  = request.POST.getlist('description[]')
            exw_values    = request.POST.getlist('exw_value[]')
            freights_list = request.POST.getlist('item_freight[]')
            ins_list      = request.POST.getlist('item_insurance[]')
            quantities    = request.POST.getlist('quantity[]')
            units         = request.POST.getlist('unit[]')
            unit_prices   = request.POST.getlist('unit_price[]')
            hs_code_ids   = request.POST.getlist('hs_code_id[]')
            duty_rates    = request.POST.getlist('item_duty_rate[]')
            gws           = request.POST.getlist('gw[]')
            nws           = request.POST.getlist('nw[]')
            pkgs_list     = request.POST.getlist('pkgs[]')

            # Pad all lists to same length as descriptions
            n = len(descriptions)
            def _pad(lst, default=''):
                return (lst + [default] * n)[:n]
            freights_list = _pad(freights_list, '0')
            ins_list      = _pad(ins_list,      '0')
            hs_code_ids   = _pad(hs_code_ids,   '')
            duty_rates    = _pad(duty_rates,     '0')
            gws           = _pad(gws,            '')
            nws           = _pad(nws,            '')
            pkgs_list     = _pad(pkgs_list,      '')
            quantities    = _pad(quantities,     '')
            units         = _pad(units,          '')
            unit_prices   = _pad(unit_prices,    '')

            # Build HS code string lookup map (id â†’ code string)
            valid_hs_ids = [int(h) for h in hs_code_ids if h and h.strip().isdigit()]
            hs_code_map  = {
                str(obj.id): obj.code
                for obj in HSCode.objects.filter(id__in=valid_hs_ids).only('id', 'code')
            } if valid_hs_ids else {}

            items_data = [
                {
                    'description':    d.strip(),
                    'exw_usd':        e,
                    'freight_usd':    f  or '0',
                    'insurance_usd':  ins or '0',
                    'quantity':       q,
                    'unit':           unit,
                    'unit_price':     unit_price,
                    'hs_code_id':     h,
                    'hs_code':        hs_code_map.get(str(h).strip(), ''),
                    'duty_rate':      dr or '0',
                    'gw':             gw,
                    'nw':             nw,
                    'pkgs':           pk,
                }
                for d, e, f, ins, q, unit, unit_price, h, dr, gw, nw, pk
                in zip(descriptions, exw_values, freights_list, ins_list,
                       quantities, units, unit_prices, hs_code_ids, duty_rates, gws, nws, pkgs_list)
                if e and float(e) > 0
            ]
            if not items_data:
                messages.error(request, 'Add at least one item with a value.')
                raise ValueError('no items')

            # â”€â”€ Proportional freight/insurance server-side distribution â”€â”€â”€â”€â”€â”€
            # If the declarant entered a global total_freight / total_insurance
            # but left all per-item values at 0, distribute proportionally by EXW.
            total_exw_for_dist = sum(Decimal(str(it['exw_usd'])) for it in items_data) or Decimal('1')
            total_freight_global   = Decimal(request.POST.get('total_freight',   '0') or '0')
            total_insurance_global = Decimal(request.POST.get('total_insurance', '0') or '0')
            all_fr_zero  = all(Decimal(str(it.get('freight_usd',   0) or 0)) == 0 for it in items_data)
            all_ins_zero = all(Decimal(str(it.get('insurance_usd', 0) or 0)) == 0 for it in items_data)
            if all_fr_zero and total_freight_global > 0:
                for it in items_data:
                    prop = Decimal(str(it['exw_usd'])) / total_exw_for_dist
                    it['freight_usd'] = float(round(total_freight_global * prop, 4))
            if all_ins_zero and total_insurance_global > 0:
                for it in items_data:
                    prop = Decimal(str(it['exw_usd'])) / total_exw_for_dist
                    it['insurance_usd'] = float(round(total_insurance_global * prop, 4))

            items, summary = compute_ecdt(
                items_data, exchange_rate,
                arrastre=arrastre, wharfage=wharfage, csf_php=csf_php_val,
                bank_charges=bank_charges
            )

            # Totals for model storage
            total_freight   = sum(Decimal(str(it.get('freight_usd',   0) or 0)) for it in items_data)
            total_insurance = sum(Decimal(str(it.get('insurance_usd', 0) or 0)) for it in items_data)
            result = summary

            # Use first item's HS code + duty rate for model-level fields
            first_hs_id   = next((it['hs_code_id'] for it in items_data if it.get('hs_code_id')), None)
            first_dr      = Decimal(str(items_data[0].get('duty_rate', 0) or 0)) if items_data else Decimal('0')
            hs_code = None
            if first_hs_id:
                try:
                    hs_code = HSCode.objects.get(id=first_hs_id)
                except HSCode.DoesNotExist:
                    pass

            total_exw = sum(Decimal(str(it['exw_usd'])) for it in items_data)

            DutyComputation.objects.update_or_create(
                shipment=shipment,
                defaults={
                    'hs_code':           hs_code,
                    'total_freight':     total_freight,
                    'total_insurance':   total_insurance,
                    'exchange_rate':     exchange_rate,
                    'duty_rate':         first_dr,
                    'declared_value':    total_exw,
                    'items_json':        json.dumps(items),
                    'dutiable_value':    summary['taxable_value'],
                    'customs_duty':      summary['customs_duties'],
                    'vat_base':          summary['vat_base'],
                    'vat_amount':        summary['vat'],
                    'brokerage_fee':     summary['brokerage_fee'],
                    'ipf':               summary['ipf'],
                    'bank_charges':      bank_charges,
                    'arrastre':          arrastre,
                    'wharfage':          wharfage,
                    'csf_usd':           csf_usd_val,
                    'container_type':    container_type or charge_mode,
                    'total_landed_cost': summary['total_landed_cost'],
                    'computed_by':       request.user,
                }
            )

            if shipment.status == 'arrived':
                old_status = shipment.status
                shipment.status = 'computed'
                if not shipment.processed_at:
                    shipment.processed_at = timezone.now()
                shipment.save(update_fields=['status', 'processed_at', 'updated_at'])
                StatusLog.objects.create(
                    shipment=shipment,
                    changed_by=request.user,
                    old_status=old_status,
                    new_status='computed',
                    notes='Duties and taxes computation completed.',
                )
                notify_shipment_status_change(
                    shipment=shipment,
                    old_status=old_status,
                    new_status='computed',
                    changed_by=request.user,
                    notes='Duties and taxes computation completed.',
                )

            # â”€â”€ Auto-run WMCDA alongside ECDT â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            try:
                wmcda_weight   = float(shipment.gross_weight or 0)
                wmcda_volume   = float(request.POST.get('cargo_volume', 0) or 0)
                wmcda_value    = float(total_exw)
                wmcda_urgency  = shipment.urgency or 'normal'
                wmcda_distance = float(request.POST.get('distance_km', 2600) or 2600)

                wmcda_scores, wmcda_recommended, wmcda_breakdown, wmcda_explanation = compute_wmcda(
                    wmcda_weight, wmcda_volume, wmcda_value, wmcda_urgency, wmcda_distance
                )

                ShippingAdvisory.objects.update_or_create(
                    shipment=shipment,
                    defaults={
                        'gross_weight':     wmcda_weight,
                        'cargo_volume':     wmcda_volume,
                        'declared_value':   wmcda_value,
                        'urgency_level':    wmcda_urgency,
                        'distance_km':      wmcda_distance,
                        'lcl_score':        wmcda_scores['lcl'],
                        'fcl_score':        wmcda_scores['fcl'],
                        'air_score':        wmcda_scores['air'],
                        'land_score':       wmcda_scores['land'],
                        'recommended_type': wmcda_recommended,
                        'computed_by':      request.user,
                    }
                )

                # â”€â”€ Historical recommendation â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
                if shipment.shipment_type:
                    past = (
                        ShippingAdvisory.objects
                        .filter(shipment__shipment_type=shipment.shipment_type,
                                recommended_type__isnull=False)
                        .exclude(shipment=shipment)
                        .values_list('recommended_type', flat=True)
                    )
                    if past:
                        from collections import Counter
                        counts   = Counter(past)
                        top_mode = counts.most_common(1)[0]
                        pct      = round(top_mode[1] / len(past) * 100)
                        wmcda_history = {
                            'total':       len(past),
                            'top_mode':    top_mode[0],
                            'top_pct':     pct,
                            'mode_label':  {'air': 'Air Freight', 'lcl': 'LCL', 'fcl': 'FCL', 'land': 'Land Freight'}.get(top_mode[0], top_mode[0].upper()),
                            'ship_type':   shipment.get_shipment_type_display(),
                        }
            except Exception as wmcda_err:
                print(f'WMCDA auto-compute error: {wmcda_err}')

            # Consignee notification is sent by notify_shipment_status_change above
            # when the status transitions to 'computed'. No duplicate needed here.

            messages.success(request, 'Computation & shipping analysis saved!')

        except ValueError:
            pass
        except Exception as e:
            messages.error(request, f'Computation error: {e}')
            items = result = None

    else:
        # â”€â”€ GET: pre-load saved data â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if existing:
            items = existing.get_items()

        if advisory_ex:
            wmcda_scores = {
                'lcl':  float(advisory_ex.lcl_score  or 0),
                'fcl':  float(advisory_ex.fcl_score  or 0),
                'air':  float(advisory_ex.air_score  or 0),
                'land': float(advisory_ex.land_score or 0),
            }
            wmcda_recommended = advisory_ex.recommended_type
            try:
                _, _, wmcda_breakdown, wmcda_explanation = compute_wmcda(
                    float(advisory_ex.gross_weight),
                    float(advisory_ex.cargo_volume),
                    float(advisory_ex.declared_value),
                    advisory_ex.urgency_level,
                    float(advisory_ex.distance_km),
                )
            except Exception:
                pass

        # OCR pre-fill â€” prefer DB-persisted ShipmentLineItem over session
        if not items and request.GET.get('ocr') == '1':
            db_items = ShipmentLineItem.objects.filter(
                shipment=shipment, source='ocr'
            ).select_related('hs_code').order_by('row_order')

            if db_items.exists():
                items = []
                for i, li in enumerate(db_items, 1):
                    hs_id   = li.hs_code_id or ''
                    hs_rate = float(li.hs_code.duty_rate) if li.hs_code else 0
                    items.append({
                        'no':           i,
                        'line_item_id': li.id,
                        'description':  li.description,
                        'exw':          float(li.total_val_usd) if li.total_val_usd else '',
                        'quantity':     float(li.quantity) if li.quantity else '1',
                        'unit':         li.unit or '',
                        'unit_price':   float(li.unit_price) if li.unit_price else '',
                        'hs_code_id':   hs_id,
                        'duty_rate':    hs_rate,
                        'dv_php':       None,
                        'cud':          None,
                        'item_freight': None,
                        'item_insurance': None,
                        'dv_usd':       None,
                        'gw': '', 'nw': '', 'pkgs': '',
                        'is_extracted': True,
                        'confidence':   float(li.confidence),
                    })

        if not items and request.GET.get('ocr') == '1':
            _ocr_sid   = request.session.get('ocr_shipment_id')
            _raw_items = request.session.get('ocr_items',  []) if _ocr_sid == shipment_id else []
            _ocr_flds  = request.session.get('ocr_fields', {}) if _ocr_sid == shipment_id else {}

            if _raw_items:
                # â”€â”€ Multi-item path: one row per extracted line item â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
                items = [
                    {
                        'no':             i,
                        'line_item_id':   '',
                        'description':    it.get('description', ''),
                        'exw':            it.get('total_value', '') or '',
                        'quantity':       it.get('quantity', '') or '1',
                        'unit':           it.get('unit', ''),
                        'unit_price':     it.get('unit_price', ''),
                        'hs_code_id':     it.get('hs_code_id', ''),
                        'duty_rate':      it.get('duty_rate', 0),
                        'dv_php':         None,
                        'cud':            None,
                        'item_freight':   None,
                        'item_insurance': None,
                        'other_charges':  None,
                        'dv_usd':         None,
                        'gw':             it.get('gross_weight', ''),
                        'nw':             it.get('net_weight', ''),
                        'pkgs':           it.get('num_packages', ''),
                        'is_extracted':   True,
                        'confidence':     it.get('confidence', 0.0),
                    }
                    for i, it in enumerate(_raw_items, 1)
                ]
            elif _ocr_flds:
                # â”€â”€ Single-total fallback: one row from merged OCR totals â”€â”€â”€â”€â”€
                def _val(k):
                    v = _ocr_flds.get(k, {})
                    return v.get('value', '') if isinstance(v, dict) else v
                items = [{
                    'no': 1, 'description': _val('description'),
                    'exw': _val('declared_value'),
                    'quantity': _val('total_quantity') or '1',
                    'unit': '', 'unit_price': '',
                    'hs_code_id': '', 'duty_rate': 0,
                    'dv_php': None, 'cud': None,
                    'item_freight': None, 'item_insurance': None,
                    'other_charges': None, 'dv_usd': None,
                    'is_extracted': True,
                    'confidence': _ocr_flds.get('description', {}).get('confidence', 0.0) if isinstance(_ocr_flds.get('description', {}), dict) else 0.0,
                }]

        # Historical on load
        if shipment.shipment_type:
            past = (
                ShippingAdvisory.objects
                .filter(shipment__shipment_type=shipment.shipment_type,
                        recommended_type__isnull=False)
                .exclude(shipment=shipment)
                .values_list('recommended_type', flat=True)
            )
            if past:
                from collections import Counter
                counts   = Counter(past)
                top_mode = counts.most_common(1)[0]
                pct      = round(top_mode[1] / len(past) * 100)
                wmcda_history = {
                    'total':       len(past),
                    'top_mode':    top_mode[0],
                    'top_pct':     pct,
                    'mode_label':  {'air': 'Air Freight', 'lcl': 'LCL', 'fcl': 'FCL', 'land': 'Land Freight'}.get(top_mode[0], top_mode[0].upper()),
                    'ship_type':   shipment.get_shipment_type_display(),
                }

    ocr_fields = request.session.get('ocr_fields', {}) if request.session.get('ocr_shipment_id') == shipment_id else {}
    ocr_items  = request.session.get('ocr_items',  []) if request.session.get('ocr_shipment_id') == shipment_id else []

    # â”€â”€ HS Code Suggestions (rule-based + historical) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Collect the richest available description text in priority order:
    # 1. shipment.description, 2. OCR item descriptions, 3. saved item descriptions
    hs_suggestions = []
    _suggest_parts = []
    if shipment.description:
        _suggest_parts.append(shipment.description)
    for _it in (ocr_items or [])[:3]:
        if _it.get('description'):
            _suggest_parts.append(_it['description'])
    if not _suggest_parts and existing:
        for _it in (existing.get_items() or [])[:3]:
            if _it.get('description'):
                _suggest_parts.append(_it['description'])
    _combined = ' '.join(_suggest_parts).strip()
    if _combined:
        hs_suggestions = suggest_hs_codes(_combined, top_n=5)
        # Persist as is_suggested records for tracking & historical learning
        for _hs in hs_suggestions:
            ShipmentHSCode.objects.get_or_create(
                shipment=shipment, hs_code=_hs,
                defaults={'is_suggested': True, 'is_confirmed': False}
            )

    # â”€â”€ Declared mode focused breakdown â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    declared_score     = None
    declared_breakdown = None
    declared_rating    = None
    if wmcda_scores and shipment.shipment_type:
        declared_score = wmcda_scores.get(shipment.shipment_type)
        if wmcda_breakdown:
            declared_breakdown = wmcda_breakdown.get(shipment.shipment_type)
        if declared_score is not None:
            if declared_score >= 0.80:
                declared_rating = 'Excellent'
            elif declared_score >= 0.65:
                declared_rating = 'Good'
            elif declared_score >= 0.50:
                declared_rating = 'Fair'
            else:
                declared_rating = 'Poor'

    # Pre-fill freight / insurance from consignee-provided values when no existing computation
    if existing:
        prefill_freight   = float(existing.total_freight   or 0)
        prefill_insurance = float(existing.total_insurance or 0)
    else:
        prefill_freight   = float(shipment.freight_cost   or 0)
        prefill_insurance = float(shipment.insurance_cost or 0)
    # ── Determine initial charge mode for template (drives section visibility) ──
    if existing:
        _ct = (existing.container_type or '').lower()
        if _ct in ('fcl', '20ft', '40ft'):
            computed_mode = 'fcl'
        elif _ct == 'air':
            computed_mode = 'air'
        elif _ct == 'land':
            computed_mode = 'land'
        else:
            computed_mode = 'lcl'
    else:
        _st = (shipment.shipment_type or 'lcl').lower()
        computed_mode = 'lcl' if _st in ('lcl', 'sea', '') else _st

    # ── Guide HS codes — set in session by save_ocr_items on the process page ──
    guide_hs_codes = []
    if str(request.session.get('guide_shipment_id', '')) == str(shipment_id):
        guide_hs_codes = request.session.get('guide_hs_codes', [])

    context = {
        'shipment':           shipment,
        'hs_codes':           hs_codes,
        'existing':           existing,
        'advisory_existing':  advisory_ex,
        'result':             result,
        'items':              items,
        'documents':          documents,
        'ocr_fields':         ocr_fields,
        'ocr_items':          ocr_items,
        'default_rate':       default_rate,
        'wmcda_scores':       wmcda_scores,
        'wmcda_recommended':  wmcda_recommended,
        'wmcda_breakdown':    wmcda_breakdown,
        'wmcda_explanation':  wmcda_explanation,
        'wmcda_history':      wmcda_history,
        'declared_score':     declared_score,
        'declared_breakdown': declared_breakdown,
        'declared_rating':    declared_rating,
        'hs_suggestions':     hs_suggestions,
        'guide_hs_codes':     guide_hs_codes,
        'computed_mode':      computed_mode,
        'prefill_freight':    prefill_freight,
        'prefill_insurance':  prefill_insurance,
    }
    context['confirmed_items'] = ShipmentLineItem.objects.filter(
        shipment=shipment, source='ocr'
    ).select_related('hs_code').order_by('row_order')
    return render(request, 'computation/compute.html', context)


# â”€â”€â”€ HS Code Suggestion Engine â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _can_download_computation(user, shipment):
    is_assigned_declarant = user.role == 'declarant' and shipment.declarant == user
    is_consignee = user.role == 'consignee' and shipment.consignee == user
    is_supervisor = user.role == 'supervisor'
    return is_assigned_declarant or is_consignee or is_supervisor


def _num(value):
    return float(value or 0)


def _resolve_report_items(computation):
    items = computation.get_items()
    hs_ids = [item.get('hs_code_id') for item in items if item.get('hs_code_id')]
    hs_map = {str(hs.id): hs for hs in HSCode.objects.filter(id__in=hs_ids)}
    report_items = []
    for item in items:
        hs = hs_map.get(str(item.get('hs_code_id'))) or computation.hs_code
        report_items.append({
            'description': item.get('description', ''),
            'quantity': item.get('quantity', ''),
            'unit': item.get('unit', ''),
            'hs_code': hs.code if hs else '',
            'duty_rate': _num(item.get('duty_rate', computation.duty_rate)),
            'dutiable_value': _num(item.get('dv_php')),
            'cud': _num(item.get('cud')),
            'unit_price': item.get('unit_price', ''),
        })
    return report_items


def _summary_rows(computation):
    csf_php = (computation.csf_usd or 0) * (computation.exchange_rate or 0)
    boc_total = (computation.customs_duty or 0) + (computation.vat_amount or 0) + Decimal('130') + (computation.ipf or 0)
    return [
        ('Dutiable Value', computation.dutiable_value or 0),
        ('Bank Charges', computation.bank_charges or 0),
        ('Customs Duties', computation.customs_duty or 0),
        ('Brokerage Fee', computation.brokerage_fee or 0),
        ('Arrastre', computation.arrastre or 0),
        ('Wharfage', computation.wharfage or 0),
        ('Container Service Fee', csf_php),
        ('Customs Documentary Stamp', Decimal('130')),
        ('Import Processing Fee', computation.ipf or 0),
        ('Total Landed Cost', computation.total_landed_cost or 0),
        ('VAT', computation.vat_amount or 0),
        ('BOC Payable', boc_total),
    ]


def _download_excel_report(shipment, computation):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_details = wb.active
    ws_details.title = 'Shipment Details'
    ws_items = wb.create_sheet('Line Items')
    ws_summary = wb.create_sheet('ECDT Summary')

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(bold=True, color='FFFFFF')
    title_font = Font(bold=True, size=14, color='1E3A5F')
    bold = Font(bold=True)
    border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )
    money_format = '#,##0.00'

    consignee = shipment.consignee.company_name or shipment.consignee.get_full_name() or shipment.consignee.username
    declarant = shipment.declarant.get_full_name() or shipment.declarant.username if shipment.declarant else ''
    prepared = computation.computed_by.get_full_name() or computation.computed_by.username if computation.computed_by else ''

    ws_details['A1'] = 'RTripleJ Customs Brokerage'
    ws_details['A1'].font = title_font
    ws_details['A2'] = 'ECDT Computation Sheet'
    detail_rows = [
        ('HAWB / BOL', shipment.hawb_number),
        ('Consignee', consignee),
        ('Declarant', declarant),
        ('Date', computation.computed_at.strftime('%Y-%m-%d') if computation.computed_at else ''),
        ('Shipment Mode', computation.container_type or shipment.shipment_type or ''),
        ('Import Type', shipment.get_import_type_display()),
        ('Exchange Rate', _num(computation.exchange_rate)),
        ('Prepared By', prepared),
    ]
    for row, (label, value) in enumerate(detail_rows, 4):
        ws_details.cell(row=row, column=1, value=label).font = bold
        ws_details.cell(row=row, column=2, value=value)
    ws_details.column_dimensions['A'].width = 22
    ws_details.column_dimensions['B'].width = 38

    item_headers = ['Description', 'Quantity', 'Unit', 'HS Code', 'Duty Rate', 'Dutiable Value', 'CUD per Item']
    for col, label in enumerate(item_headers, 1):
        cell = ws_items.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for row, item in enumerate(_resolve_report_items(computation), 2):
        values = [
            item['description'], item['quantity'], item['unit'], item['hs_code'],
            item['duty_rate'], item['dutiable_value'], item['cud'],
        ]
        for col, value in enumerate(values, 1):
            cell = ws_items.cell(row=row, column=col, value=value)
            cell.border = border
            if col in (6, 7):
                cell.number_format = money_format
            if col == 5:
                cell.number_format = '0.00"%"'
    for col, width in enumerate([42, 12, 12, 18, 12, 18, 18], 1):
        ws_items.column_dimensions[get_column_letter(col)].width = width

    ws_summary['A1'] = 'ECDT Summary'
    ws_summary['A1'].font = title_font
    ws_summary['A3'] = 'Charge'
    ws_summary['B3'] = 'Amount'
    for cell in ws_summary[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    for row, (label, value) in enumerate(_summary_rows(computation), 4):
        ws_summary.cell(row=row, column=1, value=label).border = border
        amount_cell = ws_summary.cell(row=row, column=2, value=_num(value))
        amount_cell.border = border
        amount_cell.number_format = money_format
        if label in {'Total Landed Cost', 'BOC Payable'}:
            ws_summary.cell(row=row, column=1).font = bold
            amount_cell.font = bold
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ECDT_{shipment.hawb_number}.xlsx'
    wb.save(response)
    return response


def _download_pdf_report(request, shipment, computation):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError:
        messages.error(request, 'PDF export requires reportlab. Install project requirements, then try again.')
        return redirect('declarant:process', shipment_id=shipment.id)

    buffer = tempfile.SpooledTemporaryFile()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.45 * inch, leftMargin=0.45 * inch)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('RTripleJ Customs Brokerage', styles['Title']),
        Paragraph('ECDT Computation Sheet', styles['Heading2']),
        Spacer(1, 10),
    ]

    consignee = shipment.consignee.company_name or shipment.consignee.get_full_name() or shipment.consignee.username
    declarant = shipment.declarant.get_full_name() or shipment.declarant.username if shipment.declarant else ''
    details = [
        ['HAWB / BOL', shipment.hawb_number, 'Consignee', consignee],
        ['Declarant', declarant, 'Date', computation.computed_at.strftime('%Y-%m-%d') if computation.computed_at else ''],
        ['Shipment Mode', computation.container_type or shipment.shipment_type or '', 'Exchange Rate', f'{_num(computation.exchange_rate):,.4f}'],
    ]
    detail_table = Table(details, colWidths=[1.1 * inch, 2.0 * inch, 1.1 * inch, 2.3 * inch])
    detail_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
        ('BACKGROUND', (2, 0), (2, -1), colors.whitesmoke),
    ]))
    story.extend([detail_table, Spacer(1, 12)])

    item_data = [['Description', 'Qty', 'Unit', 'HS Code', 'Duty %', 'D/V PHP', 'CUD']]
    for item in _resolve_report_items(computation):
        item_data.append([
            Paragraph(item['description'] or '', styles['BodyText']),
            item['quantity'], item['unit'], item['hs_code'],
            f"{item['duty_rate']:,.2f}", f"{item['dutiable_value']:,.2f}", f"{item['cud']:,.2f}",
        ])
    item_table = Table(item_data, colWidths=[2.25 * inch, 0.5 * inch, 0.5 * inch, 0.9 * inch, 0.6 * inch, 0.9 * inch, 0.9 * inch])
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
    ]))
    story.extend([Paragraph('Line Items', styles['Heading3']), item_table, Spacer(1, 12)])

    summary_data = [['Charge', 'Amount']]
    for label, value in _summary_rows(computation):
        summary_data.append([label, f"{_num(value):,.2f}"])
    summary_table = Table(summary_data, colWidths=[3.0 * inch, 1.5 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    story.extend([Paragraph('ECDT Summary', styles['Heading3']), summary_table])

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=ECDT_{shipment.hawb_number}.pdf'
    return response


@login_required
def download_computation(request, shipment_id):
    shipment = get_object_or_404(Shipment, id=shipment_id)
    if not _can_download_computation(request.user, shipment):
        messages.error(request, 'Access denied.')
        return redirect('accounts:login')

    computation = get_object_or_404(DutyComputation, shipment=shipment)
    if request.GET.get('format') == 'pdf':
        return _download_pdf_report(request, shipment, computation)
    return _download_excel_report(shipment, computation)


_HS_STOPWORDS = {
    'the','and','for','with','from','this','that','are','all','per',
    'each','pcs','set','unit','nos','lot','item','items','qty','piece',
    'pieces','new','used','other','various','type','types','model','grade',
    'size','kind','made','part','parts','product','products','goods',
}

def suggest_hs_codes(text, top_n=5):
    """
    Two-layer HS code recommendation engine.

    Layer 1 (Rule-based): DB-level OR prefilter on description keywords to
    narrow candidates from 9,268 rows to a small working set, then Python
    scoring with a minimum threshold of 2 matching keywords.

    Layer 2 (Historical): previously confirmed ShipmentHSCode assignments
    each contribute +0.5 to the score for that HS code.

    Returns up to top_n HSCode objects, ranked highest first.
    """
    if not text or not text.strip():
        return []

    keywords = [
        w for w in re.findall(r'[a-zA-Z]{3,}', text.lower())
        if w not in _HS_STOPWORDS
    ]
    # Deduplicate preserving order
    seen_kw = set()
    unique_keywords = []
    for kw in keywords:
        if kw not in seen_kw:
            seen_kw.add(kw)
            unique_keywords.append(kw)

    # Need at least 1 keyword. For single-word searches (e.g. “incubator”),
    # do a direct icontains match and return the best results.
    if not unique_keywords:
        return []

    # Layer 1 — DB-level OR prefilter (avoids loading all 9,268 rows per call)
    q = Q()
    for kw in unique_keywords[:12]:   # cap keyword count to keep query manageable
        q |= Q(description__icontains=kw)

    candidates = list(
        HSCode.objects.filter(q, is_active=True)
        .only('id', 'description', 'code', 'duty_rate', 'chapter')
    )
    if not candidates:
        return []

    # Score candidates in Python.
    # Note: AHTN descriptions are often very short ("Sunglasses", "Centrifuges"),
    # so requiring â‰¥2 hits would filter out many valid matches.
    # Minimum threshold = 1; higher scores naturally rank better matches first.
    scored = []
    for hs in candidates:
        hs_words = set(re.findall(r'[a-zA-Z]{3,}', hs.description.lower()))
        hits = sum(1 for kw in unique_keywords if kw in hs_words)
        if hits >= 1:
            scored.append([hs, float(hits)])

    if not scored:
        return []

    # Layer 2 â€” historical boost from confirmed past assignments
    hist = dict(
        ShipmentHSCode.objects
        .filter(is_confirmed=True)
        .values('hs_code_id')
        .annotate(n=Count('id'))
        .values_list('hs_code_id', 'n')
    )
    if hist:
        for entry in scored:
            entry[1] += hist.get(entry[0].id, 0) * 0.5

    scored.sort(key=lambda x: x[1], reverse=True)
    return [hs for hs, _ in scored[:top_n]]


# â”€â”€â”€ HS Code Suggest (AJAX) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _hs_payload(hs, source):
    return {
        'id': hs.id,
        'code': hs.code,
        'description': hs.description,
        'duty_rate': float(hs.duty_rate),
        'source': source,
    }


def _invoice_ocr_description(shipment):
    doc = shipment.documents.filter(document_type='invoice', ocr_ran_at__isnull=False).order_by('-ocr_ran_at').first()
    if not doc:
        return '', ''

    description_parts = []
    if doc.ocr_fields_json:
        try:
            fields = json.loads(doc.ocr_fields_json)
        except (TypeError, ValueError):
            fields = {}
        desc = fields.get('description')
        if isinstance(desc, dict) and desc.get('value'):
            description_parts.append(str(desc['value']))
        items = fields.get('__items__')
        if isinstance(items, list):
            for item in items:
                if item.get('description'):
                    description_parts.append(str(item['description']))
    return ' '.join(description_parts).strip(), doc.ocr_text or ''


@login_required
def hs_suggestions(request, shipment_id):
    shipment = get_object_or_404(Shipment, id=shipment_id)
    if request.user.role != 'declarant' or shipment.declarant != request.user:
        return JsonResponse({'error': 'Access denied.'}, status=403)

    q = request.GET.get('q', '').strip()
    if q:
        results = HSCode.objects.filter(
            Q(code__icontains=q) | Q(description__icontains=q),
            is_active=True,
        )[:10]
        return JsonResponse([_hs_payload(hs, 'suggested') for hs in results], safe=False)

    description, raw_text = _invoice_ocr_description(shipment)
    rows = []
    seen = set()

    direct_codes = re.findall(r'\b\d{4}(?:\.\d{2}){1,3}\b|\b\d{6,10}\b', raw_text or '')
    for code in direct_codes:
        hs = HSCode.objects.filter(Q(code=code) | Q(code__icontains=code), is_active=True).first()
        if hs and hs.id not in seen:
            rows.append(_hs_payload(hs, 'document'))
            seen.add(hs.id)
        if len(rows) >= 5:
            return JsonResponse(rows, safe=False)

    words = [
        word for word in re.findall(r'[A-Za-z]{3,}', description.lower())
        if word not in _HS_STOPWORDS
    ]
    scored = {}
    for word in words:
        for hs in HSCode.objects.filter(description__icontains=word, is_active=True)[:80]:
            scored.setdefault(hs.id, [hs, 0])
            scored[hs.id][1] += 1

    for hs, _score in sorted(scored.values(), key=lambda item: item[1], reverse=True):
        if hs.id in seen:
            continue
        rows.append(_hs_payload(hs, 'suggested'))
        seen.add(hs.id)
        if len(rows) >= 5:
            break

    return JsonResponse(rows, safe=False)


@login_required
def confirm_hs_code(request, shipment_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required.'}, status=405)

    shipment = get_object_or_404(Shipment, id=shipment_id)
    if request.user.role != 'declarant' or shipment.declarant != request.user:
        return JsonResponse({'ok': False, 'error': 'Access denied.'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except (TypeError, ValueError):
        payload = request.POST
    hs_code_value = str(payload.get('code') or '').strip()
    hs_code_id = payload.get('id')
    hs_qs = HSCode.objects.filter(is_active=True)
    hs = hs_qs.filter(id=hs_code_id).first() if hs_code_id else None
    if not hs and hs_code_value:
        hs = hs_qs.filter(code=hs_code_value).first()
    if not hs:
        return JsonResponse({'ok': False, 'error': 'HS code not found.'}, status=404)

    rel, _created = ShipmentHSCode.objects.get_or_create(
        shipment=shipment,
        hs_code=hs,
        defaults={'is_suggested': True, 'is_confirmed': True},
    )
    if not rel.is_confirmed:
        rel.is_confirmed = True
        rel.is_suggested = True
        rel.save(update_fields=['is_confirmed', 'is_suggested'])
    return JsonResponse({'ok': True})


@login_required
def hs_code_suggest(request):
    """
    AJAX endpoint for per-row live suggestions.
    GET ?q=<item description>&doc_hs=<HS from document>&context=<OCR text>&limit=<n>

    Priority order:
    1. doc_hs  — HS code explicitly printed in the invoice/packing list (highest confidence).
                 Looked up directly in the DB and returned as the first result.
    2. q alone — keyword search on item description.
    3. q + context — enriched search using full OCR raw text when q gives < 2 results.
    """
    try:
        q       = request.GET.get('q', '').strip()
        doc_hs  = request.GET.get('doc_hs', '').strip()
        context = request.GET.get('context', '').strip()[:3000]  # raw OCR text — expanded cap
        limit   = min(int(request.GET.get('limit', 5) or 5), 10)

        seen_ids = set()
        pinned   = []

        # ── Priority 1: HS code explicitly printed in the document ────────────
        if doc_hs:
            hs_obj = (
                HSCode.objects.filter(code=doc_hs, is_active=True).first()
                or HSCode.objects.filter(code__icontains=doc_hs, is_active=True).first()
                or HSCode.objects.filter(
                    code__startswith=doc_hs.replace('.', ''), is_active=True
                ).first()
            )
            if hs_obj:
                pinned.append({
                    'id':      hs_obj.id,
                    'code':    hs_obj.code,
                    'desc':    hs_obj.description[:80],
                    'rate':    float(hs_obj.duty_rate),
                    'chapter': hs_obj.chapter or '',
                    'source':  'document',
                })
                seen_ids.add(hs_obj.id)

        # ── Priority 2: OCR raw text as PRIMARY classification source ──────────
        # When raw OCR text is available it is far richer than a short extracted
        # item description.  Run the suggestion engine against the full OCR text
        # first; then re-rank results by how well they also match the typed
        # description (q), so description-specific terms bubble to the top.
        if context:
            # Combine: OCR text carries the product vocabulary; q refines it
            search_text = (context + ' ' + q).strip() if q else context
            kw_results = suggest_hs_codes(search_text, top_n=limit * 2)

            # Re-rank: items whose descriptions also match q get priority
            if q and kw_results:
                q_keywords = [
                    w for w in re.findall(r'[a-zA-Z]{3,}', q.lower())
                    if w not in _HS_STOPWORDS
                ]
                if q_keywords:
                    def _desc_hits(hs):
                        words = set(re.findall(r'[a-zA-Z]{3,}', hs.description.lower()))
                        return sum(1 for kw in q_keywords if kw in words)
                    kw_results = sorted(kw_results, key=_desc_hits, reverse=True)

            kw_results = kw_results[:limit]

            # Fallback: if OCR context produced nothing, try description alone
            if not kw_results and q:
                kw_results = suggest_hs_codes(q, top_n=limit)
        else:
            # No OCR context — use typed description only (original behaviour)
            kw_results = suggest_hs_codes(q, top_n=limit)

        remaining_slots = limit - len(pinned)
        extra = [
            {
                'id':      hs.id,
                'code':    hs.code,
                'desc':    hs.description[:80],
                'rate':    float(hs.duty_rate),
                'chapter': hs.chapter or '',
                'source':  'suggested',
            }
            for hs in kw_results
            if hs.id not in seen_ids
        ][:remaining_slots]

        return JsonResponse({'suggestions': pinned + extra})

    except Exception as e:
        print(f'[hs_code_suggest] Error: {e}')
        return JsonResponse({'suggestions': [], 'error': str(e)})


# â”€â”€â”€ Update ShipmentLineItem HS Code (AJAX PATCH) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@login_required
def update_line_item_hs(request, item_id):
    """
    PATCH /computation/line-item/<id>/hs/
    Body: { hs_code_id: <int> }
    Updates the hs_code FK on a ShipmentLineItem and returns the duty_rate.
    Only the assigned declarant may call this.
    """
    if request.method not in ('POST', 'PATCH'):
        return JsonResponse({'ok': False, 'error': 'POST/PATCH required.'}, status=405)

    item = get_object_or_404(ShipmentLineItem, id=item_id)
    shipment = item.shipment

    if request.user.role != 'declarant' or shipment.declarant != request.user:
        return JsonResponse({'ok': False, 'error': 'Access denied.'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except (TypeError, ValueError):
        payload = request.POST

    hs_code_id = payload.get('hs_code_id')
    if not hs_code_id:
        return JsonResponse({'ok': False, 'error': 'hs_code_id required.'}, status=400)

    try:
        hs = HSCode.objects.get(id=int(hs_code_id), is_active=True)
    except (HSCode.DoesNotExist, ValueError):
        return JsonResponse({'ok': False, 'error': 'HS code not found.'}, status=404)

    item.hs_code     = hs
    item.is_confirmed = True
    item.save(update_fields=['hs_code', 'is_confirmed', 'updated_at'])

    # Record the confirmation for historical boost
    ShipmentHSCode.objects.get_or_create(
        shipment=shipment, hs_code=hs,
        defaults={'is_suggested': True, 'is_confirmed': True},
    )

    return JsonResponse({
        'ok':       True,
        'hs_code':  hs.code,
        'duty_rate': float(hs.duty_rate),
    })


# â”€â”€â”€ HS Code Search â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@login_required
def hs_code_search(request):
    query   = request.GET.get('q', '')
    results = []
    if query:
        from django.db.models import Q
        results = HSCode.objects.filter(
            Q(code__icontains=query) | Q(description__icontains=query),
            is_active=True
        )[:10]
    return render(request, 'computation/hs_search.html', {
        'query': query, 'results': results,
    })


# â”€â”€â”€ Graduated WMCDA â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _lerp(x, x0, x1, y0, y1):
    if x <= x0: return y0
    if x >= x1: return y1
    return y0 + (y1 - y0) * (x - x0) / (x1 - x0)


def compute_wmcda(weight, volume, value, urgency, distance):
    # â”€â”€ Urgency factor â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # 0.0 = standard/normal, 0.5 = priority, 1.0 = urgent, 1.3 = rush
    _uf = {'standard': 0.0, 'normal': 0.0, 'priority': 0.5, 'urgent': 1.0, 'rush': 1.3}.get(urgency, 0.0)
    _is_time_critical = urgency in ('urgent', 'rush')
    _urgency_label    = {'standard': 'standard', 'normal': 'standard',
                         'priority': 'priority', 'urgent': 'urgent', 'rush': 'rush'}.get(urgency, urgency)

    # â”€â”€ Land freight viability flag â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Land freight is only practical for domestic PH routes or short ASEAN
    # cross-border routes. International sea/air routes (>1500 km) make land
    # freight impractical.
    _land_viable = distance <= 1500

    # â”€â”€ Cost scores â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # LCL: cost scales with CBM. Use volume if provided, else weight as proxy.
    if volume > 0:
        lcl_cost = max(0.20, _lerp(volume, 0, 15, 0.92, 0.28))   # ideal <5 CBM, poor >15
        fcl_cost = min(0.95, _lerp(volume, 0, 15, 0.22, 0.90))   # ideal >15 CBM (fills container)
    else:
        lcl_cost = max(0.25, _lerp(weight, 0, 1000, 0.88, 0.35))
        fcl_cost = _lerp(value, 0, 30000, 0.30, 0.88)

    air_cost  = max(0.15, _lerp(weight, 0, 500, 0.55, 0.18))     # expensive per-kg above 100 kg
    land_cost = max(0.20, _lerp(distance, 0, 1500, 0.90, 0.30)) if _land_viable else 0.15

    # â”€â”€ Time scores â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    _base_lcl_time  = max(0.30, _lerp(distance, 0, 2000, 0.72, 0.50))
    _base_fcl_time  = max(0.35, _lerp(distance, 0, 2000, 0.78, 0.55))
    _base_air_time  = 0.62
    _base_land_time = max(0.20, _lerp(distance, 0, 1500, 0.88, 0.28)) if _land_viable else 0.20

    lcl_time  = max(0.20, _base_lcl_time  - 0.37 * _uf)   # worse under urgency
    fcl_time  = max(0.25, _base_fcl_time  - 0.30 * _uf)   # worse under urgency
    air_time  = min(0.99, _base_air_time  + 0.34 * _uf)   # better under urgency
    land_time = max(0.15, _base_land_time - 0.20 * _uf) if _land_viable else 0.15

    # â”€â”€ Cargo suitability scores (weight + volume blended) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Physical weight component
    lcl_w  = _lerp(weight, 0, 2000, 0.92, 0.28)
    fcl_w  = _lerp(weight, 0, 2000, 0.18, 0.95)
    air_w  = max(0.10, _lerp(weight, 0, 300, 0.95, 0.15))
    land_w = _lerp(weight, 0, 2000, 0.70, 0.90)

    if volume > 0:
        # Volume (CBM) component â€” critical for LCL vs FCL decision
        # LCL sweet spot: <5 CBM. At 15 CBM (20ft container threshold) it's poor.
        # FCL sweet spot: >15 CBM. Below 5 CBM wastes the container.
        # Air: very harsh above 3 CBM (volumetric weight cost explodes).
        # Land: trucks are flexible; minimal volume penalty.
        lcl_v  = max(0.15, _lerp(volume, 0, 15, 0.95, 0.18))
        fcl_v  = min(0.95, _lerp(volume, 0, 15, 0.18, 0.95))
        air_v  = max(0.10, _lerp(volume, 0,  3, 0.95, 0.10))
        land_v = max(0.55, _lerp(volume, 0, 50, 0.90, 0.60))
        # Blend: 55% weight, 45% volume
        lcl_weight  = round(0.55 * lcl_w  + 0.45 * lcl_v,  3)
        fcl_weight  = round(0.55 * fcl_w  + 0.45 * fcl_v,  3)
        air_weight  = round(0.55 * air_w  + 0.45 * air_v,  3)
        land_weight = round(0.55 * land_w + 0.45 * land_v, 3) if _land_viable else 0.20
    else:
        # No volume data â€” weight only
        lcl_weight  = lcl_w
        fcl_weight  = fcl_w
        air_weight  = air_w
        land_weight = land_w if _land_viable else 0.20

    # â”€â”€ Risk scores â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Distance scores — normalised to [0,1] using 20 000 km as the practical maximum.
    # Each mode's score reflects how well-suited it is for the given routing distance:
    #   Air  — more justified at longer distances where speed vs. cost is critical.
    #   FCL  — optimal for long ocean routes; score rises with distance.
    #   LCL  — competitive at medium distances; handling overhead grows on very long routes.
    #   Land — only viable ≤1 500 km; score falls sharply beyond short-haul range.
    _D_MAX    = 20_000   # km — full trans-Pacific/Atlantic reference ceiling
    lcl_dist  = max(0.30, _lerp(distance, 0, _D_MAX, 0.80, 0.50))
    fcl_dist  = min(0.95, _lerp(distance, 0, _D_MAX, 0.55, 0.92))
    air_dist  = min(0.95, _lerp(distance, 0, _D_MAX, 0.60, 0.95))
    land_dist = max(0.10, _lerp(distance, 0, 1500,   0.92, 0.15)) if _land_viable else 0.10

    # â”€â”€ Criterion weights from SystemConfig â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    try:
        w_cost   = float(SystemConfig.get('wmcda_w_cost',     '35')) / 100
        w_time   = float(SystemConfig.get('wmcda_w_time',     '30')) / 100
        w_weight = float(SystemConfig.get('wmcda_w_weight',   '20')) / 100
        w_dist   = float(SystemConfig.get('wmcda_w_distance', '15')) / 100
    except Exception:
        w_cost, w_time, w_weight, w_dist = 0.35, 0.30, 0.20, 0.15

    def tws(c, t, wt, d):
        return round(c * w_cost + t * w_time + wt * w_weight + d * w_dist, 4)

    scores = {
        'lcl':  tws(lcl_cost,  lcl_time,  lcl_weight,  lcl_dist),
        'fcl':  tws(fcl_cost,  fcl_time,  fcl_weight,  fcl_dist),
        'air':  tws(air_cost,  air_time,  air_weight,  air_dist),
        'land': tws(land_cost, land_time, land_weight, land_dist),
    }
    recommended = max(scores, key=scores.get)

    breakdown = {
        'lcl':  {'cost': round(lcl_cost,  3), 'time': round(lcl_time,  3),
                 'weight': round(lcl_weight,  3), 'distance': round(lcl_dist,  3)},
        'fcl':  {'cost': round(fcl_cost,  3), 'time': round(fcl_time,  3),
                 'weight': round(fcl_weight,  3), 'distance': round(fcl_dist,  3)},
        'air':  {'cost': round(air_cost,  3), 'time': round(air_time,  3),
                 'weight': round(air_weight,  3), 'distance': round(air_dist,  3)},
        'land': {'cost': round(land_cost, 3), 'time': round(land_time, 3),
                 'weight': round(land_weight, 3), 'distance': round(land_dist, 3)},
    }

    weight_label = f'{weight:.0f} kg'
    value_label  = f'${value:,.0f}'
    dist_label   = f'{distance:.0f} km'
    vol_label    = f'{volume:.2f} CBM' if volume > 0 else ''

    cargo_desc = f'{weight_label}{", " + vol_label if vol_label else ""}'

    explanations = {
        'lcl': (
            f'LCL is cost-efficient for small-to-moderate cargo ({cargo_desc}). '
            f'{"Not recommended â€” slower sea transit conflicts with " + _urgency_label + " urgency." if _is_time_critical else "Suitable transit time for this urgency level."}'
        ),
        'fcl': (
            f'FCL is optimal for large or heavy cargo. '
            f'{"Volume of " + vol_label + " justifies a dedicated container. " if volume > 10 else ""}'
            f'{"Cargo of " + cargo_desc + " and value of " + value_label + " justify the container cost." if value > 10000 or weight > 500 else "May underutilize a full container for this cargo size."}'
            f'{" Sea transit may be too slow for " + _urgency_label + " urgency." if _is_time_critical else ""}'
        ),
        'air': (
            f'{"ðŸš¨ Rush â€” Air Freight only viable option for immediate delivery. " if urgency == "rush" else ""}'
            f'{"âš¡ Air Freight recommended â€” urgency requires fastest transit. " if urgency == "urgent" else ""}'
            f'{"â© Air Freight ideal for priority delivery at " + value_label + ". " if urgency == "priority" else ""}'
            f'{"Air Freight offers best security and speed for high-value goods at " + value_label + "." if value > 10000 and not _is_time_critical else ""}'
            f'{"Air Freight is competitive for this shipment profile." if not _is_time_critical and value <= 10000 else ""}'
        ),
        'land': (
            f'{"ðŸš› Land Freight is viable for this regional route (" + dist_label + ")." if distance <= 1000 else "Land Freight suited for this shorter route (" + dist_label + ")."} '
            f'Cargo of {cargo_desc} is well-suited for road transport. '
            f'{"Short-haul land routes can accommodate " + _urgency_label + " urgency." if _is_time_critical and distance <= 500 else "Suitable for this urgency level." if not _is_time_critical else "Road transit may be too slow for time-critical urgency on this route."}'
        ),
    }
    explanation = explanations.get(recommended, '')

    return scores, recommended, breakdown, explanation


# â”€â”€â”€ Shipping Advisory (auto-populated) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@login_required
def shipping_advisory(request, shipment_id):
    shipment = get_object_or_404(Shipment, id=shipment_id)

    # Only the assigned declarant may access the shipping advisory
    if request.user.role != 'declarant' or shipment.declarant != request.user:
        messages.error(request, 'Access denied.')
        return redirect('declarant:queue')

    existing = ShippingAdvisory.objects.filter(shipment=shipment).first()
    result = breakdown = explanation = None
    scores = None

    # â”€â”€ Auto-populate from shipment + computation data â”€â”€
    computation = getattr(shipment, 'computation', None)

    if existing:
        auto_weight   = float(existing.gross_weight)
        auto_volume   = float(existing.cargo_volume)
        auto_value    = float(existing.declared_value)
        auto_urgency  = existing.urgency_level
        auto_distance = float(existing.distance_km)

        # Re-derive breakdown and explanation from saved inputs so the criterion
        # table is visible on every page load, not just immediately after a POST.
        try:
            scores, result, breakdown, explanation = compute_wmcda(
                auto_weight, auto_volume, auto_value, auto_urgency, auto_distance
            )
        except Exception:
            pass
    else:
        # Pull weight from shipment model field
        auto_weight = float(shipment.gross_weight) if shipment.gross_weight else 0.0
        # Pull declared value from computation (USD) or shipment
        if computation and computation.declared_value:
            auto_value = float(computation.declared_value)
        elif shipment.declared_value:
            auto_value = float(shipment.declared_value)
        else:
            auto_value = 0.0
        auto_volume   = 0.0
        auto_urgency  = shipment.urgency
        auto_distance = 2600.0  # Default: Incheon, Korea â†’ Manila, Philippines

    # Determine which fields were auto-populated vs missing
    missing_fields = []
    if not auto_weight:
        missing_fields.append('Gross Weight (kg)')
    if not auto_value:
        missing_fields.append('Declared Value (USD)')

    auto_data = {
        'gross_weight':   auto_weight,
        'cargo_volume':   auto_volume,
        'declared_value': auto_value,
        'urgency_level':  auto_urgency,
        'distance_km':    auto_distance,
    }
    auto_sources = {
        'gross_weight':   'shipment' if (not existing and shipment.gross_weight) else ('advisory' if existing else 'manual'),
        'declared_value': 'computation' if (not existing and computation and computation.declared_value) else ('advisory' if existing else 'manual'),
        'urgency_level':  'shipment' if not existing else 'advisory',
        'distance_km':    'default' if not existing else 'advisory',
        'cargo_volume':   'advisory' if existing else 'manual',
    }

    if request.method == 'POST':
        try:
            weight   = float(request.POST.get('gross_weight', 0))
            volume   = float(request.POST.get('cargo_volume', 0))
            value    = float(request.POST.get('declared_value', 0))
            urgency  = request.POST.get('urgency_level', 'normal')
            distance = float(request.POST.get('distance_km', 2600))

            scores, recommended, breakdown, explanation = compute_wmcda(
                weight, volume, value, urgency, distance
            )

            ShippingAdvisory.objects.update_or_create(
                shipment=shipment,
                defaults={
                    'gross_weight':     weight,
                    'cargo_volume':     volume,
                    'declared_value':   value,
                    'urgency_level':    urgency,
                    'distance_km':      distance,
                    'lcl_score':        scores['lcl'],
                    'fcl_score':        scores['fcl'],
                    'air_score':        scores['air'],
                    'recommended_type': recommended,
                    'computed_by':      request.user,
                }
            )
            result = recommended
            messages.success(request, f'Recommendation: {recommended.upper()}')

            # Notify consignee of the advisory result
            try:
                from apps.notifications.utils import create_notification
                label_map = {'air': 'Air Freight', 'lcl': 'LCL', 'fcl': 'FCL', 'land': 'Land Freight'}
                create_notification(
                    recipient=shipment.consignee,
                    shipment=shipment,
                    notification_type='status_update',
                    title=f'Shipping Advisory Ready â€” {shipment.hawb_number}',
                    message=(
                        f'WMCDA Recommendation: {label_map.get(recommended, recommended.upper())}. '
                        f'{explanation[:120] if explanation else ""}'
                    ),
                )
            except Exception:
                pass

        except Exception as e:
            messages.error(request, f'Error: {e}')

    # â”€â”€ Historical advisory counts (same shipment type as this shipment) â”€â”€â”€â”€â”€â”€â”€
    wmcda_history = None
    if shipment.shipment_type:
        from collections import Counter
        past = list(
            ShippingAdvisory.objects
            .filter(
                shipment__shipment_type=shipment.shipment_type,
                recommended_type__isnull=False,
            )
            .exclude(shipment=shipment)
            .values_list('recommended_type', flat=True)
        )
        if past:
            counts   = Counter(past)
            top_mode = counts.most_common(1)[0]
            pct      = round(top_mode[1] / len(past) * 100)
            _label_map = {
                'air':  'Air Freight',
                'lcl':  'LCL',
                'fcl':  'FCL',
                'land': 'Land Freight',
            }
            wmcda_history = {
                'total':      len(past),
                'top_mode':   top_mode[0],
                'top_pct':    pct,
                'mode_label': _label_map.get(top_mode[0], top_mode[0].upper()),
                'ship_type':  shipment.get_shipment_type_display(),
                'counts':     {k: counts.get(k, 0) for k in ('air', 'lcl', 'fcl', 'land')},
            }

    context = {
        'shipment':       shipment,
        'existing':       existing,
        'result':         result,
        'scores':         scores,
        'breakdown':      breakdown,
        'explanation':    explanation,
        'auto_data':      auto_data,
        'auto_sources':   auto_sources,
        'missing_fields': missing_fields,
        'wmcda_history':  wmcda_history,
    }
    return render(request, 'computation/advisory.html', context)


# â”€â”€â”€ Save Declarant Advisory â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@login_required
def save_declarant_advisory(request, shipment_id):
    if request.method != 'POST':
        return redirect('computation:advisory', shipment_id=shipment_id)

    shipment = get_object_or_404(Shipment, id=shipment_id)

    if request.user.role != 'declarant' or shipment.declarant != request.user:
        messages.error(request, 'Access denied.')
        return redirect('declarant:queue')

    advisory = ShippingAdvisory.objects.filter(shipment=shipment).first()
    if not advisory:
        messages.error(request, 'Run the WMCDA computation first before saving an advisory.')
        return redirect('computation:advisory', shipment_id=shipment_id)

    recommendation = request.POST.get('declarant_recommendation', '').strip()
    note = request.POST.get('declarant_note', '').strip()

    valid_types = {'air', 'lcl', 'fcl', 'land', ''}
    if recommendation not in valid_types:
        messages.error(request, 'Invalid shipping type selected.')
        return redirect('computation:advisory', shipment_id=shipment_id)

    advisory.declarant_recommendation = recommendation or None
    advisory.declarant_note = note or None
    advisory.save(update_fields=['declarant_recommendation', 'declarant_note'])

    if recommendation:
        label_map = {'air': 'Air Freight', 'lcl': 'LCL', 'fcl': 'FCL', 'land': 'Land Freight'}
        mode_label = label_map.get(recommendation, recommendation.upper())
        try:
            from apps.notifications.utils import create_notification
            create_notification(
                recipient=shipment.consignee,
                shipment=shipment,
                notification_type='status_update',
                title=f'Declarant Advisory â€” {shipment.hawb_number}',
                message=(
                    f'Your declarant recommends {mode_label} for your shipment. '
                    f'{note}' if note else f'Your declarant recommends {mode_label} for your shipment.'
                ),
            )
        except Exception:
            pass
        messages.success(request, f'Advisory saved â€” {mode_label} recommended to consignee.')
    else:
        messages.success(request, 'Declarant advisory cleared.')

    return redirect('computation:advisory', shipment_id=shipment_id)
