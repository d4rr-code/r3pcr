import json
import logging
import os
import re
import threading
import uuid
from collections import defaultdict
from datetime import datetime, timedelta, date as date_type
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Avg, Sum, Min, Max, F, ExpressionWrapper, DurationField, Q
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from django.core.files.storage import default_storage
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.text import slugify
from django.core.mail import send_mail
from django.conf import settings
from django.core.paginator import Paginator
from django.http import HttpResponse

# ─── Business-day helpers ─────────────────────────────────────────────────────

URGENCY_BUSINESS_DAYS = {
    'rush': 3, 'urgent': 5, 'priority': 10, 'standard': 15, 'normal': 15,
}


def _urgency_business_days():
    values = dict(URGENCY_BUSINESS_DAYS)
    try:
        rows = {
            sc.key: sc.value
            for sc in SystemConfig.objects.filter(
                key__in=[f'urgency_days_{k}' for k in ('standard', 'priority', 'urgent', 'rush')]
            )
        }
    except Exception:
        rows = {}
    for key in ('standard', 'priority', 'urgent', 'rush'):
        try:
            days = int(rows.get(f'urgency_days_{key}', ''))
        except (TypeError, ValueError):
            continue
        if 1 <= days <= 60:
            values[key] = days
    values['normal'] = values['standard']
    return values


def _urgency_days_for(urgency):
    days = _urgency_business_days()
    return days.get(urgency or 'standard', days['standard'])


def _add_business_days(start_dt, n):
    """Return date that is n business days (Mon–Fri) after start_dt."""
    d = start_dt.date() if hasattr(start_dt, 'date') else start_dt
    added = 0
    while added < n:
        d += timedelta(days=1)
        if d.weekday() < 5:
            added += 1
    return d


def _business_days_diff(from_date, to_date):
    """Signed count of business days from from_date to to_date.
    Positive = future (days left), negative = past (overdue)."""
    from_date = from_date.date() if hasattr(from_date, 'date') else from_date
    to_date   = to_date.date()   if hasattr(to_date,   'date') else to_date
    if from_date == to_date:
        return 0
    sign = 1 if to_date > from_date else -1
    a, b = (from_date, to_date) if to_date > from_date else (to_date, from_date)
    count, d = 0, a
    while d < b:
        d += timedelta(days=1)
        if d.weekday() < 5:
            count += 1
    return sign * count


#  HS Code Section / Chapter Hierarchy
_HS_SECTIONS = [
    (1,  'I',     'Live Animals; Animal Products',                  list(range(1, 6))),
    (2,  'II',    'Vegetable Products',                             list(range(6, 15))),
    (3,  'III',   'Animal or Vegetable Fats and Oils',             [15]),
    (4,  'IV',    'Prepared Foodstuffs; Beverages; Tobacco',        list(range(16, 25))),
    (5,  'V',     'Mineral Products',                               list(range(25, 28))),
    (6,  'VI',    'Chemical or Allied Industry Products',            list(range(28, 39))),
    (7,  'VII',   'Plastics and Rubber',                            [39, 40]),
    (8,  'VIII',  'Raw Hides, Leather, Furskins',                  list(range(41, 44))),
    (9,  'IX',    'Wood, Cork, Straw',                             list(range(44, 47))),
    (10, 'X',     'Pulp of Wood, Paper, Paperboard',               list(range(47, 50))),
    (11, 'XI',    'Textiles and Textile Articles',                 list(range(50, 64))),
    (12, 'XII',   'Footwear, Headgear, Umbrellas',                 list(range(64, 68))),
    (13, 'XIII',  'Articles of Stone, Ceramics, Glass',            list(range(68, 71))),
    (14, 'XIV',   'Precious Stones, Precious Metals',              [71]),
    (15, 'XV',    'Base Metals and Articles',                      list(range(72, 84))),
    (16, 'XVI',   'Machinery and Mechanical Appliances',           [84, 85]),
    (17, 'XVII',  'Vehicles, Aircraft, Vessels',                   list(range(86, 90))),
    (18, 'XVIII', 'Optical, Photographic, Medical Instruments',    list(range(90, 93))),
    (19, 'XIX',   'Arms and Ammunition',                            [93]),
    (20, 'XX',    'Miscellaneous Manufactured Articles',            list(range(94, 97))),
    (21, 'XXI',   'Works of Art, Collectors Pieces, Antiques',     [97]),
]

def _chapter_num(chapter_str):
    """Extract numeric chapter from 'Chapter 84', '84', 'Chapter 01', '01', etc."""
    if not chapter_str:
        return None
    m = re.search(r'\d+', str(chapter_str))
    return int(m.group()) if m else None

logger = logging.getLogger(__name__)


def _send_mail_async(subject, message, from_email, recipient_list, html_message=None, log_tag=''):
    """Send email in a daemon thread; never blocks the HTTP response."""
    def _send():
        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=from_email,
                recipient_list=recipient_list,
                html_message=html_message,
            )
        except Exception as e:
            print(f'[EMAIL ERROR] {log_tag}: {e}')
    threading.Thread(target=_send, daemon=True).start()
from apps.accounts.models import User
from apps.accounts.views import _validate_phone_number
from apps.shipments.models import Shipment, HSCode, StatusLog, TariffSchedule, HSCodeRate
from apps.computation.models import DutyComputation, ShippingAdvisory
from apps.consignee.models import Feedback
from apps.notifications.utils import create_notification, notify_shipment_status_change
from apps.supervisor.exchange_rates import ensure_daily_exchange_rates
from .models import SystemConfig, Announcement, IssueReport


def supervisor_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != 'supervisor':
            return redirect('accounts:login')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


#  Dashboard 

@login_required
@supervisor_required
def dashboard(request):
    """Unified analytics/command-centre page."""
    return _analytics_context_response(request)


def _analytics_report_data(request):
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    declarant_filter = request.GET.get('declarant', '').strip()

    qs = Shipment.objects.select_related('consignee', 'declarant').all()
    if date_from:
        qs = qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(submitted_at__date__lte=date_to)
    if declarant_filter:
        qs = qs.filter(declarant__username=declarant_filter)

    total = qs.count()
    status_rows = []
    status_counts = {
        row['status']: row['count']
        for row in qs.values('status').annotate(count=Count('id'))
    }
    for key, label in Shipment.STATUS_CHOICES:
        count = status_counts.get(key, 0)
        status_rows.append([label, count, f'{round(count / total * 100, 1) if total else 0}%'])

    type_rows = []
    type_counts = {
        row['shipment_type']: row['count']
        for row in qs.values('shipment_type').annotate(count=Count('id'))
    }
    for key, label in Shipment.SHIPMENT_TYPE_CHOICES:
        count = type_counts.get(key, 0)
        type_rows.append([label, count, f'{round(count / total * 100, 1) if total else 0}%'])

    urgency_map = defaultdict(int)
    for row in qs.values('urgency').annotate(count=Count('id')):
        key = 'standard' if row['urgency'] in ('normal', 'standard', None) else row['urgency']
        urgency_map[key] += row['count']
    urgency_labels = dict(Shipment.URGENCY_CHOICES)
    urgency_rows = []
    for key in ('standard', 'priority', 'urgent', 'rush'):
        count = urgency_map.get(key, 0)
        urgency_rows.append([urgency_labels.get(key, key.title()), count, f'{round(count / total * 100, 1) if total else 0}%'])

    ids = qs.values_list('id', flat=True)
    advisory_qs = ShippingAdvisory.objects.filter(shipment_id__in=ids)
    wmcda_total = advisory_qs.filter(recommended_type__isnull=False).count()
    wmcda_labels = {'air': 'Air Freight', 'lcl': 'LCL Sea', 'fcl': 'FCL Sea'}
    wmcda_counts = {
        row['recommended_type']: row['count']
        for row in advisory_qs.values('recommended_type').annotate(count=Count('id'))
        if row['recommended_type']
    }
    wmcda_avg = advisory_qs.aggregate(
        avg_air=Avg('air_score'), avg_lcl=Avg('lcl_score'),
        avg_fcl=Avg('fcl_score'),
    )
    wmcda_rows = []
    for key, label in wmcda_labels.items():
        count = wmcda_counts.get(key, 0)
        wmcda_rows.append([
            label,
            count,
            f'{round(count / wmcda_total * 100, 1) if wmcda_total else 0}%',
            f'{round(float(wmcda_avg.get(f"avg_{key}") or 0) * 100, 1)}%',
        ])

    currency_total = qs.exclude(invoice_currency='').count()
    currency_rows = []
    for row in qs.exclude(invoice_currency='').values('invoice_currency').annotate(count=Count('id')).order_by('-count'):
        count = row['count']
        currency_rows.append([
            row['invoice_currency'] or 'USD',
            count,
            f'{round(count / currency_total * 100, 1) if currency_total else 0}%',
        ])

    cost_rows = []
    cost_qs = DutyComputation.objects.filter(shipment_id__in=ids, total_landed_cost__isnull=False)
    for key, label in Shipment.SHIPMENT_TYPE_CHOICES:
        agg = cost_qs.filter(shipment__shipment_type=key).aggregate(
            count=Count('id'), avg=Avg('total_landed_cost'), total=Sum('total_landed_cost'),
            min_val=Min('total_landed_cost'), max_val=Max('total_landed_cost')
        )
        cost_rows.append([
            label,
            agg['count'],
            round(float(agg['avg'] or 0), 2),
            round(float(agg['total'] or 0), 2),
            round(float(agg['min_val'] or 0), 2),
            round(float(agg['max_val'] or 0), 2),
        ])

    declarant_rows = []
    for dec in User.objects.filter(role='declarant').order_by('first_name', 'username'):
        dec_qs = qs.filter(declarant=dec)
        assigned = dec_qs.count()
        computed = dec_qs.filter(status__in=['computed', 'approved', 'lodgement', 'ongoing', 'assessed', 'paid', 'released', 'billed']).count()
        completed = dec_qs.filter(status='billed').count()
        revision_flags = dec_qs.filter(status__in=['for_revision', 'rejected']).count()
        if assigned or computed or completed or revision_flags:
            declarant_rows.append([
                dec.get_full_name() or dec.username,
                assigned,
                computed,
                completed,
                revision_flags,
                f'{round(completed / assigned * 100, 1) if assigned else 0}%',
            ])

    feedback_total = Feedback.objects.count()
    feedback_avg = Feedback.objects.aggregate(avg=Avg('rating'))['avg']
    feedback_positive = Feedback.objects.filter(rating__gte=4).count()

    recent_rows = []
    for s in qs.order_by('-submitted_at')[:100]:
        recent_rows.append([
            s.hawb_number,
            s.consignee.get_full_name() or s.consignee.username,
            s.get_shipment_type_display() or '',
            s.get_status_display(),
            s.get_urgency_display(),
            s.submitted_at.strftime('%Y-%m-%d'),
        ])

    return {
        'generated_at': timezone.localtime().strftime('%Y-%m-%d %H:%M'),
        'filters': {
            'date_from': date_from or 'All',
            'date_to': date_to or 'All',
            'declarant': declarant_filter or 'All',
        },
        'summary': [
            ['Total Shipments', total],
            ['Active Users', User.objects.filter(role__in=['consignee', 'declarant'], is_active=True, is_pending_approval=False).count()],
            ['Consignees', User.objects.filter(role='consignee', is_active=True, is_pending_approval=False).count()],
            ['Declarants', User.objects.filter(role='declarant', is_active=True, is_pending_approval=False).count()],
            ['WMCDA Advisories', wmcda_total],
            ['Feedback Count', feedback_total],
            ['Average Feedback Rating', round(float(feedback_avg or 0), 1)],
            ['Positive Feedback %', f'{round(feedback_positive / feedback_total * 100, 1) if feedback_total else 0}%'],
        ],
        'tables': [
            ('Status Pipeline', ['Status', 'Count', 'Share'], status_rows),
            ('Shipment Types', ['Type', 'Count', 'Share'], type_rows),
            ('Urgency Distribution', ['Urgency', 'Count', 'Share'], urgency_rows),
            ('WMCDA Recommendations', ['Mode', 'Recommended Count', 'Share', 'Average Score'], wmcda_rows),
            ('Currency Usage', ['Currency', 'Count', 'Share'], currency_rows),
            ('Landed Cost By Mode', ['Mode', 'Computations', 'Average PHP', 'Total PHP', 'Min PHP', 'Max PHP'], cost_rows),
            ('Declarant Performance', ['Declarant', 'Assigned', 'Computed+', 'Billed', 'Revision/Rejected', 'Completion %'], declarant_rows),
            ('Recent Shipments', ['HAWB', 'Consignee', 'Mode', 'Status', 'Urgency', 'Submitted'], recent_rows),
        ],
    }


@login_required
@supervisor_required
def analytics_export(request):
    fmt = (request.GET.get('format') or 'xlsx').lower()
    data = _analytics_report_data(request)
    filename_date = timezone.localtime().strftime('%Y%m%d')

    if fmt == 'pdf':
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [
            Paragraph('R3-PCR Analytics Report', styles['Title']),
            Paragraph(f"Generated: {data['generated_at']}", styles['Normal']),
            Paragraph(
                f"Filters: From {data['filters']['date_from']} to {data['filters']['date_to']} | Declarant: {data['filters']['declarant']}",
                styles['Normal'],
            ),
            Spacer(1, 12),
        ]
        for title, headers, rows in [('Executive Summary', ['Metric', 'Value'], data['summary'])] + data['tables']:
            story.append(Paragraph(title, styles['Heading2']))
            table_data = [headers] + (rows or [['No data', ''] + [''] * (len(headers) - 2)])
            tbl = Table(table_data, repeatRows=1)
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3358')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#DCE5EF')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.extend([tbl, Spacer(1, 12)])
        doc.build(story)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="R3PCR_Analytics_Report_{filename_date}.pdf"'
        return response

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    header_fill = PatternFill('solid', fgColor='1B3358')
    header_font = Font(color='FFFFFF', bold=True)

    ws.append(['R3-PCR Analytics Report'])
    ws['A1'].font = Font(bold=True, size=16)
    ws.append(['Generated', data['generated_at']])
    ws.append(['Date From', data['filters']['date_from'], 'Date To', data['filters']['date_to'], 'Declarant', data['filters']['declarant']])
    ws.append([])
    ws.append(['Metric', 'Value'])
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
    for row in data['summary']:
        ws.append(row)

    for title, headers, rows in data['tables']:
        sheet = wb.create_sheet(title[:31])
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            sheet.append(row)
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="R3PCR_Analytics_Report_{filename_date}.xlsx"'
    return response


#  User Management 

@login_required
@supervisor_required
def user_management(request):
    users   = User.objects.filter(is_pending_approval=False).order_by('role', 'username')
    pending = User.objects.filter(is_pending_approval=True).order_by('date_joined')
    user_stats = {
        'total': users.count(),
        'consignees': users.filter(role='consignee').count(),
        'declarants': users.filter(role='declarant').count(),
        'active': users.filter(is_active=True).count(),
        'inactive': users.filter(is_active=False).count(),
    }
    return render(request, 'supervisor/users.html', {
        'users':   users,
        'pending': pending,
        'user_stats': user_stats,
    })


@login_required
@supervisor_required
def approve_registration(request, user_id):
    user = get_object_or_404(User, id=user_id, is_pending_approval=True)
    if request.method == 'POST':
        user.is_active           = True
        user.is_pending_approval = False
        user.save()

        if user.email:
            _send_mail_async(
                subject='R3-PCR - Account Approved',
                message=(
                    f'Hello {user.first_name or user.username},\n\n'
                    f'Your R3-PCR account has been approved. '
                    f'You can now log in.\n\nUsername: {user.username}'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                html_message=f'''
                    <div style="font-family:Arial,sans-serif;max-width:480px;margin:0 auto;">
                        <h2 style="color:#22c55e;">Account Approved!</h2>
                        <p>Hello <strong>{user.first_name or user.username}</strong>,</p>
                        <p>Your R3-PCR account has been <strong style="color:#22c55e;">approved</strong>.
                           You can now log in.</p>
                        <p><strong>Username:</strong> {user.username}</p>
                        <p style="color:#94a3b8;font-size:12px;margin-top:20px;">
                            R3-PCR Pre-Clearance Decision Support System
                        </p>
                    </div>
                ''',
                log_tag=f'approval email to {user.username}',
            )

        messages.success(request, f'Account for {user.username} approved and activated.')
    return redirect('supervisor:users')


@login_required
@supervisor_required
def reject_registration(request, user_id):
    user = get_object_or_404(User, id=user_id, is_pending_approval=True)
    if request.method == 'POST':
        username = user.username
        email    = user.email
        name     = user.first_name or username
        if email:
            _send_mail_async(
                subject='R3-PCR - Registration Not Approved',
                message=(
                    f'Hello {name},\n\nUnfortunately your R3-PCR registration was not approved. '
                    f'Please contact the administrator for more information.'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                log_tag=f'rejection email to {username}',
            )
        user.delete()
        messages.warning(request, f'Registration for {username} rejected and removed.')
    return redirect('supervisor:users')


@login_required
@supervisor_required
def add_user(request):
    form_data = {}
    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        email      = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        role       = request.POST.get('role', '').strip()
        phone      = request.POST.get('phone_number', '').strip()
        company    = request.POST.get('company_name', '').strip()
        password   = request.POST.get('password', '')
        confirm    = request.POST.get('confirm_password', '')
        form_data  = request.POST

        errors = []
        if not all([first_name, last_name, username, email, role, password, confirm]):
            errors.append('Please complete all required fields.')
        if role not in dict(User.ROLE_CHOICES):
            errors.append('Please select a valid role.')
        if User.objects.filter(username=username).exists():
            errors.append('Username already taken.')
        if User.objects.filter(email=email).exists():
            errors.append('Email already registered.')
        if password != confirm:
            errors.append('Passwords do not match.')
        if len(password) < 8:
            errors.append('Password must be at least 8 characters.')
        phone_error = _validate_phone_number(phone)
        if phone_error:
            errors.append(phone_error)
        if not re.search(r'[a-z]', password):
            errors.append('Password must include at least one lowercase letter.')
        if not re.search(r'[A-Z]', password):
            errors.append('Password must include at least one uppercase letter.')
        if not re.search(r'\d', password):
            errors.append('Password must include at least one number.')
        if not re.search(r'[^A-Za-z0-9]', password):
            errors.append('Password must include at least one special character.')

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            User.objects.create_user(
                username=username, email=email,
                first_name=first_name, last_name=last_name,
                role=role, phone_number=phone, company_name=company,
                password=password,
            )
            messages.success(request, f'User {username} created.')
            return redirect('supervisor:users')

    return render(request, 'supervisor/add_user.html', {
        'mode': 'add',
        'form_data': form_data,
        'role_choices': User.ROLE_CHOICES,
    })


@login_required
@supervisor_required
def edit_user(request, user_id):
    edited_user = get_object_or_404(User, id=user_id)
    form_data = None

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        email      = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        phone      = request.POST.get('phone_number', '').strip()
        company    = request.POST.get('company_name', '').strip()
        form_data  = request.POST

        errors = []
        if not all([first_name, last_name, username, email]):
            errors.append('Please complete all required fields.')
        if User.objects.filter(username=username).exclude(pk=edited_user.pk).exists():
            errors.append('Username already taken.')
        if User.objects.filter(email=email).exclude(pk=edited_user.pk).exists():
            errors.append('Email already registered.')
        phone_error = _validate_phone_number(phone)
        if phone_error:
            errors.append(phone_error)

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            edited_user.first_name = first_name
            edited_user.last_name = last_name
            edited_user.email = email
            edited_user.phone_number = phone
            edited_user.company_name = company
            edited_user.username = username
            edited_user.save()
            messages.success(request, f'User {edited_user.username} updated.')
            return redirect('supervisor:users')

    return render(request, 'supervisor/add_user.html', {
        'mode': 'edit',
        'edited_user': edited_user,
        'form_data': form_data,
        'role_choices': User.ROLE_CHOICES,
    })


@login_required
@supervisor_required
def toggle_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if user == request.user:
        messages.error(request, 'You cannot deactivate yourself.')
    else:
        user.is_active = not user.is_active
        user.save()
        state = 'activated' if user.is_active else 'deactivated'
        messages.success(request, f'User {user.username} {state}.')
    return redirect('supervisor:users')


#  Analytics (merged command centre) 

@login_required
@supervisor_required
def analytics(request):
    return redirect('supervisor:dashboard')


def _analytics_context_response(request):
    all_shipments = Shipment.objects.all()

    #  Chart/KPI filters (date range + declarant)
    date_from        = request.GET.get('date_from', '').strip()
    date_to          = request.GET.get('date_to', '').strip()
    declarant_filter = request.GET.get('declarant', '').strip()
    overview_range  = request.GET.get('overview_range', 'year').strip().lower()
    if overview_range not in {'year', '6m'}:
        overview_range = 'year'

    chart_qs = all_shipments
    if date_from:
        chart_qs = chart_qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        chart_qs = chart_qs.filter(submitted_at__date__lte=date_to)
    if declarant_filter:
        chart_qs = chart_qs.filter(declarant__username=declarant_filter)

    chart_total = chart_qs.count()

    #  Shipment table filters (search + status + date) 
    q        = request.GET.get('q', '').strip()
    status_f = request.GET.get('status_f', '').strip()

    table_qs = all_shipments.order_by('-submitted_at')
    if q:
        table_qs = (
            all_shipments.filter(hawb_number__icontains=q)
            | all_shipments.filter(consignee__first_name__icontains=q)
            | all_shipments.filter(consignee__last_name__icontains=q)
            | all_shipments.filter(consignee__username__icontains=q)
        ).order_by('-submitted_at')
    if status_f:
        table_qs = table_qs.filter(status=status_f)
    if date_from:
        table_qs = table_qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        table_qs = table_qs.filter(submitted_at__date__lte=date_to)

    #  KPI strip (always all-time) 
    total_all = all_shipments.count()
    total_computed_presented = (
        StatusLog.objects.filter(new_status='computed')
        .values('shipment_id').distinct().count()
    )
    total_consignee_approved = (
        StatusLog.objects.filter(new_status='approved')
        .values('shipment_id').distinct().count()
    )
    consignee_approval_rate = (
        round(total_consignee_approved / total_computed_presented * 100, 1)
        if total_computed_presented else 0
    )

    #  Status breakdown bar chart (respects chart filters) — single grouped query
    _status_colors = {
        'incoming':    '#f59e0b', 'arrived':    '#3b82f6', 'computed':    '#8b5cf6',
        'approved':    '#22c55e', 'rejected':   '#ef4444', 'for_revision':'#f97316',
        'lodgement':   '#38bdf8', 'ongoing':    '#64748b', 'assessed':    '#14b8a6',
        'paid':        '#84cc16', 'released':   '#22d3ee', 'billed':      '#a855f7',
    }
    # Materialise chart_qs IDs once — reused for status, WMCDA and declarant sections
    _chart_ids_qs = chart_qs.values_list('id', flat=True)
    _status_counts_raw = {
        r['status']: r['count']
        for r in (
            Shipment.objects.filter(id__in=_chart_ids_qs)
            .values('status')
            .annotate(count=Count('id'))
        )
    }
    status_rows = []
    for key, label in Shipment.STATUS_CHOICES:
        count = _status_counts_raw.get(key, 0)
        status_rows.append({
            'key': key, 'label': label, 'count': count,
            'pct': round(count / chart_total * 100, 1) if chart_total else 0,
            'color': _status_colors.get(key, '#475569'),
        })
    # Dashboard display order: 4 rows x 3 columns, matching the supervisor wireframe.
    _pipeline_order = [
        'incoming', 'approved', 'assessed',
        'arrived', 'for_revision', 'paid',
        'rejected', 'lodgement', 'released',
        'computed', 'ongoing', 'billed',
    ]
    _status_map = {r['key']: r for r in status_rows}
    pipeline_rows = [_status_map[k] for k in _pipeline_order if k in _status_map]
    # bar_pct relative to max for stacked bar tooltip (not used for width  CSS does that)
    max_status = max((r['count'] for r in pipeline_rows), default=1) or 1
    for row in pipeline_rows:
        row['bar_pct'] = round(row['count'] / max_status * 100) if max_status > 0 else 0
    # Keep sorted version for any legacy references
    status_rows_sorted = sorted(pipeline_rows, key=lambda r: r['count'], reverse=True)

    # Add subtitle to each pipeline row for the Status Overview cards.
    _status_meta = {
        'incoming':     {'subtitle': 'Awaits Declarant Assignment'},
        'arrived':      {'subtitle': 'Awaits ECDT Processing'},
        'computed':     {'subtitle': 'Awaits Consignee Approval'},
        'for_revision': {'subtitle': 'Returned for Revision'},
        'rejected':     {'subtitle': 'Docs Not Complete Update From Declarant'},
        'approved':     {'subtitle': 'Proceeding to Lodgement'},
        'lodgement':    {'subtitle': 'Filed with BOC'},
        'ongoing':      {'subtitle': 'Lined Up for Final Assessment'},
        'assessed':     {'subtitle': 'Awaits Payment of D/T'},
        'paid':         {'subtitle': 'Awaits CNTR Discharge & Delivery'},
        'released':     {'subtitle': 'Awaits Final Billing'},
        'billed':       {'subtitle': 'Shipment Fully Processed End-to-End'},
    }
    for row in pipeline_rows:
        meta = _status_meta.get(row['key'], {})
        row['subtitle'] = meta.get('subtitle', '')
        _display_labels = {
            'for_revision': 'Revision',
            'rejected': 'Flags',
        }
        row['display_label'] = _display_labels.get(row['key'], row['label'])

    # WMCDA Scoreboard (respects chart filters) — use materialised IDs
    _wmcda_meta = [
        ('air',  'Air Freight',  '#f59e0b', 'AIR'),
        ('lcl',  'LCL Sea',      '#38bdf8', 'LCL'),
        ('fcl',  'FCL Sea',      '#8b5cf6', 'FCL'),
    ]
    advisory_qs = ShippingAdvisory.objects.filter(shipment_id__in=_chart_ids_qs)
    wmcda_total = advisory_qs.filter(recommended_type__isnull=False).count()
    # Batch count by recommended_type in one query
    _wmcda_type_counts = {
        r['recommended_type']: r['cnt']
        for r in advisory_qs.values('recommended_type').annotate(cnt=Count('id'))
        if r['recommended_type']
    }
    # Batch avg scores in one query
    _wmcda_avg_agg = advisory_qs.aggregate(
        avg_air=Avg('air_score'), avg_lcl=Avg('lcl_score'),
        avg_fcl=Avg('fcl_score'),
    )
    wmcda_scoreboard = []
    for key, label, color, icon in _wmcda_meta:
        count     = _wmcda_type_counts.get(key, 0)
        pct       = round(count / wmcda_total * 100, 1) if wmcda_total else 0
        avg_score = round(float(_wmcda_avg_agg.get(f'avg_{key}') or 0) * 100, 1)
        wmcda_scoreboard.append({
            'key': key, 'label': label, 'color': color, 'icon': icon,
            'count': count, 'pct': pct, 'avg_score': avg_score,
        })
    wmcda_scoreboard.sort(key=lambda x: x['count'], reverse=True)
    rank_labels = ['1st', '2nd', '3rd']
    for i, row in enumerate(wmcda_scoreboard):
        row['rank'] = rank_labels[i] if i < len(rank_labels) else f'{i+1}th'
    wmcda_max = wmcda_scoreboard[0]['count'] if wmcda_scoreboard else 1
    #  Declarant Performance (respects chart filters) — batch queries to avoid N+1
    declarants = User.objects.filter(role='declarant').order_by('first_name', 'username')

    # Single bulk load of all relevant StatusLog rows for declarant performance
    # Re-uses _chart_ids_qs (lazy queryset) already defined above
    _perf_logs = (
        StatusLog.objects
        .filter(
            shipment_id__in=_chart_ids_qs,
            new_status__in=['computed', 'arrived', 'approved', 'for_revision', 'rejected'],
        )
        .values('shipment_id', 'new_status', 'changed_at')
        .order_by('shipment_id', 'changed_at')
    )

    # Build lookup tables from the single query
    # shipment_id -> declarant_id  (from chart_qs)
    _ship_declarant = dict(
        Shipment.objects.filter(id__in=_chart_ids_qs)
                        .values_list('id', 'declarant_id')
    )
    # Group logs by declarant
    _dec_logs = defaultdict(lambda: defaultdict(list))  # dec_id -> status -> [log_dicts]
    for log in _perf_logs:
        dec_id = _ship_declarant.get(log['shipment_id'])
        if dec_id:
            _dec_logs[dec_id][log['new_status']].append(log)

    declarant_data = []
    for dec in declarants:
        logs_by_status = _dec_logs.get(dec.id, {})

        # First computed log per shipment
        computed_map = {}
        for log in logs_by_status.get('computed', []):
            sid = log['shipment_id']
            if sid not in computed_map or log['changed_at'] < computed_map[sid]['changed_at']:
                computed_map[sid] = log

        # Most recent arrived log per shipment (for speed calculation)
        arrived_map = {}
        for log in logs_by_status.get('arrived', []):
            sid = log['shipment_id']
            if sid not in arrived_map or log['changed_at'] > arrived_map[sid]['changed_at']:
                arrived_map[sid] = log

        durations = []
        for sid, c_log in computed_map.items():
            a_log = arrived_map.get(sid)
            if a_log and a_log['changed_at'] <= c_log['changed_at']:
                durations.append(c_log['changed_at'] - a_log['changed_at'])

        total_comp       = len(computed_map)
        ecdt_approved    = len(logs_by_status.get('approved', []))
        revised_rejected = (
            len(logs_by_status.get('for_revision', []))
            + len(logs_by_status.get('rejected', []))
        )
        avg_hours = None
        if durations:
            avg_hours = round(
                sum(d.total_seconds() for d in durations) / len(durations) / 3600, 1
            )
        declarant_data.append({
            'name':             dec.get_full_name() or dec.username,
            'username':         dec.username,
            'total_processed':  total_comp,
            'avg_hours':        avg_hours,
            'ecdt_approved':    ecdt_approved,
            'revised_rejected': revised_rejected,
            'approval_rate':    round(ecdt_approved / total_comp * 100, 1) if total_comp else 0,
        })

    # ── Redesigned dashboard: new context variables ──────────────────────

    # Shipment type KPI counts (all-time)
    shipment_type_counts = {
        'air':  all_shipments.filter(shipment_type='air').count(),
        'lcl':  all_shipments.filter(shipment_type='lcl').count(),
        'fcl':  all_shipments.filter(shipment_type='fcl').count(),
    }

    # Urgency distribution — normalise 'normal' alias → 'standard' (all-time)
    _urgency_raw = chart_qs.values('urgency').annotate(count=Count('id'))
    _urgency_map = {}
    for _r in _urgency_raw:
        _key = 'standard' if _r['urgency'] in ('normal', 'standard', None) else _r['urgency']
        _urgency_map[_key] = _urgency_map.get(_key, 0) + _r['count']
    urgency_counts = [
        {'key': 'standard', 'label': 'Standard', 'color': '#3b82f6', 'count': _urgency_map.get('standard', 0)},
        {'key': 'priority', 'label': 'Priority', 'color': '#f59e0b', 'count': _urgency_map.get('priority', 0)},
        {'key': 'urgent',   'label': 'Urgent',   'color': '#f97316', 'count': _urgency_map.get('urgent', 0)},
        {'key': 'rush',     'label': 'Rush',     'color': '#ef4444', 'count': _urgency_map.get('rush', 0)},
    ]
    urgency_total = sum(u['count'] for u in urgency_counts)
    urgency_chart_labels = json.dumps([u['label'] for u in urgency_counts])
    urgency_chart_data   = json.dumps([u['count'] for u in urgency_counts])
    urgency_chart_colors = json.dumps([u['color'] for u in urgency_counts])
    selected_month = (date_from[:7] if date_from else timezone.now().strftime('%Y-%m'))

    def _parse_filter_date(value):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            return None

    def _add_months(value, months):
        month_index = value.month - 1 + months
        year = value.year + month_index // 12
        month = month_index % 12 + 1
        return value.replace(year=year, month=month, day=1)

    def _period_date(value):
        if hasattr(value, 'date'):
            return value.date()
        return value

    _today = timezone.localdate()
    _from_date = _parse_filter_date(date_from)
    _to_date = _parse_filter_date(date_to)

    overview_qs = all_shipments
    if declarant_filter:
        overview_qs = overview_qs.filter(declarant__username=declarant_filter)

    if overview_range == '6m':
        _overview_start = _add_months(_today.replace(day=1), -5)
        _overview_end = _add_months(_today.replace(day=1), 1) - timedelta(days=1)
    else:
        _overview_start = _today.replace(month=1, day=1)
        _overview_end = _today.replace(month=12, day=31)

    _overview_rows = list(
        overview_qs
        .filter(submitted_at__date__gte=_overview_start, submitted_at__date__lte=_overview_end)
        .annotate(period=TruncMonth('submitted_at'))
        .values('period')
        .annotate(count=Count('id'))
        .order_by('period')
    )
    _overview_map = {
        _period_date(r['period']).replace(day=1): r['count']
        for r in _overview_rows
        if r['period']
    }
    _overview_labels = []
    _overview_data = []
    _cursor = _overview_start.replace(day=1)
    while _cursor <= _overview_end:
        _overview_labels.append(_cursor.strftime('%b %Y'))
        _overview_data.append(_overview_map.get(_cursor, 0))
        _cursor = _add_months(_cursor, 1)

    monthly_chart_labels = json.dumps(_overview_labels)
    monthly_chart_data   = json.dumps(_overview_data)

    # Pre-clearance SLA buckets stop once the shipment reaches BOC assessment.
    _preclearance_done_statuses = ['assessed', 'paid', 'released', 'billed']
    _now    = timezone.now()
    _today  = _now.date()
    _d1 = _d3 = _d5 = _d5plus = 0
    _active_qs = chart_qs.exclude(status__in=_preclearance_done_statuses)
    _due_total = _active_qs.count()
    for _s in _active_qs.values('urgency', 'submitted_at'):
        _alloc     = _urgency_days_for(_s['urgency'])
        _deadline  = _add_business_days(_s['submitted_at'], _alloc)
        _remaining = _business_days_diff(_today, _deadline)
        if _remaining <= 1:
            _d1 += 1
        elif _remaining <= 3:
            _d3 += 1
        elif _remaining <= 5:
            _d5 += 1
        else:
            _d5plus += 1
    due_date_data = {
        'one_day': _d1, 'three_days': _d3,
        'five_days': _d5, 'over_five': _d5plus,
        'total': _due_total,
    }
    due_date_chart_data   = json.dumps([_d1, _d3, _d5, _d5plus])
    due_date_chart_labels = json.dumps(['1 Day Left', '3 Days Left', '5 Days Left', '5+ Days Left'])
    due_date_chart_colors = json.dumps(['#dc0000', '#f75b5b', '#f9a1a1', '#ffd6d6'])

    # WMCDA vertical bar chart - fixed order: LCL, Air, FCL
    _wmcda_bar_order = [
        ('lcl',  'LCL Sea',      '#38bdf8'),
        ('air',  'Air Freight',  '#f59e0b'),
        ('fcl',  'FCL Sea',      '#8b5cf6'),
    ]
    _wmap = {r['key']: r for r in wmcda_scoreboard}
    wmcda_bar_labels = json.dumps([b[1] for b in _wmcda_bar_order])
    wmcda_bar_data   = json.dumps([_wmap.get(b[0], {}).get('count', 0) for b in _wmcda_bar_order])
    wmcda_bar_colors = json.dumps([b[2] for b in _wmcda_bar_order])
    wmcda_bar_keys   = json.dumps([b[0] for b in _wmcda_bar_order])

    # Top performing declarant: prioritize real processing volume, then approval quality.
    top_declarant = None
    _eligible = [d for d in declarant_data if d['total_processed'] > 0]
    if _eligible:
        top_declarant = max(_eligible, key=lambda d: (d['total_processed'], d['approval_rate'], d['ecdt_approved']))
        _name_parts = [part for part in top_declarant['name'].split() if part]
        top_declarant['initials'] = ''.join(part[0] for part in _name_parts[:2]).upper()

    # ── Currency usage breakdown ───────────────────────────────────────────────
    from django.db.models import Count as _Count
    _cur_colors = {
        'USD': '#3B82F6', 'EUR': '#8B5CF6', 'JPY': '#F59E0B',
        'HKD': '#EC4899', 'CNY': '#EF4444', 'GBP': '#14B8A6', 'SGD': '#22C55E',
    }
    _cur_qs = (
        Shipment.objects.filter(id__in=_chart_ids_qs)
        .exclude(invoice_currency='')
        .values('invoice_currency')
        .annotate(count=_Count('id'))
        .order_by('-count')
    )
    currency_total = sum(r['count'] for r in _cur_qs)
    currency_breakdown = [
        {
            'code':  r['invoice_currency'] or 'USD',
            'count': r['count'],
            'pct':   round(r['count'] / currency_total * 100, 1) if currency_total else 0,
            'color': _cur_colors.get(r['invoice_currency'] or 'USD', '#94A3B8'),
        }
        for r in _cur_qs
    ]
    currency_chart_labels = json.dumps([r['code']  for r in currency_breakdown])
    currency_chart_data   = json.dumps([r['count'] for r in currency_breakdown])
    currency_chart_colors = json.dumps([r['color'] for r in currency_breakdown])

    # Cost comparison by shipment type — avg/total landed cost per mode
    _cost_qs = DutyComputation.objects.filter(total_landed_cost__isnull=False)
    if date_from:
        _cost_qs = _cost_qs.filter(shipment__submitted_at__date__gte=date_from)
    if date_to:
        _cost_qs = _cost_qs.filter(shipment__submitted_at__date__lte=date_to)
    if declarant_filter:
        _cost_qs = _cost_qs.filter(shipment__declarant__username=declarant_filter)

    _cost_type_meta = [
        ('air',  'Air',  '#F59E0B'),
        ('lcl',  'LCL',  '#38BDF8'),
        ('fcl',  'FCL',  '#8B5CF6'),
    ]
    cost_by_type = []
    for code, label, color in _cost_type_meta:
        agg = _cost_qs.filter(shipment__shipment_type=code).aggregate(
            avg=Avg('total_landed_cost'),
            total=Sum('total_landed_cost'),
            count=Count('id'),
            min_val=Min('total_landed_cost'),
            max_val=Max('total_landed_cost'),
        )
        cost_by_type.append({
            'code': code, 'label': label, 'color': color,
            'avg':   round(float(agg['avg'] or 0), 2),
            'total': round(float(agg['total'] or 0), 2),
            'count': agg['count'],
            'min_val': round(float(agg['min_val'] or 0), 2),
            'max_val': round(float(agg['max_val'] or 0), 2),
        })

    cost_bar_labels = json.dumps([r['label'] for r in cost_by_type])
    cost_bar_data   = json.dumps([r['avg'] for r in cost_by_type])
    cost_bar_colors = json.dumps([r['color'] for r in cost_by_type])

    # Feedback summary — all-time
    _fb_qs       = Feedback.objects.all()
    _fb_total    = _fb_qs.count()
    _fb_avg      = _fb_qs.aggregate(avg=Avg('rating'))['avg']
    _fb_positive = _fb_qs.filter(rating__gte=4).count()
    feedback_summary = {
        'total':        _fb_total,
        'avg_rating':   round(float(_fb_avg), 1) if _fb_avg else 0,
        'positive':     _fb_positive,
        'positive_pct': round(_fb_positive / _fb_total * 100, 1) if _fb_total else 0,
    }
    feedback_summary['filled_stars'] = int(round(feedback_summary['avg_rating'])) if _fb_total else 0
    feedback_summary['star_rows'] = [
        {'value': i, 'filled': i <= feedback_summary['filled_stars']}
        for i in range(1, 6)
    ]

    return render(request, 'supervisor/analytics.html', {
        # KPI strip
        'total_all':                  total_all,
        'total_incoming':             all_shipments.filter(status='incoming').count(),
        'total_arrived':              all_shipments.filter(status='arrived').count(),
        'total_computed':             all_shipments.filter(status='computed').count(),
        'total_approved':             all_shipments.filter(status='approved').count(),
        'total_rejected':             all_shipments.filter(status='rejected').count(),
        'total_users':                User.objects.filter(role__in=['consignee', 'declarant'], is_active=True, is_pending_approval=False).count(),
        'total_consignees':           User.objects.filter(role='consignee', is_active=True, is_pending_approval=False).count(),
        'total_declarants':           User.objects.filter(role='declarant', is_active=True, is_pending_approval=False).count(),
        'consignee_approval_rate':    consignee_approval_rate,
        'total_computed_presented':   total_computed_presented,
        'total_consignee_approved':   total_consignee_approved,
        # chart data
        'chart_total':        chart_total,
        'status_rows':        status_rows_sorted,
        'wmcda_scoreboard':   wmcda_scoreboard,
        'wmcda_max':          wmcda_max,
        'wmcda_total':        wmcda_total,
        'declarant_data':     declarant_data,
        # filters
        'date_from':          date_from,
        'date_to':            date_to,
        'declarant_filter':   declarant_filter,
        'overview_range':     overview_range,
        'declarants':         declarants,
        # chart data
        'pipeline_rows':      pipeline_rows,
        # shipment table
        'recent':    table_qs,
        'q':         q,
        'status_f':  status_f,
        # redesigned dashboard
        'shipment_type_counts':  shipment_type_counts,
        'urgency_counts':        urgency_counts,
        'urgency_total':         urgency_total,
        'urgency_chart_labels':  urgency_chart_labels,
        'urgency_chart_data':    urgency_chart_data,
        'urgency_chart_colors':  urgency_chart_colors,
        'due_date_data':         due_date_data,
        'due_date_chart_data':   due_date_chart_data,
        'due_date_chart_labels': due_date_chart_labels,
        'due_date_chart_colors': due_date_chart_colors,
        'monthly_chart_labels':  monthly_chart_labels,
        'monthly_chart_data':    monthly_chart_data,
        'top_declarant':         top_declarant,
        'feedback_summary':      feedback_summary,
        'selected_month':        selected_month,
        'wmcda_bar_labels':      wmcda_bar_labels,
        'wmcda_bar_data':        wmcda_bar_data,
        'wmcda_bar_colors':      wmcda_bar_colors,
        'wmcda_bar_keys':        wmcda_bar_keys,
        # cost comparison
        'cost_by_type':          cost_by_type,
        'cost_bar_labels':       cost_bar_labels,
        'cost_bar_data':         cost_bar_data,
        'cost_bar_colors':       cost_bar_colors,
        # currency analytics
        'currency_breakdown':      currency_breakdown,
        'currency_total':          currency_total,
        'currency_chart_labels':   currency_chart_labels,
        'currency_chart_data':     currency_chart_data,
        'currency_chart_colors':   currency_chart_colors,
    })


#  Live Status Counts (AJAX) 

@login_required
@supervisor_required
def analytics_status_counts(request):
    from django.http import JsonResponse
    qs = Shipment.objects.all()
    total = qs.count()
    counts = {}
    max_count = 0
    for key, label in Shipment.STATUS_CHOICES:
        c = qs.filter(status=key).count()
        counts[key] = {'count': c, 'label': label}
        if c > max_count:
            max_count = c
    return JsonResponse({'counts': counts, 'total': total, 'max_count': max_count})


#  Supervisor Shipment Detail (read-only) 

@login_required
@supervisor_required
def shipment_detail(request, shipment_id):
    from apps.shipments.status_progress import build_status_progress, CONSIGNEE_STATUS_SUBLABELS
    from apps.shipments.fan import fan_assessment_has_values, fan_assessment_rows
    shipment    = get_object_or_404(Shipment, id=shipment_id)
    advisory    = getattr(shipment, 'shipping_advisory', None)
    computation = getattr(shipment, 'computation', None)
    status_logs = shipment.status_logs.order_by('-changed_at')
    sad_document = shipment.documents.filter(document_type='sad').first()
    fan_rows = fan_assessment_rows(sad_document)
    current_sublabel = CONSIGNEE_STATUS_SUBLABELS.get(shipment.status, '')
    back_url = request.GET.get('return_to') or ''
    back_label = request.GET.get('return_label') or 'Back to Dashboard'
    if not url_has_allowed_host_and_scheme(back_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
        back_url = reverse('supervisor:dashboard')
        back_label = 'Back to Dashboard'

    explanation = wmcda_scores = wmcda_breakdown = None
    declared_score = declared_breakdown = declared_rating = None

    if advisory:
        try:
            from apps.computation.views import compute_wmcda
            wmcda_scores, _, wmcda_breakdown, explanation = compute_wmcda(
                float(advisory.gross_weight), float(advisory.cargo_volume),
                float(advisory.declared_value), advisory.urgency_level,
                float(advisory.distance_km),
            )
            if wmcda_scores and shipment.shipment_type:
                declared_score = wmcda_scores.get(shipment.shipment_type)
                if wmcda_breakdown:
                    declared_breakdown = wmcda_breakdown.get(shipment.shipment_type)
                if declared_score is not None:
                    if declared_score >= 0.80:   declared_rating = 'Excellent'
                    elif declared_score >= 0.65: declared_rating = 'Good'
                    elif declared_score >= 0.50: declared_rating = 'Fair'
                    else:                        declared_rating = 'Poor'
        except Exception:
            pass

    return render(request, 'supervisor/shipment_detail.html', {
        'shipment':           shipment,
        'advisory':           advisory,
        'computation':        computation,
        'status_logs':        status_logs,
        'explanation':        explanation,
        'wmcda_scores':       wmcda_scores,
        'wmcda_breakdown':    wmcda_breakdown,
        'declared_score':     declared_score,
        'declared_breakdown': declared_breakdown,
        'declared_rating':    declared_rating,
        'status_steps':       build_status_progress(shipment.status, 'consignee'),
        'sad_document':       sad_document,
        'fan_assessment_rows': fan_rows,
        'fan_assessment_has_values': fan_assessment_has_values(fan_rows),
        'current_sublabel':   current_sublabel,
        'back_url':           back_url,
        'back_label':         back_label,
    })


#  Memos & Announcements 

def _announcement_recipients(announcement):
    return User.objects.filter(
        role__in=announcement.target_roles(),
        is_active=True,
        is_pending_approval=False,
    )


def _notify_announcement_recipients(announcement):
    recipients = _announcement_recipients(announcement)
    for recipient in recipients:
        create_notification(
            recipient=recipient,
            shipment=None,
            notification_type='announcement',
            title=announcement.title,
            message=announcement.content,
            announcement=announcement,
        )
    announcement.notified_at = timezone.now()
    announcement.save(update_fields=['notified_at', 'updated_at'])
    return recipients.count()


@login_required
@supervisor_required
def list_memos(request):
    memos = Announcement.objects.all()
    return render(request, 'supervisor/memos.html', {'memos': memos})


@login_required
@supervisor_required
def create_memo(request):
    if request.method == 'POST':
        title     = request.POST.get('title', '').strip()
        content   = request.POST.get('content', '').strip()
        category  = request.POST.get('category', 'general')
        audience  = request.POST.get('target_audience', 'all')
        is_active = request.POST.get('is_active') == '1'
        valid_audiences = {choice[0] for choice in Announcement.AUDIENCE_CHOICES}
        if audience not in valid_audiences:
            audience = 'all'
        if not title or not content:
            messages.error(request, 'Title and content are required.')
        else:
            announcement = Announcement.objects.create(
                title=title, content=content,
                category=category, target_audience=audience,
                is_active=is_active, created_by=request.user,
            )
            if announcement.is_active:
                notified_count = _notify_announcement_recipients(announcement)
                messages.success(
                    request,
                    f'Announcement "{title}" published and sent to {notified_count} user(s).',
                )
            else:
                messages.success(request, f'Announcement "{title}" saved as hidden.')
    return redirect('supervisor:memos')


@login_required
@supervisor_required
def delete_memo(request, memo_id):
    if request.method == 'POST':
        memo = get_object_or_404(Announcement, id=memo_id)
        title = memo.title
        memo.delete()
        messages.success(request, f'Announcement "{title}" deleted.')
    return redirect('supervisor:memos')


@login_required
@supervisor_required
def toggle_memo(request, memo_id):
    if request.method == 'POST':
        memo = get_object_or_404(Announcement, id=memo_id)
        memo.is_active = not memo.is_active
        memo.save()
        state = 'published' if memo.is_active else 'archived'
        if memo.is_active and not memo.notified_at:
            notified_count = _notify_announcement_recipients(memo)
            messages.success(request, f'"{memo.title}" {state} and sent to {notified_count} user(s).')
        else:
            messages.success(request, f'"{memo.title}" {state}.')
    return redirect('supervisor:memos')


#  System Configuration 

_CURRENCY_KEYS = ['rate_USD', 'rate_EUR', 'rate_JPY', 'rate_HKD', 'rate_CNY', 'rate_GBP', 'rate_SGD']

_CURRENCY_META = [
    {'key': 'rate_USD', 'code': 'USD', 'name': 'US Dollar',       'symbol': 'USD'},
    {'key': 'rate_EUR', 'code': 'EUR', 'name': 'Euro',             'symbol': 'EUR'},
    {'key': 'rate_JPY', 'code': 'JPY', 'name': 'Japanese Yen',     'symbol': 'JPY'},
    {'key': 'rate_HKD', 'code': 'HKD', 'name': 'Hong Kong Dollar', 'symbol': 'HKD'},
    {'key': 'rate_CNY', 'code': 'CNY', 'name': 'Chinese Yuan',     'symbol': 'CNY'},
    {'key': 'rate_GBP', 'code': 'GBP', 'name': 'British Pound',    'symbol': 'GBP'},
    {'key': 'rate_SGD', 'code': 'SGD', 'name': 'Singapore Dollar', 'symbol': 'SGD'},
]


def _get_config():
    from types import SimpleNamespace
    defaults = {
        'exchange_rate':  '59.1480',   # Legacy USDPHP key (kept for backward compat)
        'rate_USD':       '59.1480',
        'rate_EUR':       '65.0000',
        'rate_JPY':       '0.3900',
        'rate_HKD':       '7.5700',
        'rate_CNY':       '8.1500',
        'rate_GBP':       '74.5000',
        'rate_SGD':       '43.8000',
        'vat_rate':       '12.00',
        'wmcda_w_cost':   '35',
        'wmcda_w_time':   '30',
        'wmcda_w_weight': '20',
        'wmcda_w_distance': '15',
        'urgency_days_standard': '15',
        'urgency_days_priority': '10',
        'urgency_days_urgent':   '5',
        'urgency_days_rush':     '3',
    }
    rows   = {sc.key: sc.value for sc in SystemConfig.objects.all()}
    merged = {k: rows.get(k, v) for k, v in defaults.items()}
    return SimpleNamespace(**merged)


def _config_meta(keys):
    return {
        row.key: row
        for row in SystemConfig.objects.filter(key__in=keys).select_related('updated_by')
    }


@login_required
@supervisor_required
def config_home(request):
    """Landing page  3 large buttons to sub-sections."""
    return render(request, 'supervisor/config.html')


_BF_DEFAULT_TIERS = [
    {'max': 10000,    'fee': '1300'},
    {'max': 20000,    'fee': '2000'},
    {'max': 30000,    'fee': '2700'},
    {'max': 40000,    'fee': '3300'},
    {'max': 50000,    'fee': '3600'},
    {'max': 60000,    'fee': '4000'},
    {'max': 100000,   'fee': '4700'},
    {'max': 200000,   'fee': '5300', 'excess_rate': '0.00125'},
]

_IPF_DEFAULT_TIERS = [
    {'max': 25000,    'fee': '250'},
    {'max': 50000,    'fee': '500'},
    {'max': 250000,   'fee': '750'},
    {'max': 500000,   'fee': '1000'},
    {'max': 750000,   'fee': '1500'},
    {'max': 99999999, 'fee': '2000'},
]


def _load_tiers(key, defaults):
    try:
        raw = SystemConfig.get(key, '')
        return json.loads(raw) if raw else list(defaults)
    except Exception:
        return list(defaults)


def config_global(request):
    if request.method != 'POST':
        ensure_daily_exchange_rates(user=None)

    config   = _get_config()
    urgency_keys = ['urgency_days_standard', 'urgency_days_priority', 'urgency_days_urgent', 'urgency_days_rush']
    rate_status_keys = [
        'exchange_rates_last_success',
        'exchange_rates_last_attempt',
        'exchange_rates_last_error',
        'exchange_rates_source',
    ]
    all_keys = _CURRENCY_KEYS + ['exchange_rate', 'vat_rate'] + urgency_keys + rate_status_keys
    meta     = _config_meta(all_keys)

    if request.method == 'POST':
        for key in _CURRENCY_KEYS + ['vat_rate']:
            val = request.POST.get(key, '').strip()
            if val:
                SystemConfig.objects.update_or_create(
                    key=key, defaults={'value': val, 'updated_by': request.user}
                )
        # Keep legacy exchange_rate in sync with rate_USD
        usd_val = request.POST.get('rate_USD', '').strip()
        if usd_val:
            SystemConfig.objects.update_or_create(
                key='exchange_rate', defaults={'value': usd_val, 'updated_by': request.user}
            )
        # Document template URLs
        for tmpl_key in ['invoice_template_url', 'packing_list_template_url']:
            tmpl_val = request.POST.get(tmpl_key, '').strip()
            SystemConfig.objects.update_or_create(
                key=tmpl_key, defaults={'value': tmpl_val, 'updated_by': request.user}
            )
        for key in urgency_keys:
            val = request.POST.get(key, '').strip()
            try:
                days = int(val)
            except (TypeError, ValueError):
                messages.error(request, 'Urgency business days must be whole numbers.')
                return redirect('supervisor:config_global')
            if not 1 <= days <= 60:
                messages.error(request, 'Urgency business days must be between 1 and 60.')
                return redirect('supervisor:config_global')
            SystemConfig.objects.update_or_create(
                key=key, defaults={'value': str(days), 'updated_by': request.user}
            )
        messages.success(request, 'Global parameters saved.')
        return redirect('supervisor:config_global')

    # Build currency rows for template
    currency_rows = []
    for row in _CURRENCY_META:
        currency_rows.append({
            **row,
            'value': getattr(config, row['key'], '0.0000'),
            'meta':  meta.get(row['key']),
        })

    urgency_rows = [
        {'key': 'urgency_days_standard', 'label': 'Standard', 'value': config.urgency_days_standard, 'meta': meta.get('urgency_days_standard'), 'color': '#3b82f6'},
        {'key': 'urgency_days_priority', 'label': 'Priority', 'value': config.urgency_days_priority, 'meta': meta.get('urgency_days_priority'), 'color': '#f59e0b'},
        {'key': 'urgency_days_urgent',   'label': 'Urgent',   'value': config.urgency_days_urgent,   'meta': meta.get('urgency_days_urgent'),   'color': '#f97316'},
        {'key': 'urgency_days_rush',     'label': 'Rush',     'value': config.urgency_days_rush,     'meta': meta.get('urgency_days_rush'),     'color': '#ef4444'},
    ]

    return render(request, 'supervisor/config_global.html', {
        'config':        config,
        'config_meta':   meta,
        'currency_rows': currency_rows,
        'urgency_rows':  urgency_rows,
        'invoice_template_url':      SystemConfig.get('invoice_template_url', ''),
        'packing_list_template_url': SystemConfig.get('packing_list_template_url', ''),
    })


@login_required
@supervisor_required
def config_fees(request):
    """Brokerage Fee and Import Processing Fee tier editor."""
    if request.method == 'POST':
        # BF tiers
        bf_tiers  = _load_tiers('bf_tiers', _BF_DEFAULT_TIERS)
        bf_changed = False
        for i, tier in enumerate(bf_tiers):
            fee_val = request.POST.get(f'bf_fee_{i}', '').strip()
            if fee_val:
                bf_tiers[i]['fee'] = fee_val
                bf_changed = True
            if 'excess_rate' in tier:
                er_val = request.POST.get('bf_excess_rate', '').strip()
                if er_val:
                    bf_tiers[i]['excess_rate'] = er_val
                    bf_changed = True
        if bf_changed:
            SystemConfig.objects.update_or_create(
                key='bf_tiers',
                defaults={'value': json.dumps(bf_tiers), 'updated_by': request.user}
            )
        # IPF tiers
        ipf_tiers  = _load_tiers('ipf_tiers', _IPF_DEFAULT_TIERS)
        ipf_changed = False
        for i, tier in enumerate(ipf_tiers):
            fee_val = request.POST.get(f'ipf_fee_{i}', '').strip()
            if fee_val:
                ipf_tiers[i]['fee'] = fee_val
                ipf_changed = True
        if ipf_changed:
            SystemConfig.objects.update_or_create(
                key='ipf_tiers',
                defaults={'value': json.dumps(ipf_tiers), 'updated_by': request.user}
            )
        messages.success(request, 'Fee schedules saved.')
        return redirect('supervisor:config_fees')

    bf_tiers  = _load_tiers('bf_tiers',  _BF_DEFAULT_TIERS)
    ipf_tiers = _load_tiers('ipf_tiers', _IPF_DEFAULT_TIERS)

    bf_rows, prev = [], 0
    for i, tier in enumerate(bf_tiers):
        bf_rows.append({
            'index': i, 'from_val': prev + 1, 'max_val': tier['max'],
            'fee': tier['fee'], 'is_last': i == len(bf_tiers) - 1,
            'excess_rate': tier.get('excess_rate', ''),
        })
        prev = tier['max']

    ipf_rows, prev = [], 0
    for i, tier in enumerate(ipf_tiers):
        ipf_rows.append({
            'index': i, 'from_val': prev + 1, 'max_val': tier['max'],
            'fee': tier['fee'], 'is_last': i == len(ipf_tiers) - 1,
        })
        prev = tier['max']

    return render(request, 'supervisor/config_fees.html', {
        'bf_rows': bf_rows, 'ipf_rows': ipf_rows,
    })


@login_required
@supervisor_required
def fetch_exchange_rates(request):
    """Force-refresh live PHP-based rates and save to SystemConfig."""
    from django.http import JsonResponse

    result = ensure_daily_exchange_rates(user=request.user, force=True)
    if result.get('error'):
        return JsonResponse({'ok': False, 'error': result['error']}, status=500)
    return JsonResponse({
        'ok': True,
        'rates': result.get('rates', {}),
        'source': result.get('source', ''),
    })


@login_required
@supervisor_required
def config_wmcda(request):
    config = _get_config()
    meta   = _config_meta(['wmcda_w_cost', 'wmcda_w_time', 'wmcda_w_weight', 'wmcda_w_distance'])
    if request.method == 'POST':
        for key in ('wmcda_w_cost', 'wmcda_w_time', 'wmcda_w_weight', 'wmcda_w_distance'):
            val = request.POST.get(key, '').strip()
            if val:
                SystemConfig.objects.update_or_create(
                    key=key, defaults={'value': val, 'updated_by': request.user}
                )
        messages.success(request, 'WMCDA weights saved.')
        return redirect('supervisor:config_wmcda')
    return render(request, 'supervisor/config_wmcda.html', {'config': config, 'config_meta': meta})


def _selected_tariff_schedule(request):
    schedule_id = (request.POST.get('schedule') or request.GET.get('schedule') or '').strip()
    schedules = list(TariffSchedule.objects.all())
    selected = None
    if schedule_id:
        try:
            selected = next((s for s in schedules if s.id == int(schedule_id)), None)
        except ValueError:
            selected = None
    if selected is None:
        selected = next((s for s in schedules if s.is_active), None)
    if selected is None and schedules:
        selected = schedules[0]
    return schedules, selected


def _apply_schedule_rates(hs_codes, schedule):
    if not hs_codes:
        return hs_codes
    rate_map = {}
    if schedule:
        ids = [hs.id for hs in hs_codes]
        rate_map = {
            rate.hs_code_id: rate
            for rate in HSCodeRate.objects.filter(schedule=schedule, hs_code_id__in=ids)
        }
    for hs in hs_codes:
        schedule_rate = rate_map.get(hs.id)
        hs.display_duty_rate = schedule_rate.duty_rate if schedule_rate else hs.duty_rate
        hs.has_schedule_rate = bool(schedule_rate)
    return hs_codes


def _clean_tariff_text(value):
    if value is None:
        return ''
    return ' '.join(str(value).strip().split())


def _parse_tariff_rate(value):
    if value is None or value == '':
        return None
    try:
        rate = Decimal(str(value).replace('%', '').strip())
    except (InvalidOperation, ValueError):
        return None
    if rate < 0 or rate > 100:
        return None
    return rate


def _unique_tariff_code(base):
    base = slugify(base or 'tariff-schedule')[:70] or 'tariff-schedule'
    candidate = base
    suffix = 2
    while TariffSchedule.objects.filter(code=candidate).exists():
        candidate = f'{base[:65]}-{suffix}'
        suffix += 1
    return candidate


def _read_tariff_workbook(path, rate_column):
    try:
        import openpyxl
    except ImportError:
        raise ValueError('openpyxl is required to import tariff schedules.')

    try:
        with default_storage.open(path, 'rb') as f:
            workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f'Could not open workbook: {exc}')

    try:
        if 'Tariff Codes' not in workbook.sheetnames:
            raise ValueError('Sheet "Tariff Codes" was not found.')

        worksheet = workbook['Tariff Codes']
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [_clean_tariff_text(value).lower() for value in next(rows)]
        except StopIteration:
            raise ValueError('Sheet "Tariff Codes" is empty.')

        required = ['ahtn_code', 'description', 'chapter']
        missing = [col for col in required if col not in headers]
        if missing:
            raise ValueError(f'Missing required column(s): {", ".join(missing)}')

        available_rate_columns = [col for col in headers if col == 'mfn_rate' or col.startswith('mfn_')]
        selected_rate_column = (rate_column or '').strip().lower()
        if not selected_rate_column:
            selected_rate_column = 'mfn_rate' if 'mfn_rate' in headers else (available_rate_columns[-1] if available_rate_columns else '')
        if selected_rate_column not in headers:
            raise ValueError(
                f'Rate column "{selected_rate_column or "(blank)"}" was not found. '
                f'Available MFN columns: {", ".join(available_rate_columns) or "none"}.'
            )

        idx = {name: headers.index(name) for name in set(required + [selected_rate_column])}
        records = {}
        stats = {
            'total_rows': 0,
            'valid_rows': 0,
            'new_codes': 0,
            'existing_codes': 0,
            'blank_or_invalid_rates': 0,
            'missing_required': 0,
            'duplicate_rows': 0,
        }
        warnings = []

        existing_codes = set(HSCode.objects.values_list('code', flat=True))

        for row_number, row in enumerate(rows, start=2):
            stats['total_rows'] += 1
            code = _clean_tariff_text(row[idx['ahtn_code']] if len(row) > idx['ahtn_code'] else '')
            description = _clean_tariff_text(row[idx['description']] if len(row) > idx['description'] else '')
            chapter = _clean_tariff_text(row[idx['chapter']] if len(row) > idx['chapter'] else '')
            rate = _parse_tariff_rate(row[idx[selected_rate_column]] if len(row) > idx[selected_rate_column] else None)

            if not code or not description or not chapter:
                stats['missing_required'] += 1
                if len(warnings) < 12:
                    warnings.append(f'Row {row_number}: missing HS code, description, or chapter.')
                continue
            if rate is None:
                stats['blank_or_invalid_rates'] += 1
                if len(warnings) < 12:
                    warnings.append(f'Row {row_number}: blank or invalid rate for {code}.')
                continue
            if code in records:
                stats['duplicate_rows'] += 1
                if len(warnings) < 12:
                    warnings.append(f'Row {row_number}: duplicate HS code {code}; latest row will be used.')

            records[code] = {
                'code': code,
                'description': description,
                'chapter': chapter,
                'duty_rate': rate,
                'source_row': row_number,
            }

        stats['valid_rows'] = len(records)
        stats['existing_codes'] = sum(1 for code in records if code in existing_codes)
        stats['new_codes'] = stats['valid_rows'] - stats['existing_codes']
        sample_records = list(records.values())[:10]

        return {
            'rate_column': selected_rate_column,
            'available_rate_columns': available_rate_columns,
            'stats': stats,
            'warnings': warnings,
            'records': records,
            'sample_records': sample_records,
        }
    finally:
        workbook.close()


@login_required
@supervisor_required
def config_hscodes_sections(request):
    """Show all 21 HS sections with chapter/code counts."""
    q = request.GET.get('q', '').strip()
    tariff_schedules, selected_schedule = _selected_tariff_schedule(request)
    hs_list = HSCode.objects.filter(is_active=True).values('chapter')
    chapter_counts = {}
    for hs in hs_list:
        ch = _chapter_num(hs['chapter'])
        if ch:
            chapter_counts[ch] = chapter_counts.get(ch, 0) + 1

    sections = []
    for num, roman, title, chapters in _HS_SECTIONS:
        total_codes = sum(chapter_counts.get(ch, 0) for ch in chapters)
        has_data    = sum(1 for ch in chapters if chapter_counts.get(ch, 0) > 0)
        sections.append({
            'num': num, 'roman': roman, 'title': title,
            'total_chapters': len(chapters), 'chapters_with_codes': has_data,
            'total_codes': total_codes,
        })

    search_results = []
    if q:
        search_results = list(
            HSCode.objects.filter(
                Q(code__icontains=q) | Q(description__icontains=q),
                is_active=True,
            ).order_by('code')[:60]
        )
        _apply_schedule_rates(search_results, selected_schedule)
        for hs in search_results:
            hs.chapter_num = _chapter_num(hs.chapter)

    return render(request, 'supervisor/config_hscodes.html', {
        'sections': sections,
        'q': q,
        'search_results': search_results,
        'tariff_schedules': tariff_schedules,
        'selected_schedule': selected_schedule,
    })


@login_required
@supervisor_required
def upload_tariff_schedule(request):
    preview = None
    pending = request.session.get('pending_tariff_import')

    if request.method == 'POST':
        action = request.POST.get('action', 'preview')

        if action == 'confirm':
            if not pending:
                messages.error(request, 'No tariff import is waiting for confirmation.')
                return redirect('supervisor:upload_tariff_schedule')

            try:
                parsed = _read_tariff_workbook(pending['path'], pending['rate_column'])
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect('supervisor:upload_tariff_schedule')

            if parsed['stats']['valid_rows'] == 0:
                messages.error(request, 'No valid tariff rows were found to import.')
                return redirect('supervisor:upload_tariff_schedule')

            make_active = bool(pending.get('make_active'))
            schedule = TariffSchedule.objects.create(
                name=pending['schedule_name'],
                code=_unique_tariff_code(pending.get('schedule_code') or pending['schedule_name']),
                rate_basis='mfn',
                effective_from=pending.get('effective_from') or None,
                effective_to=pending.get('effective_to') or None,
                is_active=make_active,
                source_file=pending.get('original_filename', ''),
                notes=(
                    f'Imported from tariff upload using column {parsed["rate_column"]}. '
                    f'Valid rows: {parsed["stats"]["valid_rows"]}.'
                ),
                imported_by=request.user,
                imported_at=timezone.now(),
            )

            hs_by_code = {
                hs.code: hs
                for hs in HSCode.objects.filter(code__in=parsed['records'].keys())
            }
            new_hs = []
            for code, record in parsed['records'].items():
                if code not in hs_by_code:
                    new_hs.append(HSCode(
                        code=code,
                        description=record['description'],
                        chapter=record['chapter'],
                        duty_rate=record['duty_rate'] if make_active else Decimal('0'),
                        is_active=True,
                    ))
            if new_hs:
                HSCode.objects.bulk_create(new_hs, batch_size=1000)
                hs_by_code.update({
                    hs.code: hs
                    for hs in HSCode.objects.filter(code__in=[obj.code for obj in new_hs])
                })

            rates = []
            current_updates = []
            for code, record in parsed['records'].items():
                hs = hs_by_code.get(code)
                if not hs:
                    continue
                hs.description = record['description']
                hs.chapter = record['chapter']
                hs.is_active = True
                if make_active:
                    hs.duty_rate = record['duty_rate']
                current_updates.append(hs)
                rates.append(HSCodeRate(
                    hs_code=hs,
                    schedule=schedule,
                    duty_rate=record['duty_rate'],
                    source_row=record['source_row'],
                    updated_by=request.user,
                ))

            HSCode.objects.bulk_update(
                current_updates,
                ['description', 'chapter', 'is_active', 'duty_rate'] if make_active else ['description', 'chapter', 'is_active'],
                batch_size=1000,
            )
            HSCodeRate.objects.bulk_create(rates, batch_size=1000)

            request.session.pop('pending_tariff_import', None)
            if default_storage.exists(pending['path']):
                default_storage.delete(pending['path'])

            messages.success(
                request,
                f'Tariff schedule "{schedule.name}" imported with {len(rates)} rate row(s).'
            )
            return redirect(f'{reverse("supervisor:config_hscodes_sections")}?schedule={schedule.id}')

        upload = request.FILES.get('tariff_file')
        schedule_name = request.POST.get('schedule_name', '').strip()
        schedule_code = request.POST.get('schedule_code', '').strip()
        rate_column = request.POST.get('rate_column', '').strip().lower()
        effective_from = request.POST.get('effective_from', '').strip()
        effective_to = request.POST.get('effective_to', '').strip()
        make_active = bool(request.POST.get('make_active'))

        if not upload or not schedule_name:
            messages.error(request, 'Please choose an Excel file and enter a schedule name.')
            return redirect('supervisor:upload_tariff_schedule')
        if TariffSchedule.objects.filter(name=schedule_name).exists():
            messages.error(request, 'A tariff schedule with that name already exists.')
            return redirect('supervisor:upload_tariff_schedule')

        filename = f'tariff_uploads/{uuid.uuid4().hex}_{os.path.basename(upload.name)}'
        saved_path = default_storage.save(filename, upload)
        try:
            parsed = _read_tariff_workbook(saved_path, rate_column)
        except ValueError as exc:
            if default_storage.exists(saved_path):
                default_storage.delete(saved_path)
            messages.error(request, str(exc))
            return redirect('supervisor:upload_tariff_schedule')

        request.session['pending_tariff_import'] = {
            'path': saved_path,
            'original_filename': upload.name,
            'schedule_name': schedule_name,
            'schedule_code': schedule_code,
            'rate_column': parsed['rate_column'],
            'effective_from': effective_from,
            'effective_to': effective_to,
            'make_active': make_active,
        }
        preview = {
            'schedule_name': schedule_name,
            'schedule_code': schedule_code,
            'effective_from': effective_from,
            'effective_to': effective_to,
            'make_active': make_active,
            'filename': upload.name,
            **parsed,
        }

    return render(request, 'supervisor/upload_tariff_schedule.html', {
        'preview': preview,
        'pending': pending,
    })


@login_required
@supervisor_required
def config_hscodes_section(request, section_num):
    """List chapters in one section."""
    from apps.declarant.views import _CHAPTER_TITLES
    tariff_schedules, selected_schedule = _selected_tariff_schedule(request)

    section_data = next((s for s in _HS_SECTIONS if s[0] == section_num), None)
    if not section_data:
        messages.error(request, 'Section not found.')
        return redirect('supervisor:config_hscodes_sections')

    num, roman, title, chapters = section_data
    hs_list = HSCode.objects.filter(is_active=True).values('chapter', 'code')
    chapter_map = {}
    for hs in hs_list:
        ch = _chapter_num(hs['chapter'])
        if ch and ch in chapters:
            chapter_map.setdefault(ch, {'count': 0, 'samples': []})
            chapter_map[ch]['count'] += 1
            if len(chapter_map[ch]['samples']) < 3:
                chapter_map[ch]['samples'].append(hs['code'])

    chapter_list = [
        {
            'num': ch, 'num_str': str(ch).zfill(2),
            'title': _CHAPTER_TITLES.get(ch, ''),
            'count': chapter_map.get(ch, {}).get('count', 0),
            'samples': chapter_map.get(ch, {}).get('samples', []),
        }
        for ch in chapters
    ]
    return render(request, 'supervisor/config_hscodes_section.html', {
        'section_num': num, 'section_roman': roman, 'section_title': title,
        'chapters': chapter_list,
        'tariff_schedules': tariff_schedules,
        'selected_schedule': selected_schedule,
    })


@login_required
@supervisor_required
def config_hscodes_chapter(request, chapter_num):
    """View/edit all HS codes in a specific chapter."""
    q = request.GET.get('q', '').strip()
    tariff_schedules, selected_schedule = _selected_tariff_schedule(request)
    section_data = next(
        ((num, roman, title) for num, roman, title, chs in _HS_SECTIONS if chapter_num in chs),
        (None, '', '')
    )
    section_num, section_roman, section_title = section_data

    all_hs   = list(HSCode.objects.filter(is_active=True).order_by('code'))
    hs_codes = [hs for hs in all_hs if _chapter_num(hs.chapter) == chapter_num]
    _apply_schedule_rates(hs_codes, selected_schedule)

    if request.method == 'POST':
        hs_ids   = request.POST.getlist('hs_id[]')
        hs_rates = request.POST.getlist('hs_rate[]')
        updated  = 0
        for hs_id, rate in zip(hs_ids, hs_rates):
            try:
                hs       = HSCode.objects.get(id=int(hs_id))
                rate_val = float(rate)
                if 0 <= rate_val <= 100:
                    if selected_schedule:
                        HSCodeRate.objects.update_or_create(
                            hs_code=hs,
                            schedule=selected_schedule,
                            defaults={
                                'duty_rate': rate_val,
                                'updated_by': request.user,
                            },
                        )
                    if not selected_schedule or selected_schedule.is_active:
                        hs.duty_rate = rate_val
                        hs.save(update_fields=['duty_rate'])
                    updated += 1
            except (HSCode.DoesNotExist, ValueError):
                pass
        messages.success(request, f'{updated} duty rate{"s" if updated != 1 else ""} saved.')
        redirect_url = reverse('supervisor:config_hscodes_chapter', args=[chapter_num])
        if selected_schedule:
            redirect_url = f'{redirect_url}?schedule={selected_schedule.id}'
        return redirect(redirect_url)

    return render(request, 'supervisor/config_hscodes_chapter.html', {
        'chapter_num': chapter_num,
        'chapter_num_str': str(chapter_num).zfill(2),
        'section_num': section_num, 'section_roman': section_roman,
        'section_title': section_title, 'hs_codes': hs_codes,
        'q': q,
        'tariff_schedules': tariff_schedules,
        'selected_schedule': selected_schedule,
    })


# Keep old URL working (redirect to new home)
@login_required
@supervisor_required
def system_config(request):
    return redirect('supervisor:config_home')


#  Shipment Admin Actions 

@login_required
@supervisor_required
def reset_shipment(request, shipment_id):
    if request.method == 'POST':
        shipment   = get_object_or_404(Shipment, id=shipment_id)
        old_status = shipment.status
        hawb       = shipment.hawb_number

        shipment.status        = 'incoming'
        shipment.declarant     = None
        shipment.processed_at  = None
        shipment.save()

        DutyComputation.objects.filter(shipment=shipment).delete()

        StatusLog.objects.create(
            shipment=shipment,
            changed_by=request.user,
            old_status=old_status,
            new_status='incoming',
            notes='Reset to incoming by supervisor. Computation cleared.',
        )
        messages.success(request, f'Shipment {hawb} reset to Incoming.')
    return redirect('supervisor:dashboard')


@login_required
@supervisor_required
def update_shipment_status(request, shipment_id):
    if request.method == 'POST':
        shipment = get_object_or_404(Shipment, id=shipment_id)
        new_status = request.POST.get('status', '').strip()
        notes = request.POST.get('notes', '').strip()
        allowed = {'approved', 'rejected', 'for_revision'}

        if new_status not in allowed:
            messages.error(request, 'Invalid supervisor status.')
            return redirect('supervisor:dashboard')

        old_status = shipment.status
        shipment.status = new_status
        if new_status in {'approved', 'rejected'} and not shipment.processed_at:
            shipment.processed_at = timezone.now()
        shipment.save()

        StatusLog.objects.create(
            shipment=shipment,
            changed_by=request.user,
            old_status=old_status,
            new_status=new_status,
            notes=notes or f'Supervisor marked shipment {shipment.get_status_display()}.',
        )
        notify_shipment_status_change(
            shipment=shipment,
            old_status=old_status,
            new_status=new_status,
            changed_by=request.user,
            notes=notes,
        )
        messages.success(request, f'Shipment {shipment.hawb_number} marked {shipment.get_status_display()}.')

    return redirect('supervisor:dashboard')


@login_required
@supervisor_required
def delete_shipment(request, shipment_id):
    if request.method == 'POST':
        shipment = get_object_or_404(Shipment, id=shipment_id)
        hawb     = shipment.hawb_number

        # Persist audit record to server logs BEFORE deleting.
        # StatusLog can't survive (CASCADE), so we write to the application log
        # which is retained by Railway and can be reviewed later.
        logger.warning(
            'AUDIT: Shipment %s (consignee=%s, status=%s) permanently deleted by supervisor %s at %s',
            hawb,
            shipment.consignee.username,
            shipment.status,
            request.user.username,
            timezone.now().isoformat(),
        )

        shipment.delete()
        messages.success(request, f'Shipment {hawb} permanently deleted.')
    return redirect('supervisor:dashboard')


#  Feedback Management 

@login_required
@supervisor_required
def manage_feedbacks(request):
    feedbacks = Feedback.objects.select_related('consignee', 'shipment').order_by('-created_at')
    return render(request, 'supervisor/feedbacks.html', {'feedbacks': feedbacks})


@login_required
@supervisor_required
def approve_feedback(request, feedback_id):
    if request.method == 'POST':
        fb = get_object_or_404(Feedback, id=feedback_id)
        fb.is_approved = True
        fb.save()
        messages.success(request, 'Feedback approved  it will now appear on the landing page.')
    return redirect('supervisor:feedbacks')


@login_required
@supervisor_required
def reject_feedback(request, feedback_id):
    if request.method == 'POST':
        fb = get_object_or_404(Feedback, id=feedback_id)
        fb.delete()
        messages.success(request, 'Feedback removed.')
    return redirect('supervisor:feedbacks')


#  System Issue Reports

@login_required
@supervisor_required
def issue_reports(request):
    status_f = request.GET.get('status', '').strip()
    priority_f = request.GET.get('priority', '').strip()
    role_f = request.GET.get('role', '').strip()

    reports = IssueReport.objects.select_related(
        'reporter', 'related_shipment', 'handled_by'
    )
    if status_f:
        reports = reports.filter(status=status_f)
    if priority_f:
        reports = reports.filter(priority=priority_f)
    if role_f:
        reports = reports.filter(reporter_role=role_f)

    counts = {
        'open': IssueReport.objects.filter(status='open').count(),
        'in_review': IssueReport.objects.filter(status='in_review').count(),
        'resolved': IssueReport.objects.filter(status='resolved').count(),
        'closed': IssueReport.objects.filter(status='closed').count(),
    }

    return render(request, 'supervisor/issue_reports.html', {
        'reports': reports,
        'counts': counts,
        'status_choices': IssueReport.STATUS_CHOICES,
        'priority_choices': IssueReport.PRIORITY_CHOICES,
        'active_status': status_f,
        'active_priority': priority_f,
        'active_role': role_f,
    })


@login_required
@supervisor_required
def update_issue_report(request, report_id):
    if request.method != 'POST':
        return redirect('supervisor:issue_reports')

    issue = get_object_or_404(IssueReport, id=report_id)
    old_status = issue.status
    status = request.POST.get('status', issue.status).strip()
    note = request.POST.get('supervisor_note', '').strip()
    valid_statuses = {choice[0] for choice in IssueReport.STATUS_CHOICES}

    if status not in valid_statuses:
        messages.error(request, 'Please choose a valid issue status.')
        return redirect('supervisor:issue_reports')

    issue.status = status
    issue.supervisor_note = note
    issue.handled_by = request.user
    if status in {'resolved', 'closed'} and old_status not in {'resolved', 'closed'}:
        issue.resolved_at = timezone.now()
    elif status not in {'resolved', 'closed'}:
        issue.resolved_at = None
    issue.save(update_fields=[
        'status', 'supervisor_note', 'handled_by', 'resolved_at', 'updated_at',
    ])

    create_notification(
        recipient=issue.reporter,
        shipment=issue.related_shipment,
        notification_type='general',
        title='Issue Report Updated',
        message=(
            f'Your issue report "{issue.title}" is now '
            f'{issue.get_status_display()}. {note or ""}'
        ).strip(),
    )
    messages.success(request, 'Issue report updated and reporter notified.')
    return redirect('supervisor:issue_reports')


#  Shipment Records (dedicated browse page)

@login_required
@supervisor_required
def shipment_records(request):
    q              = request.GET.get('q', '').strip()
    status_f       = request.GET.get('status_f', '').strip()
    stype_f        = request.GET.get('stype', '').strip()
    import_type_f  = request.GET.get('import_type', '').strip()
    date_from      = request.GET.get('date_from', '').strip()
    date_to        = request.GET.get('date_to', '').strip()

    all_shipments = Shipment.objects.select_related('consignee', 'declarant')
    qs = all_shipments.order_by('-submitted_at')
    if q:
        qs = (
            all_shipments.filter(hawb_number__icontains=q)
            | all_shipments.filter(consignee__first_name__icontains=q)
            | all_shipments.filter(consignee__last_name__icontains=q)
            | all_shipments.filter(consignee__username__icontains=q)
        ).select_related('consignee', 'declarant').order_by('-submitted_at')
    if status_f:
        qs = qs.filter(status=status_f)
    if stype_f:
        qs = qs.filter(shipment_type=stype_f)
    if import_type_f:
        qs = qs.filter(import_type=import_type_f)
    if date_from:
        qs = qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(submitted_at__date__lte=date_to)

    stat_qs = all_shipments
    urgency_counts = {
        'standard': stat_qs.filter(urgency__in=['standard', 'normal']).count(),
        'priority': stat_qs.filter(urgency='priority').count(),
        'urgent': stat_qs.filter(urgency='urgent').count(),
        'rush': stat_qs.filter(urgency='rush').count(),
    }
    shipment_type_counts = {
        'air': stat_qs.filter(shipment_type='air').count(),
        'lcl': stat_qs.filter(shipment_type='lcl').count(),
        'fcl': stat_qs.filter(shipment_type='fcl').count(),
    }
    flagged_shipments = qs.filter(has_deficiency=True)
    revision_shipments = qs.filter(status='for_revision')
    flagged_count = flagged_shipments.count()
    revision_count = revision_shipments.count()
    total_shipments = stat_qs.count()
    status_summary = [
        {'key': 'arrived', 'label': 'Arrived', 'count': stat_qs.filter(status='arrived').count(), 'color': '#f59e0b'},
        {'key': 'lodgement', 'label': 'Lodgement', 'count': stat_qs.filter(status='lodgement').count(), 'color': '#06b6d4'},
        {'key': 'paid', 'label': 'Paid', 'count': stat_qs.filter(status='paid').count(), 'color': '#166534'},
        {'key': 'computed', 'label': 'Computed', 'count': stat_qs.filter(status='computed').count(), 'color': '#3b82f6'},
        {'key': 'ongoing', 'label': 'Ongoing', 'count': stat_qs.filter(status='ongoing').count(), 'color': '#f97316'},
        {'key': 'released', 'label': 'Released', 'count': stat_qs.filter(status='released').count(), 'color': '#14b8a6'},
        {'key': 'approved', 'label': 'Approved', 'count': stat_qs.filter(status='approved').count(), 'color': '#22c55e'},
        {'key': 'assessed', 'label': 'Assessed', 'count': stat_qs.filter(status='assessed').count(), 'color': '#8b5cf6'},
        {'key': 'billed', 'label': 'Billed', 'count': stat_qs.filter(status='billed').count(), 'color': '#64748b'},
    ]
    shipping_type_overview = [
        {'key': 'fcl', 'label': 'Full Container Load', 'count': shipment_type_counts['fcl'], 'color': '#6f8b9b'},
        {'key': 'air', 'label': 'Airfreight', 'count': shipment_type_counts['air'], 'color': '#24466e'},
        {'key': 'lcl', 'label': 'Less Container Load', 'count': shipment_type_counts['lcl'], 'color': '#f59e0b'},
    ]
    shipment_type_filter_choices = [
        ('air', 'Air'),
        ('lcl', 'LCL'),
        ('fcl', 'FCL'),
    ]
    for row in shipping_type_overview:
        row['pct'] = round(row['count'] / total_shipments * 100) if total_shipments else 0

    def paginate_records(queryset, param_name):
        paginator = Paginator(queryset, 6)
        page_obj = paginator.get_page(request.GET.get(param_name, 1))

        def page_url(page_number):
            query = request.GET.copy()
            query[param_name] = page_number
            return f'?{query.urlencode()}'

        page_links = [
            {
                'number': number,
                'url': page_url(number),
                'current': number == page_obj.number,
            }
            for number in paginator.page_range
        ]
        return {
            'records': page_obj.object_list,
            'page_obj': page_obj,
            'page_links': page_links,
            'prev_url': page_url(page_obj.previous_page_number()) if page_obj.has_previous() else '',
            'next_url': page_url(page_obj.next_page_number()) if page_obj.has_next() else '',
        }

    def annotate_hold_preview(records):
        today = timezone.localdate()
        for shipment in records:
            start_at  = shipment.deficiency_flagged_at or shipment.submitted_at
            due_date  = _add_business_days(start_at, 3)
            days_left = _business_days_diff(today, due_date)
            if days_left < 0:
                shipment.hold_due_label = f'Overdue {abs(days_left)} Business Day{"s" if abs(days_left) != 1 else ""}'
            elif days_left == 0:
                shipment.hold_due_label = 'Due Today'
            elif days_left == 1:
                shipment.hold_due_label = '1 Business Day Left'
            else:
                shipment.hold_due_label = f'{days_left} Days Left'

    shipments_page = paginate_records(qs, 'shipments_page')
    flagged_page = paginate_records(flagged_shipments, 'flagged_page')
    revision_page = paginate_records(revision_shipments, 'revision_page')
    annotate_hold_preview(flagged_page['records'])

    return render(request, 'supervisor/shipment_records.html', {
        'shipments':      shipments_page['records'],
        'shipments_page': shipments_page,
        'flagged_shipments': flagged_page['records'],
        'flagged_page': flagged_page,
        'flagged_count': flagged_count,
        'revision_shipments': revision_page['records'],
        'revision_page': revision_page,
        'revision_count': revision_count,
        'total_shipments': total_shipments,
        'shipment_type_counts': shipment_type_counts,
        'status_summary': status_summary,
        'shipping_type_overview': shipping_type_overview,
        'urgency_counts': urgency_counts,
        'q':                   q,
        'status_f':            status_f,
        'stype_f':             stype_f,
        'import_type_f':       import_type_f,
        'date_from':           date_from,
        'date_to':             date_to,
        'STATUS_CHOICES':      Shipment.STATUS_CHOICES,
        'TYPE_CHOICES':        shipment_type_filter_choices,
        'IMPORT_TYPE_CHOICES': Shipment.IMPORT_TYPE_CHOICES,
    })


#  Client Lists

@login_required
@supervisor_required
def consignee_list(request):
    q  = request.GET.get('q', '').strip()
    qs = User.objects.filter(role='consignee', is_pending_approval=False).order_by('first_name', 'username')
    if q:
        qs = qs.filter(
            Q(username__icontains=q) | Q(first_name__icontains=q)
            | Q(last_name__icontains=q) | Q(email__icontains=q)
        )
    qs = qs.annotate(shipment_count=Count('shipments'))
    terminal_statuses = ['paid', 'released', 'billed']
    consignee_rows = []
    for consignee in qs:
        shipments = Shipment.objects.filter(consignee=consignee).select_related('declarant').order_by('-submitted_at')
        consignee_rows.append({
            'user': consignee,
            'name': consignee.get_full_name() or consignee.username,
            'company': consignee.company_name or '-',
            'total_shipments': shipments.count(),
            'in_progress_shipments': shipments.exclude(status__in=terminal_statuses).count(),
            'completed_or_billed_shipments': shipments.filter(status__in=terminal_statuses).count(),
            'flagged_shipments': shipments.filter(has_deficiency=True).count(),
        })
    return render(request, 'supervisor/consignee_list.html', {
        'consignees': consignee_rows,
        'total_consignees': qs.count(),
        'q': q,
    })


@login_required
@supervisor_required
def consignee_detail(request, user_id):
    consignee = get_object_or_404(
        User,
        id=user_id,
        role='consignee',
        is_pending_approval=False,
    )
    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    shipments_qs = (
        Shipment.objects
        .filter(consignee=consignee)
        .select_related('consignee', 'declarant')
        .order_by('-submitted_at')
    )
    if q:
        shipments_qs = shipments_qs.filter(
            Q(hawb_number__icontains=q)
            | Q(import_type__icontains=q)
            | Q(status__icontains=q)
            | Q(declarant__first_name__icontains=q)
            | Q(declarant__last_name__icontains=q)
            | Q(declarant__username__icontains=q)
        )
    if date_from:
        shipments_qs = shipments_qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        shipments_qs = shipments_qs.filter(submitted_at__date__lte=date_to)

    all_shipments = Shipment.objects.filter(consignee=consignee)
    terminal_statuses = ['paid', 'released', 'billed']
    paginator = Paginator(shipments_qs, 6)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    def page_url(page_number):
        query = request.GET.copy()
        query['page'] = page_number
        return f'?{query.urlencode()}'

    page_links = [
        {
            'number': number,
            'url': page_url(number),
            'current': number == page_obj.number,
        }
        for number in paginator.page_range
    ]
    shipments_page = {
        'records': page_obj.object_list,
        'page_obj': page_obj,
        'page_links': page_links,
        'prev_url': page_url(page_obj.previous_page_number()) if page_obj.has_previous() else '',
        'next_url': page_url(page_obj.next_page_number()) if page_obj.has_next() else '',
    }

    return render(request, 'supervisor/consignee_detail.html', {
        'consignee': consignee,
        'shipments': shipments_page['records'],
        'shipments_page': shipments_page,
        'detail_return_url': request.get_full_path(),
        'detail_return_label': 'Back to Shipment Records',
        'total_shipments': all_shipments.count(),
        'active_shipments': all_shipments.exclude(status__in=terminal_statuses).count(),
        'flagged_shipments': all_shipments.filter(has_deficiency=True).count(),
        'q': q,
        'date_from': date_from,
        'date_to': date_to,
    })


@login_required
@supervisor_required
def declarant_list(request):
    q  = request.GET.get('q', '').strip()
    qs = User.objects.filter(role='declarant', is_pending_approval=False).order_by('first_name', 'username')
    if q:
        qs = qs.filter(
            Q(username__icontains=q) | Q(first_name__icontains=q)
            | Q(last_name__icontains=q) | Q(email__icontains=q)
        )
    terminal_statuses = ['paid', 'released', 'billed']
    declarant_rows = []
    for declarant in qs:
        shipments = Shipment.objects.filter(declarant=declarant).select_related('consignee').order_by('-submitted_at')
        cleared_statuses = ['approved', 'released', 'billed']
        cleared = shipments.filter(status__in=cleared_statuses).count()
        handled_consignees = shipments.values('consignee_id').distinct().count()
        active = shipments.exclude(status__in=terminal_statuses).count()
        revised = shipments.filter(status='for_revision').count()
        current = shipments.exclude(status__in=terminal_statuses).first() or shipments.first()

        completed_durations = []
        for shipment in shipments.filter(status__in=cleared_statuses):
            end_log = (
                StatusLog.objects
                .filter(shipment=shipment, new_status__in=cleared_statuses)
                .order_by('-changed_at')
                .first()
            )
            end_at = (
                end_log.changed_at if end_log else
                shipment.processed_at or shipment.updated_at
            )
            if end_at and shipment.submitted_at and end_at >= shipment.submitted_at:
                completed_durations.append(end_at - shipment.submitted_at)
        avg_days = None
        if completed_durations:
            avg_days = round(
                sum(duration.total_seconds() for duration in completed_durations)
                / len(completed_durations) / 86400,
                1,
            )

        total = shipments.count()
        clearance_rate = round(cleared / total * 100) if total else 0
        declarant_rows.append({
            'user': declarant,
            'name': declarant.get_full_name() or declarant.username,
            'initial': (declarant.first_name or declarant.username or '?')[:1].upper(),
            'cleared_shipments': cleared,
            'handled_consignees': handled_consignees,
            'active_count': active,
            'revised_count': revised,
            'average_clearance_days': avg_days,
            'current_shipment': current,
            'clearance_rate': clearance_rate,
        })

    top_declarants = sorted(
        declarant_rows,
        key=lambda row: (
            row['cleared_shipments'],
            row['clearance_rate'],
            -(row['average_clearance_days'] or 9999),
        ),
        reverse=True,
    )[:3]
    rank_labels = ['1st', '2nd', '3rd']
    for index, row in enumerate(top_declarants):
        row['rank'] = rank_labels[index]

    return render(request, 'supervisor/declarant_list.html', {
        'declarants': declarant_rows,
        'top_declarants': top_declarants,
        'q': q,
    })


@login_required
@supervisor_required
def declarant_detail(request, user_id):
    declarant = get_object_or_404(
        User,
        id=user_id,
        role='declarant',
        is_pending_approval=False,
    )
    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    shipments_qs = (
        Shipment.objects
        .filter(declarant=declarant)
        .select_related('consignee', 'declarant')
        .order_by('-submitted_at')
    )
    if q:
        shipments_qs = shipments_qs.filter(
            Q(hawb_number__icontains=q)
            | Q(import_type__icontains=q)
            | Q(status__icontains=q)
            | Q(consignee__first_name__icontains=q)
            | Q(consignee__last_name__icontains=q)
            | Q(consignee__username__icontains=q)
        )
    if date_from:
        shipments_qs = shipments_qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        shipments_qs = shipments_qs.filter(submitted_at__date__lte=date_to)

    all_shipments = (
        Shipment.objects
        .filter(declarant=declarant)
        .select_related('consignee', 'declarant')
    )
    cleared_statuses = ['approved', 'released', 'billed']
    terminal_statuses = ['paid', 'released', 'billed']
    preclearance_done_statuses = ['assessed', 'paid', 'released', 'billed']
    now = timezone.now()

    status_counts = {
        row['status']: row['count']
        for row in all_shipments.values('status').annotate(count=Count('id'))
    }
    status_colors = {
        'incoming': '#9DB0C5', 'arrived': '#f59e0b', 'computed': '#2F7FD6',
        'approved': '#20B86F', 'rejected': '#ef4444', 'for_revision': '#F2C715',
        'lodgement': '#06b6d4', 'ongoing': '#FF6A00', 'assessed': '#7c3aed',
        'paid': '#166534', 'released': '#14b8a6', 'billed': '#687481',
    }
    status_display = {
        'for_revision': 'Revision',
        'rejected': 'Flags',
    }
    status_subtitles = {
        'incoming': 'Awaits Declarant Assignment',
        'arrived': 'Awaits ECDT Processing',
        'computed': 'Awaits Consignee Approval',
        'for_revision': 'Returned from Consignee',
        'rejected': 'Rejected by Consignee',
        'approved': 'Proceeding to Lodgement',
        'lodgement': 'Filed with BOC',
        'ongoing': 'For final assessment',
        'assessed': 'Awaits payment',
        'paid': 'Payment received',
        'released': 'Released shipment',
        'billed': 'Fully processed',
    }
    status_order = [
        'incoming', 'approved', 'assessed',
        'arrived', 'for_revision', 'paid',
        'rejected', 'lodgement', 'released',
        'computed', 'ongoing', 'billed',
    ]
    status_rows = []
    total_shipments = all_shipments.count()
    for key in status_order:
        label = dict(Shipment.STATUS_CHOICES).get(key, key.title())
        count = status_counts.get(key, 0)
        status_rows.append({
            'key': key,
            'label': status_display.get(key, label),
            'subtitle': status_subtitles.get(key, ''),
            'count': count,
            'pct': round(count / total_shipments * 100, 1) if total_shipments else 0,
            'color': status_colors.get(key, '#64748B'),
        })

    type_meta = [
        ('fcl', 'Full Container Load (FCL)', '#6F8B9B'),
        ('air', 'Airfreight', '#24466E'),
        ('lcl', 'Less Container Load (LCL)', '#F59E0B'),
    ]
    type_counts = {
        row['shipment_type']: row['count']
        for row in all_shipments.values('shipment_type').annotate(count=Count('id'))
    }
    type_rows = [
        {'key': key, 'label': label, 'color': color, 'count': type_counts.get(key, 0)}
        for key, label, color in type_meta
    ]

    monthly_durations = defaultdict(list)
    completed_durations = []
    for shipment in all_shipments.filter(status__in=cleared_statuses):
        end_log = (
            StatusLog.objects
            .filter(shipment=shipment, new_status__in=cleared_statuses)
            .order_by('-changed_at')
            .first()
        )
        end_at = end_log.changed_at if end_log else shipment.processed_at or shipment.updated_at
        if end_at and shipment.submitted_at and end_at >= shipment.submitted_at:
            days = (end_at - shipment.submitted_at).total_seconds() / 86400
            completed_durations.append(days)
            if shipment.submitted_at.year == now.year:
                monthly_durations[shipment.submitted_at.month].append(days)

    average_clearance_days = round(sum(completed_durations) / len(completed_durations), 1) if completed_durations else None
    on_time_count = sum(1 for days in completed_durations if days <= 3)
    on_time_rate = round(on_time_count / len(completed_durations) * 100) if completed_durations else 0
    trend_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    trend_data = [
        round(sum(monthly_durations[month]) / len(monthly_durations[month]), 1)
        if monthly_durations.get(month) else 0
        for month in range(1, 13)
    ]

    due_buckets = {'one_day': 0, 'three_days': 0, 'five_days': 0, 'over_five': 0}
    _today_d = now.date()
    for shipment in all_shipments.exclude(status__in=preclearance_done_statuses):
        alloc     = _urgency_days_for(shipment.urgency)
        deadline  = _add_business_days(shipment.submitted_at, alloc)
        remaining = _business_days_diff(_today_d, deadline)
        if remaining <= 1:
            due_buckets['one_day'] += 1
        elif remaining <= 3:
            due_buckets['three_days'] += 1
        elif remaining <= 5:
            due_buckets['five_days'] += 1
        else:
            due_buckets['over_five'] += 1
    due_total = sum(due_buckets.values())

    paginator = Paginator(shipments_qs, 6)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    def page_url(page_number):
        query = request.GET.copy()
        query['page'] = page_number
        return f'?{query.urlencode()}'

    page_links = [
        {'number': number, 'url': page_url(number), 'current': number == page_obj.number}
        for number in paginator.page_range
    ]
    shipments_page = {
        'records': page_obj.object_list,
        'page_obj': page_obj,
        'page_links': page_links,
        'prev_url': page_url(page_obj.previous_page_number()) if page_obj.has_previous() else '',
        'next_url': page_url(page_obj.next_page_number()) if page_obj.has_next() else '',
    }

    return render(request, 'supervisor/declarant_detail.html', {
        'declarant': declarant,
        'shipments': shipments_page['records'],
        'shipments_page': shipments_page,
        'total_shipments': total_shipments,
        'active_shipments': all_shipments.exclude(status__in=terminal_statuses).count(),
        'cleared_shipments': all_shipments.filter(status__in=cleared_statuses).count(),
        'handled_consignees': all_shipments.values('consignee_id').distinct().count(),
        'average_clearance_days': average_clearance_days,
        'on_time_rate': on_time_rate,
        'status_rows': status_rows,
        'type_rows': type_rows,
        'type_chart_labels': json.dumps([row['label'] for row in type_rows]),
        'type_chart_data': json.dumps([row['count'] for row in type_rows]),
        'type_chart_colors': json.dumps([row['color'] for row in type_rows]),
        'trend_labels': json.dumps(trend_labels),
        'trend_data': json.dumps(trend_data),
        'trend_year': now.year,
        'due_data': due_buckets,
        'due_total': due_total,
        'due_chart_labels': json.dumps(['1 Day Left', '3 Days Left', '5 Days Left', '5+ Days Left']),
        'due_chart_data': json.dumps([due_buckets['one_day'], due_buckets['three_days'], due_buckets['five_days'], due_buckets['over_five']]),
        'due_chart_colors': json.dumps(['#dc0000', '#f75b5b', '#f9a1a1', '#ffd6d6']),
        'detail_return_url': request.get_full_path(),
        'detail_return_label': 'Back to Declarant Records',
        'q': q,
        'date_from': date_from,
        'date_to': date_to,
    })
