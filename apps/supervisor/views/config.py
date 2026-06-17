import json
import logging
import os
import uuid
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.core.files.storage import default_storage
from django.utils import timezone
from django.utils.text import slugify
from apps.shipments.models import HSCode, TariffSchedule, HSCodeRate
from apps.supervisor.exchange_rates import ensure_daily_exchange_rates
from ..models import SystemConfig

logger = logging.getLogger(__name__)

from .common import *  # noqa: F401,F403

_CURRENCY_KEYS = ['rate_USD', 'rate_EUR', 'rate_JPY', 'rate_HKD', 'rate_CNY', 'rate_GBP', 'rate_SGD']

_CURRENCY_META = [
    {'key': 'rate_USD', 'code': 'USD', 'name': 'US Dollar',       'symbol': 'USD'},
    {'key': 'rate_EUR', 'code': 'EUR', 'name': 'Euro',             'symbol': 'EUR'},
    {'key': 'rate_JPY', 'code': 'JPY', 'name': 'Japanese Yen',     'symbol': 'JPY'},
    {'key': 'rate_HKD', 'code': 'HKD', 'name': 'Hong Kong Dollar', 'symbol': 'HKD'},
    {'key': 'rate_CNY', 'code': 'CNY', 'name': 'Chinese Yuan',     'symbol': 'CNY'},
    {'key': 'rate_GBP', 'code': 'GBP', 'name': 'British Pound',    'symbol': 'GBP'},
    {'key': 'rate_SGD', 'code': 'SGD', 'name': 'Singapore Dollar', 'symbol': 'SGD'},
]


def _get_config():
    from types import SimpleNamespace
    defaults = {
        'exchange_rate':  '59.1480',   # Legacy USDPHP key (kept for backward compat)
        'rate_USD':       '59.1480',
        'rate_EUR':       '65.0000',
        'rate_JPY':       '0.3900',
        'rate_HKD':       '7.5700',
        'rate_CNY':       '8.1500',
        'rate_GBP':       '74.5000',
        'rate_SGD':       '43.8000',
        'vat_rate':       '12.00',
        'wmcda_w_cost':   '35',
        'wmcda_w_time':   '30',
        'wmcda_w_weight': '20',
        'wmcda_w_distance': '15',
        'urgency_days_standard': '15',
        'urgency_days_priority': '10',
        'urgency_days_urgent':   '5',
        'urgency_days_rush':     '3',
    }
    rows   = {sc.key: sc.value for sc in SystemConfig.objects.all()}
    merged = {k: rows.get(k, v) for k, v in defaults.items()}
    return SimpleNamespace(**merged)


def _config_meta(keys):
    return {
        row.key: row
        for row in SystemConfig.objects.filter(key__in=keys).select_related('updated_by')
    }


@login_required
@supervisor_required
def config_home(request):
    """Landing page  3 large buttons to sub-sections."""
    return render(request, 'supervisor/config.html')


_BF_DEFAULT_TIERS = [
    {'max': 10000,    'fee': '1300'},
    {'max': 20000,    'fee': '2000'},
    {'max': 30000,    'fee': '2700'},
    {'max': 40000,    'fee': '3300'},
    {'max': 50000,    'fee': '3600'},
    {'max': 60000,    'fee': '4000'},
    {'max': 100000,   'fee': '4700'},
    {'max': 200000,   'fee': '5300', 'excess_rate': '0.00125'},
]

_IPF_DEFAULT_TIERS = [
    {'max': 25000,    'fee': '250'},
    {'max': 50000,    'fee': '500'},
    {'max': 250000,   'fee': '750'},
    {'max': 500000,   'fee': '1000'},
    {'max': 750000,   'fee': '1500'},
    {'max': 99999999, 'fee': '2000'},
]


def _load_tiers(key, defaults):
    try:
        raw = SystemConfig.get(key, '')
        return json.loads(raw) if raw else list(defaults)
    except Exception:
        return list(defaults)


def config_global(request):
    if request.method != 'POST':
        ensure_daily_exchange_rates(user=None)

    config   = _get_config()
    urgency_keys = ['urgency_days_standard', 'urgency_days_priority', 'urgency_days_urgent', 'urgency_days_rush']
    rate_status_keys = [
        'exchange_rates_last_success',
        'exchange_rates_last_attempt',
        'exchange_rates_last_error',
        'exchange_rates_source',
    ]
    all_keys = _CURRENCY_KEYS + ['exchange_rate', 'vat_rate'] + urgency_keys + rate_status_keys
    meta     = _config_meta(all_keys)

    if request.method == 'POST':
        for key in _CURRENCY_KEYS + ['vat_rate']:
            val = request.POST.get(key, '').strip()
            if val:
                SystemConfig.objects.update_or_create(
                    key=key, defaults={'value': val, 'updated_by': request.user}
                )
        # Keep legacy exchange_rate in sync with rate_USD
        usd_val = request.POST.get('rate_USD', '').strip()
        if usd_val:
            SystemConfig.objects.update_or_create(
                key='exchange_rate', defaults={'value': usd_val, 'updated_by': request.user}
            )
        # Document template URLs
        for tmpl_key in ['invoice_template_url', 'packing_list_template_url']:
            tmpl_val = request.POST.get(tmpl_key, '').strip()
            SystemConfig.objects.update_or_create(
                key=tmpl_key, defaults={'value': tmpl_val, 'updated_by': request.user}
            )
        for key in urgency_keys:
            val = request.POST.get(key, '').strip()
            try:
                days = int(val)
            except (TypeError, ValueError):
                messages.error(request, 'Urgency business days must be whole numbers.')
                return redirect('supervisor:config_global')
            if not 1 <= days <= 60:
                messages.error(request, 'Urgency business days must be between 1 and 60.')
                return redirect('supervisor:config_global')
            SystemConfig.objects.update_or_create(
                key=key, defaults={'value': str(days), 'updated_by': request.user}
            )
        messages.success(request, 'Global parameters saved.')
        return redirect('supervisor:config_global')

    # Build currency rows for template
    currency_rows = []
    for row in _CURRENCY_META:
        currency_rows.append({
            **row,
            'value': getattr(config, row['key'], '0.0000'),
            'meta':  meta.get(row['key']),
        })

    urgency_rows = [
        {'key': 'urgency_days_standard', 'label': 'Standard', 'value': config.urgency_days_standard, 'meta': meta.get('urgency_days_standard'), 'color': '#3b82f6'},
        {'key': 'urgency_days_priority', 'label': 'Priority', 'value': config.urgency_days_priority, 'meta': meta.get('urgency_days_priority'), 'color': '#f59e0b'},
        {'key': 'urgency_days_urgent',   'label': 'Urgent',   'value': config.urgency_days_urgent,   'meta': meta.get('urgency_days_urgent'),   'color': '#f97316'},
        {'key': 'urgency_days_rush',     'label': 'Rush',     'value': config.urgency_days_rush,     'meta': meta.get('urgency_days_rush'),     'color': '#ef4444'},
    ]

    return render(request, 'supervisor/config_global.html', {
        'config':        config,
        'config_meta':   meta,
        'currency_rows': currency_rows,
        'urgency_rows':  urgency_rows,
        'invoice_template_url':      SystemConfig.get('invoice_template_url', ''),
        'packing_list_template_url': SystemConfig.get('packing_list_template_url', ''),
    })


@login_required
@supervisor_required
def config_fees(request):
    """Brokerage Fee and Import Processing Fee tier editor."""
    if request.method == 'POST':
        # BF tiers
        bf_tiers  = _load_tiers('bf_tiers', _BF_DEFAULT_TIERS)
        bf_changed = False
        for i, tier in enumerate(bf_tiers):
            fee_val = request.POST.get(f'bf_fee_{i}', '').strip()
            if fee_val:
                bf_tiers[i]['fee'] = fee_val
                bf_changed = True
            if 'excess_rate' in tier:
                er_val = request.POST.get('bf_excess_rate', '').strip()
                if er_val:
                    bf_tiers[i]['excess_rate'] = er_val
                    bf_changed = True
        if bf_changed:
            SystemConfig.objects.update_or_create(
                key='bf_tiers',
                defaults={'value': json.dumps(bf_tiers), 'updated_by': request.user}
            )
        # IPF tiers
        ipf_tiers  = _load_tiers('ipf_tiers', _IPF_DEFAULT_TIERS)
        ipf_changed = False
        for i, tier in enumerate(ipf_tiers):
            fee_val = request.POST.get(f'ipf_fee_{i}', '').strip()
            if fee_val:
                ipf_tiers[i]['fee'] = fee_val
                ipf_changed = True
        if ipf_changed:
            SystemConfig.objects.update_or_create(
                key='ipf_tiers',
                defaults={'value': json.dumps(ipf_tiers), 'updated_by': request.user}
            )
        messages.success(request, 'Fee schedules saved.')
        return redirect('supervisor:config_fees')

    bf_tiers  = _load_tiers('bf_tiers',  _BF_DEFAULT_TIERS)
    ipf_tiers = _load_tiers('ipf_tiers', _IPF_DEFAULT_TIERS)

    bf_rows, prev = [], 0
    for i, tier in enumerate(bf_tiers):
        bf_rows.append({
            'index': i, 'from_val': prev + 1, 'max_val': tier['max'],
            'fee': tier['fee'], 'is_last': i == len(bf_tiers) - 1,
            'excess_rate': tier.get('excess_rate', ''),
        })
        prev = tier['max']

    ipf_rows, prev = [], 0
    for i, tier in enumerate(ipf_tiers):
        ipf_rows.append({
            'index': i, 'from_val': prev + 1, 'max_val': tier['max'],
            'fee': tier['fee'], 'is_last': i == len(ipf_tiers) - 1,
        })
        prev = tier['max']

    return render(request, 'supervisor/config_fees.html', {
        'bf_rows': bf_rows, 'ipf_rows': ipf_rows,
    })


@login_required
@supervisor_required
def fetch_exchange_rates(request):
    """Force-refresh live PHP-based rates and save to SystemConfig."""
    from django.http import JsonResponse

    result = ensure_daily_exchange_rates(user=request.user, force=True)
    if result.get('error'):
        return JsonResponse({'ok': False, 'error': result['error']}, status=500)
    return JsonResponse({
        'ok': True,
        'rates': result.get('rates', {}),
        'source': result.get('source', ''),
    })


@login_required
@supervisor_required
def config_wmcda(request):
    config = _get_config()
    meta   = _config_meta(['wmcda_w_cost', 'wmcda_w_time', 'wmcda_w_weight', 'wmcda_w_distance'])
    if request.method == 'POST':
        for key in ('wmcda_w_cost', 'wmcda_w_time', 'wmcda_w_weight', 'wmcda_w_distance'):
            val = request.POST.get(key, '').strip()
            if val:
                SystemConfig.objects.update_or_create(
                    key=key, defaults={'value': val, 'updated_by': request.user}
                )
        messages.success(request, 'WMCDA weights saved.')
        return redirect('supervisor:config_wmcda')
    return render(request, 'supervisor/config_wmcda.html', {'config': config, 'config_meta': meta})


def _selected_tariff_schedule(request):
    schedule_id = (request.POST.get('schedule') or request.GET.get('schedule') or '').strip()
    schedules = list(TariffSchedule.objects.all())
    selected = None
    if schedule_id:
        try:
            selected = next((s for s in schedules if s.id == int(schedule_id)), None)
        except ValueError:
            selected = None
    if selected is None:
        selected = next((s for s in schedules if s.is_active), None)
    if selected is None and schedules:
        selected = schedules[0]
    return schedules, selected


def _apply_schedule_rates(hs_codes, schedule):
    if not hs_codes:
        return hs_codes
    rate_map = {}
    if schedule:
        ids = [hs.id for hs in hs_codes]
        rate_map = {
            rate.hs_code_id: rate
            for rate in HSCodeRate.objects.filter(schedule=schedule, hs_code_id__in=ids)
        }
    for hs in hs_codes:
        schedule_rate = rate_map.get(hs.id)
        hs.display_duty_rate = schedule_rate.duty_rate if schedule_rate else hs.duty_rate
        hs.has_schedule_rate = bool(schedule_rate)
    return hs_codes


def _clean_tariff_text(value):
    if value is None:
        return ''
    return ' '.join(str(value).strip().split())


def _parse_tariff_rate(value):
    if value is None or value == '':
        return None
    try:
        rate = Decimal(str(value).replace('%', '').strip())
    except (InvalidOperation, ValueError):
        return None
    if rate < 0 or rate > 100:
        return None
    return rate


def _unique_tariff_code(base):
    base = slugify(base or 'tariff-schedule')[:70] or 'tariff-schedule'
    candidate = base
    suffix = 2
    while TariffSchedule.objects.filter(code=candidate).exists():
        candidate = f'{base[:65]}-{suffix}'
        suffix += 1
    return candidate


def _read_tariff_workbook(path, rate_column):
    try:
        import openpyxl
    except ImportError:
        raise ValueError('openpyxl is required to import tariff schedules.')

    try:
        with default_storage.open(path, 'rb') as f:
            workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f'Could not open workbook: {exc}')

    try:
        if 'Tariff Codes' not in workbook.sheetnames:
            raise ValueError('Sheet "Tariff Codes" was not found.')

        worksheet = workbook['Tariff Codes']
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [_clean_tariff_text(value).lower() for value in next(rows)]
        except StopIteration:
            raise ValueError('Sheet "Tariff Codes" is empty.')

        required = ['ahtn_code', 'description', 'chapter']
        missing = [col for col in required if col not in headers]
        if missing:
            raise ValueError(f'Missing required column(s): {", ".join(missing)}')

        available_rate_columns = [col for col in headers if col == 'mfn_rate' or col.startswith('mfn_')]
        selected_rate_column = (rate_column or '').strip().lower()
        if not selected_rate_column:
            selected_rate_column = 'mfn_rate' if 'mfn_rate' in headers else (available_rate_columns[-1] if available_rate_columns else '')
        if selected_rate_column not in headers:
            raise ValueError(
                f'Rate column "{selected_rate_column or "(blank)"}" was not found. '
                f'Available MFN columns: {", ".join(available_rate_columns) or "none"}.'
            )

        idx = {name: headers.index(name) for name in set(required + [selected_rate_column])}
        records = {}
        stats = {
            'total_rows': 0,
            'valid_rows': 0,
            'new_codes': 0,
            'existing_codes': 0,
            'blank_or_invalid_rates': 0,
            'missing_required': 0,
            'duplicate_rows': 0,
        }
        warnings = []

        existing_codes = set(HSCode.objects.values_list('code', flat=True))

        for row_number, row in enumerate(rows, start=2):
            stats['total_rows'] += 1
            code = _clean_tariff_text(row[idx['ahtn_code']] if len(row) > idx['ahtn_code'] else '')
            description = _clean_tariff_text(row[idx['description']] if len(row) > idx['description'] else '')
            chapter = _clean_tariff_text(row[idx['chapter']] if len(row) > idx['chapter'] else '')
            rate = _parse_tariff_rate(row[idx[selected_rate_column]] if len(row) > idx[selected_rate_column] else None)

            if not code or not description or not chapter:
                stats['missing_required'] += 1
                if len(warnings) < 12:
                    warnings.append(f'Row {row_number}: missing HS code, description, or chapter.')
                continue
            if rate is None:
                stats['blank_or_invalid_rates'] += 1
                if len(warnings) < 12:
                    warnings.append(f'Row {row_number}: blank or invalid rate for {code}.')
                continue
            if code in records:
                stats['duplicate_rows'] += 1
                if len(warnings) < 12:
                    warnings.append(f'Row {row_number}: duplicate HS code {code}; latest row will be used.')

            records[code] = {
                'code': code,
                'description': description,
                'chapter': chapter,
                'duty_rate': rate,
                'source_row': row_number,
            }

        stats['valid_rows'] = len(records)
        stats['existing_codes'] = sum(1 for code in records if code in existing_codes)
        stats['new_codes'] = stats['valid_rows'] - stats['existing_codes']
        sample_records = list(records.values())[:10]

        return {
            'rate_column': selected_rate_column,
            'available_rate_columns': available_rate_columns,
            'stats': stats,
            'warnings': warnings,
            'records': records,
            'sample_records': sample_records,
        }
    finally:
        workbook.close()


@login_required
@supervisor_required
def config_hscodes_sections(request):
    """Show all 21 HS sections with chapter/code counts."""
    q = request.GET.get('q', '').strip()
    tariff_schedules, selected_schedule = _selected_tariff_schedule(request)
    hs_list = HSCode.objects.filter(is_active=True).values('chapter')
    chapter_counts = {}
    for hs in hs_list:
        ch = _chapter_num(hs['chapter'])
        if ch:
            chapter_counts[ch] = chapter_counts.get(ch, 0) + 1

    sections = []
    for num, roman, title, chapters in _HS_SECTIONS:
        total_codes = sum(chapter_counts.get(ch, 0) for ch in chapters)
        has_data    = sum(1 for ch in chapters if chapter_counts.get(ch, 0) > 0)
        sections.append({
            'num': num, 'roman': roman, 'title': title,
            'total_chapters': len(chapters), 'chapters_with_codes': has_data,
            'total_codes': total_codes,
        })

    search_results = []
    if q:
        search_results = list(
            HSCode.objects.filter(
                Q(code__icontains=q) | Q(description__icontains=q),
                is_active=True,
            ).order_by('code')[:60]
        )
        _apply_schedule_rates(search_results, selected_schedule)
        for hs in search_results:
            hs.chapter_num = _chapter_num(hs.chapter)

    return render(request, 'supervisor/config_hscodes.html', {
        'sections': sections,
        'q': q,
        'search_results': search_results,
        'tariff_schedules': tariff_schedules,
        'selected_schedule': selected_schedule,
    })


@login_required
@supervisor_required
def upload_tariff_schedule(request):
    preview = None
    pending = request.session.get('pending_tariff_import')

    if request.method == 'POST':
        action = request.POST.get('action', 'preview')

        if action == 'confirm':
            if not pending:
                messages.error(request, 'No tariff import is waiting for confirmation.')
                return redirect('supervisor:upload_tariff_schedule')

            try:
                parsed = _read_tariff_workbook(pending['path'], pending['rate_column'])
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect('supervisor:upload_tariff_schedule')

            if parsed['stats']['valid_rows'] == 0:
                messages.error(request, 'No valid tariff rows were found to import.')
                return redirect('supervisor:upload_tariff_schedule')

            make_active = bool(pending.get('make_active'))
            schedule = TariffSchedule.objects.create(
                name=pending['schedule_name'],
                code=_unique_tariff_code(pending.get('schedule_code') or pending['schedule_name']),
                rate_basis='mfn',
                effective_from=pending.get('effective_from') or None,
                effective_to=pending.get('effective_to') or None,
                is_active=make_active,
                source_file=pending.get('original_filename', ''),
                notes=(
                    f'Imported from tariff upload using column {parsed["rate_column"]}. '
                    f'Valid rows: {parsed["stats"]["valid_rows"]}.'
                ),
                imported_by=request.user,
                imported_at=timezone.now(),
            )

            hs_by_code = {
                hs.code: hs
                for hs in HSCode.objects.filter(code__in=parsed['records'].keys())
            }
            new_hs = []
            for code, record in parsed['records'].items():
                if code not in hs_by_code:
                    new_hs.append(HSCode(
                        code=code,
                        description=record['description'],
                        chapter=record['chapter'],
                        duty_rate=record['duty_rate'] if make_active else Decimal('0'),
                        is_active=True,
                    ))
            if new_hs:
                HSCode.objects.bulk_create(new_hs, batch_size=1000)
                hs_by_code.update({
                    hs.code: hs
                    for hs in HSCode.objects.filter(code__in=[obj.code for obj in new_hs])
                })

            rates = []
            current_updates = []
            for code, record in parsed['records'].items():
                hs = hs_by_code.get(code)
                if not hs:
                    continue
                hs.description = record['description']
                hs.chapter = record['chapter']
                hs.is_active = True
                if make_active:
                    hs.duty_rate = record['duty_rate']
                current_updates.append(hs)
                rates.append(HSCodeRate(
                    hs_code=hs,
                    schedule=schedule,
                    duty_rate=record['duty_rate'],
                    source_row=record['source_row'],
                    updated_by=request.user,
                ))

            HSCode.objects.bulk_update(
                current_updates,
                ['description', 'chapter', 'is_active', 'duty_rate'] if make_active else ['description', 'chapter', 'is_active'],
                batch_size=1000,
            )
            HSCodeRate.objects.bulk_create(rates, batch_size=1000)

            request.session.pop('pending_tariff_import', None)
            if default_storage.exists(pending['path']):
                default_storage.delete(pending['path'])

            messages.success(
                request,
                f'Tariff schedule "{schedule.name}" imported with {len(rates)} rate row(s).'
            )
            return redirect(f'{reverse("supervisor:config_hscodes_sections")}?schedule={schedule.id}')

        upload = request.FILES.get('tariff_file')
        schedule_name = request.POST.get('schedule_name', '').strip()
        schedule_code = request.POST.get('schedule_code', '').strip()
        rate_column = request.POST.get('rate_column', '').strip().lower()
        effective_from = request.POST.get('effective_from', '').strip()
        effective_to = request.POST.get('effective_to', '').strip()
        make_active = bool(request.POST.get('make_active'))

        if not upload or not schedule_name:
            messages.error(request, 'Please choose an Excel file and enter a schedule name.')
            return redirect('supervisor:upload_tariff_schedule')
        if TariffSchedule.objects.filter(name=schedule_name).exists():
            messages.error(request, 'A tariff schedule with that name already exists.')
            return redirect('supervisor:upload_tariff_schedule')

        filename = f'tariff_uploads/{uuid.uuid4().hex}_{os.path.basename(upload.name)}'
        saved_path = default_storage.save(filename, upload)
        try:
            parsed = _read_tariff_workbook(saved_path, rate_column)
        except ValueError as exc:
            if default_storage.exists(saved_path):
                default_storage.delete(saved_path)
            messages.error(request, str(exc))
            return redirect('supervisor:upload_tariff_schedule')

        request.session['pending_tariff_import'] = {
            'path': saved_path,
            'original_filename': upload.name,
            'schedule_name': schedule_name,
            'schedule_code': schedule_code,
            'rate_column': parsed['rate_column'],
            'effective_from': effective_from,
            'effective_to': effective_to,
            'make_active': make_active,
        }
        preview = {
            'schedule_name': schedule_name,
            'schedule_code': schedule_code,
            'effective_from': effective_from,
            'effective_to': effective_to,
            'make_active': make_active,
            'filename': upload.name,
            **parsed,
        }

    return render(request, 'supervisor/upload_tariff_schedule.html', {
        'preview': preview,
        'pending': pending,
    })


@login_required
@supervisor_required
def config_hscodes_section(request, section_num):
    """List chapters in one section."""
    from apps.declarant.views import _CHAPTER_TITLES
    tariff_schedules, selected_schedule = _selected_tariff_schedule(request)

    section_data = next((s for s in _HS_SECTIONS if s[0] == section_num), None)
    if not section_data:
        messages.error(request, 'Section not found.')
        return redirect('supervisor:config_hscodes_sections')

    num, roman, title, chapters = section_data
    hs_list = HSCode.objects.filter(is_active=True).values('chapter', 'code')
    chapter_map = {}
    for hs in hs_list:
        ch = _chapter_num(hs['chapter'])
        if ch and ch in chapters:
            chapter_map.setdefault(ch, {'count': 0, 'samples': []})
            chapter_map[ch]['count'] += 1
            if len(chapter_map[ch]['samples']) < 3:
                chapter_map[ch]['samples'].append(hs['code'])

    chapter_list = [
        {
            'num': ch, 'num_str': str(ch).zfill(2),
            'title': _CHAPTER_TITLES.get(ch, ''),
            'count': chapter_map.get(ch, {}).get('count', 0),
            'samples': chapter_map.get(ch, {}).get('samples', []),
        }
        for ch in chapters
    ]
    return render(request, 'supervisor/config_hscodes_section.html', {
        'section_num': num, 'section_roman': roman, 'section_title': title,
        'chapters': chapter_list,
        'tariff_schedules': tariff_schedules,
        'selected_schedule': selected_schedule,
    })


@login_required
@supervisor_required
def config_hscodes_chapter(request, chapter_num):
    """View/edit all HS codes in a specific chapter."""
    q = request.GET.get('q', '').strip()
    tariff_schedules, selected_schedule = _selected_tariff_schedule(request)
    section_data = next(
        ((num, roman, title) for num, roman, title, chs in _HS_SECTIONS if chapter_num in chs),
        (None, '', '')
    )
    section_num, section_roman, section_title = section_data

    all_hs   = list(HSCode.objects.filter(is_active=True).order_by('code'))
    hs_codes = [hs for hs in all_hs if _chapter_num(hs.chapter) == chapter_num]
    _apply_schedule_rates(hs_codes, selected_schedule)

    if request.method == 'POST':
        hs_ids   = request.POST.getlist('hs_id[]')
        hs_rates = request.POST.getlist('hs_rate[]')
        updated  = 0
        for hs_id, rate in zip(hs_ids, hs_rates):
            try:
                hs       = HSCode.objects.get(id=int(hs_id))
                rate_val = float(rate)
                if 0 <= rate_val <= 100:
                    if selected_schedule:
                        HSCodeRate.objects.update_or_create(
                            hs_code=hs,
                            schedule=selected_schedule,
                            defaults={
                                'duty_rate': rate_val,
                                'updated_by': request.user,
                            },
                        )
                    if not selected_schedule or selected_schedule.is_active:
                        hs.duty_rate = rate_val
                        hs.save(update_fields=['duty_rate'])
                    updated += 1
            except (HSCode.DoesNotExist, ValueError):
                pass
        messages.success(request, f'{updated} duty rate{"s" if updated != 1 else ""} saved.')
        redirect_url = reverse('supervisor:config_hscodes_chapter', args=[chapter_num])
        if selected_schedule:
            redirect_url = f'{redirect_url}?schedule={selected_schedule.id}'
        return redirect(redirect_url)

    return render(request, 'supervisor/config_hscodes_chapter.html', {
        'chapter_num': chapter_num,
        'chapter_num_str': str(chapter_num).zfill(2),
        'section_num': section_num, 'section_roman': section_roman,
        'section_title': section_title, 'hs_codes': hs_codes,
        'q': q,
        'tariff_schedules': tariff_schedules,
        'selected_schedule': selected_schedule,
    })


# Keep old URL working (redirect to new home)
@login_required
@supervisor_required
def system_config(request):
    return redirect('supervisor:config_home')


#  Shipment Admin Actions 

