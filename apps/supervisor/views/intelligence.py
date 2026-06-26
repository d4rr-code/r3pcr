from collections import defaultdict
from io import BytesIO
import json
import warnings

from django.core.cache import cache
from django.db.models import Avg, Count, F
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from apps.computation.models import ShipmentLineItem, ShippingAdvisory
from apps.computation.views.hs_codes import suggest_hs_codes
from apps.shipments.models import Shipment, ShipmentHSCode, StatusLog

from .common import supervisor_required


TERMINAL_STATUSES = {'released', 'billed'}
REQUIRED_DOCUMENTS = {'invoice', 'packing_list', 'airway_bill'}
IN_PROGRESS_STATUSES = {'arrived', 'computed', 'for_revision', 'lodgement', 'ongoing', 'assessed'}
YEARLY_FORECAST_HISTORY_YEARS = 10
DEFAULT_RISK_WEIGHTS = {
    'status_age_high': 25,
    'status_age_medium': 15,
    'kpi_delayed': 30,
    'kpi_due_soon': 10,
    'deficiency': 25,
    'for_revision': 20,
    'missing_doc': 10,
    'missing_docs_cap': 25,
    'urgent': 10,
    'priority': 5,
    'waiting_action': 8,
}
FORECAST_MODEL_COLORS = {
    'arima': '#F97316',
    'holt_winters': '#8B5CF6',
    'seasonal_naive': '#0EA5E9',
    'moving_average': '#14B8A6',
}


def _status_label(status):
    return dict(Shipment.STATUS_CHOICES).get(status, status.replace('_', ' ').title())


def _duration_days(start, end):
    if not start or not end or end <= start:
        return 0
    return (end - start).total_seconds() / 86400


def _stage_metrics(shipments):
    shipments = list(shipments)
    if not shipments:
        return [], None, 0, 0

    shipment_map = {s.id: s for s in shipments}
    logs_by_shipment = defaultdict(list)
    logs = (
        StatusLog.objects
        .filter(shipment_id__in=shipment_map)
        .order_by('shipment_id', 'changed_at')
    )
    for log in logs:
        logs_by_shipment[log.shipment_id].append(log)

    stage_totals = defaultdict(lambda: {'days': 0.0, 'count': 0})
    completed_totals = []
    for shipment in shipments:
        logs = logs_by_shipment.get(shipment.id, [])
        if logs:
            first = logs[0]
            initial_status = first.old_status or 'incoming'
            initial_days = _duration_days(shipment.submitted_at, first.changed_at)
            if initial_days:
                stage_totals[initial_status]['days'] += initial_days
                stage_totals[initial_status]['count'] += 1

            for index, log in enumerate(logs):
                next_at = logs[index + 1].changed_at if index + 1 < len(logs) else None
                if next_at:
                    end_at = next_at
                elif shipment.status in TERMINAL_STATUSES:
                    end_at = shipment.updated_at
                else:
                    # Open/current stage age belongs in delay-risk scoring, not
                    # historical bottleneck averages. Including it makes old
                    # active revisions look like a 100+ day process average.
                    continue
                days = _duration_days(log.changed_at, end_at)
                if days:
                    stage_totals[log.new_status]['days'] += days
                    stage_totals[log.new_status]['count'] += 1
        else:
            if shipment.status not in TERMINAL_STATUSES:
                continue
            end_at = shipment.updated_at
            days = _duration_days(shipment.submitted_at, end_at)
            if days:
                stage_totals[shipment.status]['days'] += days
                stage_totals[shipment.status]['count'] += 1

        if shipment.status in TERMINAL_STATUSES:
            total_days = _duration_days(shipment.submitted_at, shipment.updated_at)
            if total_days:
                completed_totals.append(total_days)

    stage_rows = []
    for status, data in stage_totals.items():
        avg_days = data['days'] / data['count'] if data['count'] else 0
        stage_rows.append({
            'status': status,
            'label': _status_label(status),
            'avg_days': round(avg_days, 1),
            'count': data['count'],
            'total_days': round(data['days'], 1),
        })
    stage_rows.sort(key=lambda row: row['avg_days'], reverse=True)

    bottleneck = stage_rows[0] if stage_rows else None
    avg_total = round(sum(completed_totals) / len(completed_totals), 1) if completed_totals else 0
    median_total = 0
    if completed_totals:
        ordered = sorted(completed_totals)
        midpoint = len(ordered) // 2
        if len(ordered) % 2:
            median_total = round(ordered[midpoint], 1)
        else:
            median_total = round((ordered[midpoint - 1] + ordered[midpoint]) / 2, 1)
    return stage_rows[:8], bottleneck, avg_total, median_total


def _current_status_age_days(shipment):
    log = (
        StatusLog.objects
        .filter(shipment=shipment, new_status=shipment.status)
        .order_by('-changed_at')
        .first()
    )
    start = log.changed_at if log else shipment.submitted_at
    return _duration_days(start, timezone.now())


def _risk_factor(shipment):
    uploaded = set(shipment.documents.values_list('document_type', flat=True))
    missing_docs = len(REQUIRED_DOCUMENTS - uploaded)
    if shipment.has_deficiency:
        return 'deficiency'
    if shipment.status == 'for_revision':
        return 'for_revision'
    if missing_docs:
        return 'missing_docs'
    if shipment.urgency in {'urgent', 'rush'}:
        return 'urgent'
    if shipment.urgency == 'priority':
        return 'priority'
    if shipment.status in {'computed', 'approved'}:
        return 'waiting_action'
    return 'baseline'


def _shipment_total_days(shipment):
    return _duration_days(shipment.submitted_at, shipment.updated_at)


def _completed_kpi_performance(shipments):
    total = 0
    on_time = 0
    late = 0
    for shipment in shipments:
        if shipment.status not in TERMINAL_STATUSES or not shipment.updated_at:
            continue
        eta_end = shipment.kpi_eta_end
        if not eta_end:
            continue
        total += 1
        completed_date = timezone.localtime(shipment.updated_at).date()
        if completed_date <= eta_end:
            on_time += 1
        else:
            late += 1
    return {
        'total': total,
        'on_time': on_time,
        'late': late,
        'rate': round(on_time / total * 100, 1) if total else 0,
    }


def _trained_delay_model(shipments):
    completed = [
        shipment for shipment in shipments
        if shipment.status in TERMINAL_STATUSES and shipment.submitted_at and shipment.updated_at
    ]
    weights = DEFAULT_RISK_WEIGHTS.copy()
    factor_counts = defaultdict(int)
    delayed_counts = defaultdict(int)
    delayed_total = 0

    for shipment in completed:
        target = shipment.kpi_target_days
        if not target:
            continue
        total_days = _shipment_total_days(shipment)
        if not total_days:
            continue
        factor = _risk_factor(shipment)
        is_delayed = total_days > target[1]
        factor_counts[factor] += 1
        delayed_counts[factor] += int(is_delayed)
        delayed_total += int(is_delayed)

    sample_count = sum(factor_counts.values())
    if sample_count < 8:
        return {
            'weights': weights,
            'sample_count': sample_count,
            'delayed_rate': 0,
            'source': 'Fallback rules',
            'summary': 'Using expert-defined weights until more completed shipments are available for training.',
        }

    baseline_rate = delayed_total / sample_count if sample_count else 0
    trained_map = {
        'deficiency': ('deficiency', 25),
        'for_revision': ('for_revision', 20),
        'missing_docs': ('missing_doc', 10),
        'urgent': ('urgent', 10),
        'priority': ('priority', 5),
        'waiting_action': ('waiting_action', 8),
    }
    for factor, (weight_key, fallback) in trained_map.items():
        count = factor_counts[factor]
        if count < 3:
            weights[weight_key] = fallback
            continue
        factor_rate = delayed_counts[factor] / count
        lift = factor_rate - baseline_rate
        learned_weight = fallback + round(lift * 40)
        weights[weight_key] = max(4, min(35, learned_weight))

    weights['kpi_delayed'] = max(weights['kpi_delayed'], 30)
    weights['status_age_high'] = max(18, min(30, weights['status_age_high'] + round(baseline_rate * 8)))

    return {
        'weights': weights,
        'sample_count': sample_count,
        'delayed_rate': round(baseline_rate * 100, 1),
        'source': 'Historical training',
        'summary': (
            'Weights are adjusted from completed shipment history by comparing each risk factor '
            'against the historical delayed-completion rate.'
        ),
    }


def _risk_for_shipment(shipment, model):
    score = 0
    reasons = []
    actions = []
    age_days = _current_status_age_days(shipment)
    weights = model['weights']

    if age_days >= 5:
        score += weights['status_age_high']
        reasons.append(f'{round(age_days, 1)} days in {_status_label(shipment.status)}')
        actions.append('Review current workflow stage for delay.')
    elif age_days >= 3:
        score += weights['status_age_medium']
        reasons.append(f'{round(age_days, 1)} days in current status')

    timing_status = getattr(shipment, 'kpi_timing_status', '')
    if timing_status == 'delayed':
        score += weights['kpi_delayed']
        reasons.append('Past KPI target')
        actions.append('Escalate KPI-delayed shipment.')
    elif timing_status == 'due_soon':
        score += weights['kpi_due_soon']
        reasons.append('KPI window ending soon')

    if shipment.has_deficiency:
        score += weights['deficiency']
        reasons.append('Document deficiency flagged')
        actions.append('Request or review revised documents.')
    if shipment.status == 'for_revision':
        score += weights['for_revision']
        reasons.append('Waiting for consignee revision')
        actions.append('Follow up document resubmission.')

    uploaded = set(shipment.documents.values_list('document_type', flat=True))
    missing = sorted(REQUIRED_DOCUMENTS - uploaded)
    if missing:
        score += min(len(missing) * weights['missing_doc'], weights['missing_docs_cap'])
        reasons.append('Missing ' + ', '.join(label.replace('_', ' ') for label in missing))
        actions.append('Verify required pre-clearance documents.')

    if shipment.urgency in {'urgent', 'rush'}:
        score += weights['urgent']
        reasons.append(f'{shipment.get_urgency_display()} urgency')
    elif shipment.urgency == 'priority':
        score += weights['priority']
        reasons.append('Priority shipment')

    if shipment.status in {'computed', 'approved'}:
        score += weights['waiting_action']
        reasons.append('Awaiting next clearance action')
        actions.append('Move shipment to the next filing step if verified.')

    score = min(score, 100)
    if score >= 70:
        label = 'High'
    elif score >= 40:
        label = 'Medium'
    else:
        label = 'Low'

    return {
        'shipment': shipment,
        'score': score,
        'label': label,
        'age_days': round(age_days, 1),
        'reasons': reasons[:4] or ['No major delay factors detected'],
        'action': actions[0] if actions else 'Continue normal monitoring.',
    }


def _risk_rows(shipments, risk_filter='high'):
    active_shipments = [
        shipment for shipment in shipments
        if shipment.status not in TERMINAL_STATUSES
    ]
    model = _trained_delay_model(shipments)
    rows = [_risk_for_shipment(shipment, model) for shipment in active_shipments]
    rows.sort(key=lambda row: row['score'], reverse=True)
    distribution = {
        'high': sum(1 for row in rows if row['label'] == 'High'),
        'medium': sum(1 for row in rows if row['label'] == 'Medium'),
        'low': sum(1 for row in rows if row['label'] == 'Low'),
    }
    visible_rows = rows
    if risk_filter in {'high', 'medium', 'low'}:
        visible_rows = [
            row for row in rows
            if row['label'].lower() == risk_filter
        ]
    return visible_rows[:12], distribution, model


def _add_months(date_value, months):
    month = date_value.month - 1 + months
    year = date_value.year + month // 12
    month = month % 12 + 1
    return date_value.replace(year=year, month=month, day=1)


def _month_start(date_value):
    return date_value.replace(day=1)


def _forecast_next_periods(period_counts, periods):
    history = list(period_counts)
    forecasts = []
    for _index in range(periods):
        recent = history[-3:] if len(history) >= 3 else history
        if len(recent) >= 3:
            projected = round((recent[-1] * 0.5) + (recent[-2] * 0.3) + (recent[-3] * 0.2))
        elif recent:
            projected = round(sum(recent) / len(recent))
        else:
            projected = 0
        forecasts.append(projected)
        history.append(projected)
    return forecasts


def _seasonal_naive_forecast(period_counts, periods, unit='month'):
    counts = [int(value or 0) for value in period_counts]
    season_length = 12 if unit == 'month' else 1
    if len(counts) < season_length:
        return None, 'Needs at least one full seasonal cycle.'
    forecasts = []
    for index in range(periods):
        source_index = len(counts) - season_length + (index % season_length)
        forecasts.append(max(0, counts[source_index]))
    return forecasts, 'Uses the same period from the previous cycle.'


def _holt_winters_forecast(period_counts, periods, unit='month'):
    counts = [int(value or 0) for value in period_counts]
    if unit == 'month':
        if len(counts) < 24 or sum(counts) < 12:
            return None, 'Needs at least 24 monthly points for seasonality.'
        model_kwargs = {'trend': 'add', 'seasonal': 'add', 'seasonal_periods': 12}
    else:
        if len(counts) < 5 or sum(counts) < 5:
            return None, 'Needs at least 5 yearly points.'
        model_kwargs = {'trend': 'add', 'seasonal': None}
    cache_key = f"supervisor:holt_winters:{unit}:{periods}:{','.join(str(value) for value in counts)}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached, 'Exponential smoothing with trend and seasonality.'
    try:
        from statsmodels.tsa.holtwinters import ExponentialSmoothing
        model = ExponentialSmoothing(counts, initialization_method='estimated', **model_kwargs)
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            fit = model.fit(optimized=True)
        raw = fit.forecast(periods)
        forecasts = [max(0, int(round(float(value)))) for value in raw]
        cache.set(cache_key, forecasts, 60 * 15)
        return forecasts, 'Exponential smoothing with trend and seasonality.'
    except Exception:
        return None, 'Model could not fit the available history.'


def _arima_forecast(period_counts, periods, unit='month'):
    """Forecast counts with ARIMA when enough history exists.

    Returns (forecasts, model_source) or (None, reason) when the caller should
    use the deterministic moving-average fallback.  statsmodels is optional so
    the app can still run in lightweight local/test environments.
    """
    counts = [int(value or 0) for value in period_counts]
    min_points = 18 if unit == 'month' else 5
    if len(counts) < min_points or sum(counts) < min_points:
        return None, 'Moving average fallback'
    cache_key = f"supervisor:arima:{unit}:{periods}:{','.join(str(value) for value in counts)}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached, 'ARIMA forecast'
    try:
        from statsmodels.tsa.arima.model import ARIMA
        model = ARIMA(counts, order=(1, 1, 1))
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            fit = model.fit()
        raw = fit.forecast(steps=periods)
        forecasts = [max(0, int(round(float(value)))) for value in raw]
        cache.set(cache_key, forecasts, 60 * 15)
        return forecasts, 'ARIMA forecast'
    except Exception:
        return None, 'Moving average fallback'


FORECAST_MODELS = [
    {
        'key': 'arima',
        'label': 'ARIMA',
        'source': 'ARIMA forecast',
        'method': _arima_forecast,
        'description': 'Trend-based time-series model.',
    },
    {
        'key': 'holt_winters',
        'label': 'Holt-Winters',
        'source': 'Holt-Winters forecast',
        'method': _holt_winters_forecast,
        'description': 'Exponential smoothing with trend and seasonality.',
    },
    {
        'key': 'seasonal_naive',
        'label': 'Seasonal Naive',
        'source': 'Seasonal naive forecast',
        'method': _seasonal_naive_forecast,
        'description': 'Uses the same period from the previous cycle.',
    },
    {
        'key': 'moving_average',
        'label': 'Weighted Moving Average',
        'source': 'Moving average forecast',
        'method': lambda counts, periods, unit='month': (
            _forecast_next_periods(counts, periods),
            'Weighted recent-period baseline.',
        ),
        'description': 'Weighted recent-period baseline.',
    },
]


def _backtest_forecast_model(method, period_counts, unit):
    counts = [int(value or 0) for value in period_counts]
    if len(counts) < 6:
        return None
    folds = min(3, max(1, len(counts) // 4))
    errors = []
    for index in range(len(counts) - folds, len(counts)):
        training = counts[:index]
        if not training:
            continue
        predicted, _note = method(training, 1, unit)
        if not predicted:
            continue
        errors.append(abs(counts[index] - predicted[0]))
    if not errors:
        return None
    return round(sum(errors) / len(errors), 1)


def _forecast_model_comparison(period_counts, periods, unit='month'):
    rows = []
    for model in FORECAST_MODELS:
        forecasts, note = model['method'](period_counts, periods, unit)
        available = bool(forecasts)
        mae = _backtest_forecast_model(model['method'], period_counts, unit) if available else None
        rows.append({
            'key': model['key'],
            'label': model['label'],
            'source': model['source'],
            'description': model['description'],
            'forecasts': forecasts or [],
            'total': sum(forecasts or []),
            'mae': mae,
            'mae_label': f'{mae} MAE' if mae is not None else 'Not enough data',
            'available': available,
            'note': note,
            'status': 'Compared' if available else 'Unavailable',
        })

    sample_total = sum(int(value or 0) for value in period_counts)
    moving_average_row = next((row for row in rows if row['key'] == 'moving_average' and row['available']), None)
    compared = [row for row in rows if row['available'] and row['mae'] is not None]
    if sample_total < 10 and moving_average_row:
        recommended = moving_average_row
    elif compared:
        recommended = min(compared, key=lambda row: (row['mae'], row['key'] != 'arima'))
    else:
        available_rows = [row for row in rows if row['available']]
        recommended = next((row for row in available_rows if row['key'] == 'arima'), None)
        recommended = recommended or moving_average_row
        recommended = recommended or (available_rows[0] if available_rows else None)

    if recommended:
        for row in rows:
            if row['key'] == recommended['key']:
                row['status'] = 'Recommended'
        recommended_model = {
            'key': recommended['key'],
            'label': recommended['label'],
            'source': recommended['source'],
            'mae': recommended['mae'],
        }
        forecasts = recommended['forecasts']
        model_source = recommended['source']
    else:
        forecasts = _forecast_next_periods(period_counts, periods)
        model_source = 'Moving average forecast'
        recommended_model = {
            'key': 'moving_average',
            'label': 'Weighted Moving Average',
            'source': model_source,
            'mae': None,
        }
    return forecasts, model_source, recommended_model, rows


def _coerce_forecast_year(value, default_year):
    try:
        year = int(value)
    except (TypeError, ValueError):
        return default_year
    return max(2000, min(default_year + 5, year))


def _coerce_forecast_window(forecast_unit='month', forecast_periods=1):
    unit = 'year' if str(forecast_unit).lower() == 'year' else 'month'
    allowed = {'month': {1, 3, 6, 12}, 'year': {1, 2, 3}}
    try:
        periods = int(forecast_periods)
    except (TypeError, ValueError):
        periods = 1
    if periods not in allowed[unit]:
        periods = 3 if unit == 'year' else 1
    return unit, periods


def _coerce_forecast_model(value):
    allowed = {'all'} | {model['key'] for model in FORECAST_MODELS}
    value = str(value or 'all').lower()
    return value if value in allowed else 'all'


def _month_count(shipments, month_start):
    next_month = _add_months(month_start, 1)
    return sum(
        1 for shipment in shipments
        if shipment.submitted_at and month_start <= timezone.localtime(shipment.submitted_at).date() < next_month
    )


def _year_count(shipments, year_start):
    next_year = year_start.replace(year=year_start.year + 1)
    return sum(
        1 for shipment in shipments
        if shipment.submitted_at and year_start <= timezone.localtime(shipment.submitted_at).date() < next_year
    )


def _workload_forecast(shipments, forecast_periods=1, forecast_unit='month', forecast_year=None, forecast_model='all'):
    forecast_unit = 'year' if str(forecast_unit).lower() == 'year' else 'month'
    forecast_model = _coerce_forecast_model(forecast_model)
    today = timezone.localdate()
    latest_data_year = None
    for shipment in shipments:
        if shipment.submitted_at:
            year = timezone.localtime(shipment.submitted_at).date().year
            latest_data_year = max(latest_data_year or year, year)
    selected_year = _coerce_forecast_year(forecast_year, latest_data_year or today.year)
    if forecast_unit == 'year':
        forecast_periods = 3
        history_start = today.replace(
            year=selected_year - YEARLY_FORECAST_HISTORY_YEARS + 1,
            month=1,
            day=1,
        )
        periods = [
            history_start.replace(year=history_start.year + index)
            for index in range(YEARLY_FORECAST_HISTORY_YEARS)
        ]
        counts = [_year_count(shipments, period) for period in periods]
        model_counts = counts
        future_periods = [today.replace(year=selected_year + index + 1, month=1, day=1) for index in range(forecast_periods)]
        period_labels = [period.strftime('%Y') for period in periods]
        future_labels = [period.strftime('%Y') for period in future_periods]
        forecast_label = f'{selected_year + 1}-{selected_year + forecast_periods}'
        history_label = f'Last {YEARLY_FORECAST_HISTORY_YEARS} years'
        chart_history_label = 'Historical yearly volume'
        recent_label = 'yearly'
    else:
        selected_start = today.replace(year=selected_year, month=1, day=1)
        periods = [selected_start.replace(month=index) for index in range(1, 13)]
        if selected_year < today.year:
            actual_periods = periods
        elif selected_year == today.year:
            actual_periods = periods[:today.month]
        else:
            actual_periods = []
        forecast_anchor = actual_periods[-1] if actual_periods else _add_months(selected_start, -1)
        future_periods = [_add_months(forecast_anchor, index + 1) for index in range(3)]
        model_end = forecast_anchor
        model_start = _add_months(model_end, -35)
        model_periods = [_add_months(model_start, index) for index in range(36)]
        model_counts = [_month_count(shipments, period) for period in model_periods]
        forecast_periods = 3
        counts = [_month_count(shipments, period) for period in actual_periods]
        period_labels = [period.strftime('%b') for period in actual_periods]
        future_labels = [period.strftime('%b') for period in future_periods]
        forecast_label = 'Next 3 months'
        history_label = 'Last 36 months'
        chart_history_label = 'Historical monthly volume'
        recent_label = 'monthly'

    sample_total = sum(model_counts)
    recommended_forecasts, recommended_source, recommended_model, model_comparison = _forecast_model_comparison(
        model_counts,
        forecast_periods,
        forecast_unit,
    )
    selected_row = None
    if forecast_model != 'all':
        selected_row = next(
            (row for row in model_comparison if row['key'] == forecast_model and row['available']),
            None,
        )
    display_row = selected_row or next(
        (row for row in model_comparison if row['key'] == recommended_model['key']),
        None,
    )
    forecasts = display_row['forecasts'] if display_row else recommended_forecasts
    model_source = display_row['source'] if display_row else recommended_source
    displayed_model = {
        'key': display_row['key'] if display_row else recommended_model['key'],
        'label': display_row['label'] if display_row else recommended_model['label'],
        'source': model_source,
        'mae': display_row['mae'] if display_row else recommended_model['mae'],
    }
    projected = sum(forecasts)
    recent_counts = counts[-3:] if len(counts) >= 3 else counts
    recent_average = sum(recent_counts) / len(recent_counts) if recent_counts else 0
    previous = counts[-1] if counts else 0
    active_backlog = sum(1 for shipment in shipments if shipment.status not in TERMINAL_STATUSES)
    period_projection = projected / forecast_periods if forecast_periods else projected
    confidence_margin = max(1, round(max(recent_average, period_projection, 1) * 0.2)) * forecast_periods
    projected_low = max(0, projected - confidence_margin)
    projected_high = projected + confidence_margin
    period_rows = [
        {
            'label': period.strftime('%Y') if forecast_unit == 'year' else period.strftime('%b %Y'),
            'date_range': 'Actual',
            'count': count,
        }
        for period, count in zip((periods if forecast_unit == 'year' else actual_periods), counts)
    ]
    forecast_rows = [
        {
            'label': period.strftime('%Y') if forecast_unit == 'year' else period.strftime('%b %Y'),
            'date_range': 'Forecast',
            'count': count,
        }
        for period, count in zip(future_periods, forecasts)
    ]
    if previous:
        trend_pct = round(((period_projection - previous) / previous) * 100, 1)
    elif period_projection:
        trend_pct = 100
    else:
        trend_pct = 0

    if recent_average and period_projection >= recent_average * 1.25:
        level = 'Heavy'
        interpretation = f'Incoming workload is projected above the recent {recent_label} baseline.'
        action = 'Prepare extra declarant capacity and monitor incoming queue assignments.'
        confidence = 'Moderate'
    elif recent_average and period_projection <= recent_average * 0.75:
        level = 'Light'
        interpretation = f'Incoming workload is projected below the recent {recent_label} baseline.'
        action = 'Use available time for HS review, document checks, and backlog cleanup.'
        confidence = 'Moderate'
    else:
        level = 'Normal'
        interpretation = f'Incoming workload is close to the recent {recent_label} baseline.'
        action = 'Maintain current declarant assignment and continue daily monitoring.'
        confidence = 'High' if recommended_model['mae'] is not None and sample_total >= 24 else 'Moderate'
    if sample_total < 10 or recommended_model['mae'] is None:
        confidence = 'Low'
        if sample_total < 10:
            interpretation = f'The {history_label.lower()} volume is limited, so this forecast should be treated as an indicative estimate.'
        action = 'Use this as a directional signal and keep monitoring incoming volume.'

    def forecast_series_values(model_forecasts):
        if forecast_unit == 'month':
            if actual_periods:
                values = [None] * (len(actual_periods) - 1) + [counts[-1]] + model_forecasts
            else:
                values = model_forecasts
        else:
            values = [None] * (len(counts) - 1) + [counts[-1]] + model_forecasts
        return (values + [None] * len(labels))[:len(labels)]

    if forecast_unit == 'month':
        labels = period_labels + future_labels
        actual_count_map = {period: count for period, count in zip(actual_periods, counts)}
        historical_values = [actual_count_map.get(period) for period in actual_periods] + [None] * forecast_periods
    else:
        labels = period_labels + future_labels
        historical_values = counts + [None] * forecast_periods
    forecast_values = forecast_series_values(forecasts)
    visible_model_rows = [
        row for row in model_comparison
        if row['available'] and (forecast_model == 'all' or row['key'] == forecast_model)
    ]
    forecast_datasets = [
        {
            'key': row['key'],
            'label': row['label'],
            'values': forecast_series_values(row['forecasts']),
            'color': FORECAST_MODEL_COLORS.get(row['key'], '#14B8A6'),
            'recommended': row['key'] == recommended_model['key'],
            'selected': row['key'] == displayed_model['key'],
        }
        for row in visible_model_rows
    ]

    return {
        'forecast_months': forecast_periods,
        'forecast_periods': forecast_periods,
        'forecast_unit': forecast_unit,
        'forecast_year': selected_year,
        'forecast_model': forecast_model,
        'forecast_label': forecast_label,
        'history_label': history_label,
        'model_source': model_source,
        'recommended_model': recommended_model,
        'displayed_model': displayed_model,
        'model_comparison': model_comparison,
        'projected_next_7_days': projected,
        'projected_period_total': projected,
        'projected_low': projected_low,
        'projected_high': projected_high,
        'recent_average': round(recent_average, 1),
        'trend_pct': trend_pct,
        'level': level,
        'confidence': confidence,
        'sample_total': sample_total,
        'active_backlog': active_backlog,
        'pressure_total': active_backlog + projected,
        'interpretation': interpretation,
        'action': action,
        'period_rows': period_rows[-6:] + forecast_rows,
        'chart': {
            'labels': labels,
            'historical_label': chart_history_label,
            'historical_values': historical_values,
            'forecast_values': forecast_values,
            'forecast_datasets': forecast_datasets,
        },
    }


def _hs_review_rows():
    historical_count = ShipmentHSCode.objects.filter(is_confirmed=True).count()
    top_hs_codes = list(
        ShipmentHSCode.objects
        .filter(is_confirmed=True)
        .values('hs_code__code', 'hs_code__description')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    avg_confidence = (
        ShipmentLineItem.objects
        .exclude(confidence=0)
        .aggregate(avg=Avg('confidence'))['avg']
    )

    candidates = (
        ShipmentLineItem.objects
        .select_related('shipment', 'hs_code')
        .order_by('-updated_at')[:80]
    )
    rows = []
    for item in candidates:
        if not item.description:
            continue
        needs_review = not item.hs_code_id or float(item.confidence or 0) < 0.65 or not item.is_confirmed
        if not needs_review:
            continue
        suggestions = suggest_hs_codes(item.description, top_n=3)
        rows.append({
            'item': item,
            'suggestions': suggestions,
            'confidence_pct': item.confidence_pct,
            'needs_code': not item.hs_code_id,
        })
        if len(rows) >= 10:
            break

    return {
        'rows': rows,
        'historical_count': historical_count,
        'top_hs_codes': top_hs_codes,
        'avg_confidence_pct': round(float(avg_confidence or 0) * 100, 1),
        'review_count': len(rows),
    }


def _shipping_advisory_support(shipments):
    shipment_ids = [shipment.id for shipment in shipments]
    advisory_qs = ShippingAdvisory.objects.filter(
        shipment_id__in=shipment_ids,
        recommended_type__in=['air', 'lcl', 'fcl'],
    ).select_related('shipment')
    labels = {
        'air': 'Air Freight',
        'lcl': 'LCL - Less Container Load',
        'fcl': 'FCL - Full Container Load',
    }
    rows = []
    total = advisory_qs.count()
    for key in ['air', 'lcl', 'fcl']:
        count = advisory_qs.filter(recommended_type=key).count()
        rows.append([
            labels[key],
            count,
            f'{round(count / total * 100, 1) if total else 0}%',
        ])
    agreement_total = advisory_qs.exclude(shipment__shipment_type__isnull=True).exclude(
        shipment__shipment_type=''
    ).count()
    agreement_count = advisory_qs.filter(
        shipment__shipment_type__isnull=False,
        shipment__shipment_type__in=['air', 'lcl', 'fcl'],
        recommended_type=F('shipment__shipment_type'),
    ).count()
    return {
        'total': total,
        'rows': rows,
        'agreement_total': agreement_total,
        'agreement_rate': round(agreement_count / agreement_total * 100, 1) if agreement_total else 0,
    }


def _intelligence_context(risk_filter='high', forecast_periods=1, forecast_unit='month', forecast_year=None, forecast_model='all'):
    shipments = (
        Shipment.objects
        .select_related('consignee', 'declarant')
        .prefetch_related('documents')
        .order_by('-submitted_at')
    )
    shipments_list = list(shipments[:2000])

    stage_rows, bottleneck, avg_total_days, median_total_days = _stage_metrics(shipments_list)
    risk_filter = risk_filter if risk_filter in {'high', 'medium', 'low', 'all'} else 'high'
    risk_rows, risk_distribution, delay_model = _risk_rows(shipments_list, risk_filter)
    hs_review = _hs_review_rows()
    workload_forecast = _workload_forecast(shipments_list, forecast_periods, forecast_unit, forecast_year, forecast_model)
    shipping_advisory_support = _shipping_advisory_support(shipments_list)
    delayed_count = sum(
        1 for shipment in shipments_list
        if shipment.status not in TERMINAL_STATUSES
        and getattr(shipment, 'kpi_timing_status', '') == 'delayed'
    )
    completed_kpi = _completed_kpi_performance(shipments_list)
    on_time_rate = completed_kpi['rate']

    stage_chart = {
        'labels': [row['label'] for row in stage_rows[:6]],
        'values': [row['avg_days'] for row in stage_rows[:6]],
    }
    risk_chart = {
        'labels': ['High', 'Medium', 'Low'],
        'values': [
            risk_distribution['high'],
            risk_distribution['medium'],
            risk_distribution['low'],
        ],
    }
    hs_chart = {
        'labels': [row['hs_code__code'] for row in hs_review['top_hs_codes']],
        'values': [row['count'] for row in hs_review['top_hs_codes']],
    }
    model_chart = {
        'labels': ['Status age', 'KPI delayed', 'Deficiency', 'Missing docs', 'Urgency'],
        'values': [
            delay_model['weights']['status_age_high'],
            delay_model['weights']['kpi_delayed'],
            delay_model['weights']['deficiency'],
            delay_model['weights']['missing_doc'],
            delay_model['weights']['urgent'],
        ],
    }

    return {
        'stage_rows': stage_rows,
        'bottleneck': bottleneck,
        'avg_total_days': avg_total_days,
        'median_total_days': median_total_days,
        'risk_rows': risk_rows,
        'risk_distribution': risk_distribution,
        'risk_filter': risk_filter,
        'risk_filters': [
            {'key': 'high', 'label': 'High', 'count': risk_distribution['high']},
            {'key': 'medium', 'label': 'Medium', 'count': risk_distribution['medium']},
            {'key': 'low', 'label': 'Low', 'count': risk_distribution['low']},
            {'key': 'all', 'label': 'All', 'count': sum(risk_distribution.values())},
        ],
        'delay_model': delay_model,
        'workload_forecast': workload_forecast,
        'forecast_model_options': [{'key': 'all', 'label': 'All Models'}] + [
            {'key': model['key'], 'label': model['label']}
            for model in FORECAST_MODELS
        ],
        'forecast_year_options': list(range(timezone.localdate().year - 5, timezone.localdate().year + 4)),
        'hs_review': hs_review,
        'shipping_advisory_support': shipping_advisory_support,
        'delayed_count': delayed_count,
        'on_time_rate': on_time_rate,
        'completed_kpi_total': completed_kpi['total'],
        'completed_kpi_on_time': completed_kpi['on_time'],
        'completed_kpi_late': completed_kpi['late'],
        'stage_chart_json': json.dumps(stage_chart),
        'risk_chart_json': json.dumps(risk_chart),
        'hs_chart_json': json.dumps(hs_chart),
        'model_chart_json': json.dumps(model_chart),
        'workload_chart_json': json.dumps(workload_forecast['chart']),
        'generated_at': timezone.localtime(),
    }


@supervisor_required
def intelligence(request):
    return render(
        request,
        'supervisor/intelligence.html',
        _intelligence_context(
            request.GET.get('risk', 'high'),
            request.GET.get('forecast_months', 1),
            request.GET.get('forecast_unit', 'month'),
            request.GET.get('forecast_year'),
            request.GET.get('forecast_model', 'all'),
        ),
    )


def _export_rows(context):
    decision_support = [
        ['Purpose', 'Supports supervisor and declarant decisions during pre-clearance; it does not replace licensed customs-broker judgment.'],
        ['Projected Incoming Workload', 'Forecasts expected workload so supervisors can plan declarant capacity and monitor queues.'],
        ['Delay Risk', 'Ranks active shipments that may need attention based on status age, KPI timing, deficiencies, missing documents, and urgency.'],
        ['HS Code Review', 'Highlights line items that may require Harmonized System code review before final computation.'],
        ['Shipping Type Advisory', 'Uses Weighted Multi-Criteria Decision Analysis to recommend Air Freight, LCL, or FCL based on shipment profile.'],
    ]
    ecdt_explanation = [
        ['Input', 'Line-item EXW/FOB value, freight, insurance, HS code duty rate, fees, and exchange rates.'],
        ['Dutiable Value', 'EXW/FOB value plus freight and insurance, converted to PHP.'],
        ['Customs Duty', 'Dutiable value multiplied by the selected HS code duty rate.'],
        ['Total Landed Cost', 'Dutiable value plus customs duty, brokerage fee, IPF, CDS, arrastre, wharfage, and bank charges.'],
        ['BOC Payable', 'Customs duty plus VAT, IPF, CDS, and FCL container security fee when applicable.'],
    ]
    mcda_explanation = [
        ['Alternatives', 'Air Freight, LCL - Less Container Load, and FCL - Full Container Load.'],
        ['Criteria', 'Cost, transit time/urgency, gross weight or cargo size, and origin distance.'],
        ['Weights', 'Supervisor-configured weights, optionally derived using Saaty AHP and checked with Consistency Ratio.'],
        ['Output', 'A recommended shipment type with comparative scores for each alternative.'],
        ['Scope', 'The result is advisory and may be overridden when operational factors outside the model apply.'],
    ]
    summary = [
        ['Average Processing Time', f"{context['avg_total_days']} days"],
        ['Median Processing Time', f"{context['median_total_days']} days"],
        ['On-Time Rate', f"{context['on_time_rate']}%"],
        ['Completed KPI Samples', context['completed_kpi_total']],
        ['Completed On Time', context['completed_kpi_on_time']],
        ['Completed Late', context['completed_kpi_late']],
        ['High Risk Shipments', context['risk_distribution']['high']],
        ['Medium Risk Shipments', context['risk_distribution']['medium']],
        ['Low Risk Shipments', context['risk_distribution']['low']],
        ['Active Delayed Shipments', context['delayed_count']],
        [f"Projected Incoming {context['workload_forecast']['forecast_label']}", context['workload_forecast']['projected_period_total']],
        ['Projected Workload Range', f"{context['workload_forecast']['projected_low']} - {context['workload_forecast']['projected_high']}"],
        ['Projected Workload Level', context['workload_forecast']['level']],
        ['Forecast Confidence', context['workload_forecast']['confidence']],
        ['Forecast Model', context['workload_forecast']['model_source']],
        ['Active Backlog', context['workload_forecast']['active_backlog']],
        ['Forecast Interpretation', context['workload_forecast']['interpretation']],
        ['Delay Model Source', context['delay_model']['source']],
        ['Delay Model Training Samples', context['delay_model']['sample_count']],
        ['HS Records Confirmed', context['hs_review']['historical_count']],
        ['HS Items Needing Review', context['hs_review']['review_count']],
        ['Shipping Advisory Samples', context['shipping_advisory_support']['total']],
        ['Declared vs Recommended Agreement', f"{context['shipping_advisory_support']['agreement_rate']}%"],
    ]
    advisory_rows = context['shipping_advisory_support']['rows']
    stages = [
        [row['label'], row['avg_days'], row['count'], row['total_days']]
        for row in context['stage_rows']
    ]
    risks = [
        [
            row['shipment'].hawb_number,
            row['shipment'].get_status_display(),
            row['label'],
            row['score'],
            '; '.join(row['reasons']),
            row['action'],
        ]
        for row in context['risk_rows']
    ]
    hs_rows = [
        [
            row['item'].description,
            row['item'].hs_code.code if row['item'].hs_code else 'No code selected',
            f"{row['confidence_pct']}%",
            ', '.join(hs.code for hs in row['suggestions']) or 'No suggestion found',
            row['item'].shipment.hawb_number,
        ]
        for row in context['hs_review']['rows']
    ]
    return summary, decision_support, ecdt_explanation, mcda_explanation, advisory_rows, stages, risks, hs_rows


@supervisor_required
def intelligence_export(request):
    fmt = (request.GET.get('format') or 'xlsx').lower()
    context = _intelligence_context(
        request.GET.get('risk', 'high'),
        request.GET.get('forecast_months', 1),
        request.GET.get('forecast_unit', 'month'),
        request.GET.get('forecast_year'),
        request.GET.get('forecast_model', 'all'),
    )
    summary, decision_support, ecdt_explanation, mcda_explanation, advisory_rows, stages, risks, hs_rows = _export_rows(context)
    filename_date = timezone.localtime().strftime('%Y%m%d')

    if fmt == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=24,
            leftMargin=24,
            topMargin=24,
            bottomMargin=24,
        )
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle('cell', fontName='Helvetica', fontSize=8, leading=10)
        story = [
            Paragraph('R3-PCR Pre-Clearance Intelligence Report', styles['Title']),
            Paragraph(f"Generated: {context['generated_at'].strftime('%b %d, %Y %I:%M %p')}", styles['Normal']),
            Spacer(1, 12),
        ]
        tables = [
            ('Executive Summary', ['Metric', 'Value'], summary),
            ('Decision Support Scope', ['Area', 'Explanation'], decision_support),
            ('ECDT Computation Basis', ['Step', 'Explanation'], ecdt_explanation),
            ('Shipping Type Advisory - MCDA Basis', ['Area', 'Explanation'], mcda_explanation),
            ('Shipping Advisory Recommendation Mix', ['Recommended Type', 'Count', 'Share'], advisory_rows),
            ('Processing Bottlenecks', ['Stage', 'Avg Days', 'Transitions', 'Total Days'], stages),
            ('Delay Risk', ['Shipment', 'Status', 'Risk', 'Score', 'Reasons', 'Action'], risks),
            ('HS Code Review', ['Description', 'Current HS', 'Confidence', 'Suggestions', 'Shipment'], hs_rows),
        ]
        for title, headers, rows in tables:
            story.append(Paragraph(title, styles['Heading2']))
            body = rows or [['No data'] + [''] * (len(headers) - 1)]
            table_data = [headers] + [
                [Paragraph(str(cell), cell_style) for cell in row]
                for row in body
            ]
            table = Table(table_data, repeatRows=1, hAlign='LEFT')
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3358')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#DCE5EF')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.extend([table, Spacer(1, 12)])
        doc.build(story)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="R3PCR_Pre_Clearance_Intelligence_{filename_date}.pdf"'
        return response

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_fill = PatternFill('solid', fgColor='1B3358')
    header_font = Font(color='FFFFFF', bold=True)

    sheets = [
        ('Summary', ['Metric', 'Value'], summary),
        ('Decision Support', ['Area', 'Explanation'], decision_support),
        ('ECDT Basis', ['Step', 'Explanation'], ecdt_explanation),
        ('MCDA Advisory', ['Area', 'Explanation'], mcda_explanation),
        ('Advisory Mix', ['Recommended Type', 'Count', 'Share'], advisory_rows),
        ('Bottlenecks', ['Stage', 'Average Days', 'Transitions', 'Total Days'], stages),
        ('Delay Risk', ['Shipment', 'Status', 'Risk', 'Score', 'Reasons', 'Action'], risks),
        ('HS Review', ['Description', 'Current HS', 'Confidence', 'Suggestions', 'Shipment'], hs_rows),
    ]
    for index, (title, headers, rows) in enumerate(sheets):
        sheet = wb.active if index == 0 else wb.create_sheet(title)
        sheet.title = title
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            sheet.append(row)
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 55)

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="R3PCR_Pre_Clearance_Intelligence_{filename_date}.xlsx"'
    return response
