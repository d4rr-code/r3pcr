import logging
from collections import defaultdict
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Avg, Sum, Min, Max
from django.utils import timezone
from django.http import HttpResponse
from apps.accounts.models import User
from apps.shipments.models import Shipment, StatusLog
from apps.computation.models import DutyComputation, ShippingAdvisory
from apps.consignee.models import Feedback

logger = logging.getLogger(__name__)

from .common import *  # noqa: F401,F403
from .analytics_sections import *  # noqa: F401,F403

@login_required
@supervisor_required
def dashboard(request):
    """Unified analytics/command-centre page."""
    return _analytics_context_response(request)


def _analytics_report_data(request):
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    declarant_filter = request.GET.get('declarant', '').strip()

    qs = Shipment.objects.select_related('consignee', 'declarant').all()
    if date_from:
        qs = qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(submitted_at__date__lte=date_to)
    if declarant_filter:
        qs = qs.filter(declarant__username=declarant_filter)

    total = qs.count()
    status_rows = []
    status_counts = {
        row['status']: row['count']
        for row in qs.values('status').annotate(count=Count('id'))
    }
    for key, label in Shipment.STATUS_CHOICES:
        count = status_counts.get(key, 0)
        status_rows.append([label, count, f'{round(count / total * 100, 1) if total else 0}%'])

    type_rows = []
    type_counts = {
        row['shipment_type']: row['count']
        for row in qs.values('shipment_type').annotate(count=Count('id'))
    }
    for key, label in Shipment.SHIPMENT_TYPE_CHOICES:
        count = type_counts.get(key, 0)
        type_rows.append([label, count, f'{round(count / total * 100, 1) if total else 0}%'])

    urgency_map = defaultdict(int)
    for row in qs.values('urgency').annotate(count=Count('id')):
        key = 'standard' if row['urgency'] in ('normal', 'standard', None) else row['urgency']
        urgency_map[key] += row['count']
    urgency_labels = dict(Shipment.URGENCY_CHOICES)
    urgency_rows = []
    for key in ('standard', 'priority', 'urgent', 'rush'):
        count = urgency_map.get(key, 0)
        urgency_rows.append([urgency_labels.get(key, key.title()), count, f'{round(count / total * 100, 1) if total else 0}%'])

    ids = qs.values_list('id', flat=True)
    advisory_qs = ShippingAdvisory.objects.filter(shipment_id__in=ids)
    wmcda_total = advisory_qs.filter(recommended_type__isnull=False).count()
    wmcda_labels = {'air': 'Air Freight', 'lcl': 'LCL Sea', 'fcl': 'FCL Sea'}
    wmcda_counts = {
        row['recommended_type']: row['count']
        for row in advisory_qs.values('recommended_type').annotate(count=Count('id'))
        if row['recommended_type']
    }
    wmcda_avg = advisory_qs.aggregate(
        avg_air=Avg('air_score'), avg_lcl=Avg('lcl_score'),
        avg_fcl=Avg('fcl_score'),
    )
    wmcda_rows = []
    for key, label in wmcda_labels.items():
        count = wmcda_counts.get(key, 0)
        wmcda_rows.append([
            label,
            count,
            f'{round(count / wmcda_total * 100, 1) if wmcda_total else 0}%',
            f'{round(float(wmcda_avg.get(f"avg_{key}") or 0) * 100, 1)}%',
        ])

    currency_total = qs.exclude(invoice_currency='').count()
    currency_rows = []
    for row in qs.exclude(invoice_currency='').values('invoice_currency').annotate(count=Count('id')).order_by('-count'):
        count = row['count']
        currency_rows.append([
            row['invoice_currency'] or 'USD',
            count,
            f'{round(count / currency_total * 100, 1) if currency_total else 0}%',
        ])

    cost_rows = []
    cost_qs = DutyComputation.objects.filter(shipment_id__in=ids, total_landed_cost__isnull=False)
    for key, label in Shipment.SHIPMENT_TYPE_CHOICES:
        agg = cost_qs.filter(shipment__shipment_type=key).aggregate(
            count=Count('id'), avg=Avg('total_landed_cost'), total=Sum('total_landed_cost'),
            min_val=Min('total_landed_cost'), max_val=Max('total_landed_cost')
        )
        cost_rows.append([
            label,
            agg['count'],
            round(float(agg['avg'] or 0), 2),
            round(float(agg['total'] or 0), 2),
            round(float(agg['min_val'] or 0), 2),
            round(float(agg['max_val'] or 0), 2),
        ])

    declarant_rows = []
    for dec in User.objects.filter(role='declarant').order_by('first_name', 'username'):
        dec_qs = qs.filter(declarant=dec)
        assigned = dec_qs.count()
        computed = dec_qs.filter(status__in=['computed', 'approved', 'lodgement', 'ongoing', 'assessed', 'paid', 'released', 'billed']).count()
        completed = dec_qs.filter(status='billed').count()
        revision_flags = dec_qs.filter(status__in=['for_revision', 'rejected']).count()
        if assigned or computed or completed or revision_flags:
            declarant_rows.append([
                dec.get_full_name() or dec.username,
                assigned,
                computed,
                completed,
                revision_flags,
                f'{round(completed / assigned * 100, 1) if assigned else 0}%',
            ])

    feedback_qs = Feedback.objects.filter(shipment_id__in=ids)
    feedback_total = feedback_qs.count()
    feedback_avg = feedback_qs.aggregate(avg=Avg('rating'))['avg']
    feedback_positive = feedback_qs.filter(rating__gte=4).count()

    recent_rows = []
    for s in qs.order_by('-submitted_at')[:100]:
        recent_rows.append([
            s.hawb_number,
            s.consignee.get_full_name() or s.consignee.username,
            s.get_shipment_type_display() or '',
            s.get_status_display(),
            s.get_urgency_display(),
            s.submitted_at.strftime('%Y-%m-%d'),
        ])

    return {
        'generated_at': timezone.localtime().strftime('%Y-%m-%d %H:%M'),
        'filters': {
            'date_from': date_from or 'All',
            'date_to': date_to or 'All',
            'declarant': declarant_filter or 'All',
        },
        'summary': [
            ['Total Shipments', total],
            ['Active Users', User.objects.filter(role__in=['consignee', 'declarant'], is_active=True, is_pending_approval=False).count()],
            ['Consignees', User.objects.filter(role='consignee', is_active=True, is_pending_approval=False).count()],
            ['Declarants', User.objects.filter(role='declarant', is_active=True, is_pending_approval=False).count()],
            ['MCDA Advisories', wmcda_total],
            ['Feedback Count', feedback_total],
            ['Average Feedback Rating', round(float(feedback_avg or 0), 1)],
            ['Positive Feedback %', f'{round(feedback_positive / feedback_total * 100, 1) if feedback_total else 0}%'],
        ],
        'tables': [
            ('Status Pipeline', ['Status', 'Count', 'Share'], status_rows),
            ('Shipment Types', ['Type', 'Count', 'Share'], type_rows),
            ('Urgency Distribution', ['Urgency', 'Count', 'Share'], urgency_rows),
            ('MCDA Recommendations', ['Mode', 'Recommended Count', 'Share', 'Average Score'], wmcda_rows),
            ('Currency Usage', ['Currency', 'Count', 'Share'], currency_rows),
            ('Landed Cost By Mode', ['Mode', 'Computations', 'Average PHP', 'Total PHP', 'Min PHP', 'Max PHP'], cost_rows),
            ('Declarant Performance', ['Declarant', 'Assigned', 'Computed+', 'Billed', 'Revision/Rejected', 'Completion %'], declarant_rows),
            ('Recent Shipments', ['HAWB', 'Consignee', 'Mode', 'Status', 'Urgency', 'Submitted'], recent_rows),
        ],
    }


def _pdf_col_widths(table_data, avail_width):
    """Distribute the available page width across columns in proportion to each
    column's widest cell, so report tables fill the page (left-aligned) instead
    of shrinking to content and floating in the centre. Every column keeps a
    sensible minimum share so narrow columns stay readable."""
    ncols = len(table_data[0])
    natural = [1] * ncols
    for row in table_data:
        for i in range(ncols):
            cell = row[i] if i < len(row) else ''
            natural[i] = max(natural[i], len(str(cell if cell is not None else '')))
    # Apply a floor so a very short column (e.g. a count) still reads cleanly.
    floor = max(natural) * 0.18
    natural = [max(n, floor) for n in natural]
    total = sum(natural)
    return [avail_width * (n / total) for n in natural]


@login_required
@supervisor_required
def analytics_export(request):
    fmt = (request.GET.get('format') or 'xlsx').lower()
    data = _analytics_report_data(request)
    filename_date = timezone.localtime().strftime('%Y%m%d')

    if fmt == 'pdf':
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle('cell', fontName='Helvetica', fontSize=8, leading=10)
        story = [
            Paragraph('R3-PCR Analytics Report', styles['Title']),
            Paragraph(f"Generated: {data['generated_at']}", styles['Normal']),
            Paragraph(
                f"Filters: From {data['filters']['date_from']} to {data['filters']['date_to']} | Declarant: {data['filters']['declarant']}",
                styles['Normal'],
            ),
            Spacer(1, 12),
        ]
        for title, headers, rows in [('Executive Summary', ['Metric', 'Value'], data['summary'])] + data['tables']:
            story.append(Paragraph(title, styles['Heading2']))
            body = rows or [['No data', ''] + [''] * (len(headers) - 2)]
            col_widths = _pdf_col_widths([headers] + body, doc.width)
            # Header cells stay plain strings (styled by TableStyle); body cells
            # are wrapped Paragraphs so long text wraps within its column.
            table_data = [list(headers)] + [
                [Paragraph(str(c if c is not None else ''), cell_style)
                 for c in (list(row) + [''] * (len(headers) - len(row)))]
                for row in body
            ]
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign='LEFT')
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3358')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#DCE5EF')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.extend([tbl, Spacer(1, 12)])
        doc.build(story)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="R3PCR_Analytics_Report_{filename_date}.pdf"'
        return response

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    header_fill = PatternFill('solid', fgColor='1B3358')
    header_font = Font(color='FFFFFF', bold=True)

    ws.append(['R3-PCR Analytics Report'])
    ws['A1'].font = Font(bold=True, size=16)
    ws.append(['Generated', data['generated_at']])
    ws.append(['Date From', data['filters']['date_from'], 'Date To', data['filters']['date_to'], 'Declarant', data['filters']['declarant']])
    ws.append([])
    ws.append(['Metric', 'Value'])
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
    for row in data['summary']:
        ws.append(row)

    for title, headers, rows in data['tables']:
        sheet = wb.create_sheet(title[:31])
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            sheet.append(row)
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="R3PCR_Analytics_Report_{filename_date}.xlsx"'
    return response


#  User Management 


@login_required
@supervisor_required
def analytics(request):
    return redirect('supervisor:dashboard')



def _analytics_filters(request, all_shipments):
    """Parse GET filters and build the chart + shipment-table querysets."""
    date_from        = request.GET.get('date_from', '').strip()
    date_to          = request.GET.get('date_to', '').strip()
    declarant_filter = request.GET.get('declarant', '').strip()
    overview_range   = request.GET.get('overview_range', 'year').strip().lower()
    if overview_range not in {'all', 'year', '6m'}:
        overview_range = 'year'

    chart_qs = all_shipments
    if date_from:
        chart_qs = chart_qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        chart_qs = chart_qs.filter(submitted_at__date__lte=date_to)
    if declarant_filter:
        chart_qs = chart_qs.filter(declarant__username=declarant_filter)

    q        = request.GET.get('q', '').strip()
    status_f = request.GET.get('status_f', '').strip()

    table_qs = all_shipments.order_by('-submitted_at')
    if q:
        table_qs = (
            all_shipments.filter(hawb_number__icontains=q)
            | all_shipments.filter(consignee__first_name__icontains=q)
            | all_shipments.filter(consignee__last_name__icontains=q)
            | all_shipments.filter(consignee__username__icontains=q)
        ).order_by('-submitted_at')
    if status_f:
        table_qs = table_qs.filter(status=status_f)
    if date_from:
        table_qs = table_qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        table_qs = table_qs.filter(submitted_at__date__lte=date_to)

    return {
        'date_from': date_from, 'date_to': date_to,
        'declarant_filter': declarant_filter, 'overview_range': overview_range,
        'q': q, 'status_f': status_f,
        'chart_qs': chart_qs, 'chart_total': chart_qs.count(),
        'table_qs': table_qs,
    }


def _kpi_strip():
    """All-time computed/approved presentation counts + consignee approval rate."""
    total_computed_presented = (
        StatusLog.objects.filter(new_status='computed')
        .values('shipment_id').distinct().count()
    )
    total_consignee_approved = (
        StatusLog.objects.filter(new_status='approved')
        .values('shipment_id').distinct().count()
    )
    return {
        'total_computed_presented': total_computed_presented,
        'total_consignee_approved': total_consignee_approved,
        'consignee_approval_rate': (
            round(total_consignee_approved / total_computed_presented * 100, 1)
            if total_computed_presented else 0
        ),
    }


def _analytics_context_response(request):
    all_shipments = Shipment.objects.all()

    # Filters + chart/table querysets
    _f = _analytics_filters(request, all_shipments)
    date_from        = _f['date_from']
    date_to          = _f['date_to']
    declarant_filter = _f['declarant_filter']
    overview_range   = _f['overview_range']
    q                = _f['q']
    status_f         = _f['status_f']
    chart_qs         = _f['chart_qs']
    chart_total      = _f['chart_total']
    table_qs         = _f['table_qs']

    # KPI strip shipment counts respect the active analytics filters.
    filtered_shipments = chart_qs
    total_all = chart_total
    _kpi = _kpi_strip()
    total_computed_presented = _kpi['total_computed_presented']
    total_consignee_approved = _kpi['total_consignee_approved']
    consignee_approval_rate  = _kpi['consignee_approval_rate']

    # Materialise chart_qs IDs once — reused for status, MCDA and declarant sections
    _chart_ids_qs = chart_qs.values_list('id', flat=True)

    # Status breakdown bar chart (respects chart filters)
    _status = _status_breakdown(_chart_ids_qs, chart_total)
    pipeline_rows      = _status['pipeline_rows']
    status_rows_sorted = _status['status_rows_sorted']

    # MCDA scoreboard + declared-vs-recommended agreement matrix
    advisory_qs = ShippingAdvisory.objects.filter(shipment_id__in=_chart_ids_qs)
    _sb = _wmcda_scoreboard(advisory_qs)
    wmcda_scoreboard = _sb['wmcda_scoreboard']
    wmcda_max        = _sb['wmcda_max']
    wmcda_total      = _sb['wmcda_total']
    _cmp = _wmcda_comparison(advisory_qs)
    wmcda_comparison_rows      = _cmp['wmcda_comparison_rows']
    wmcda_comparison_agreement = _cmp['wmcda_comparison_agreement']
    wmcda_comparison_total     = _cmp['wmcda_comparison_total']

    # Declarant Performance (respects chart filters)
    declarants = User.objects.filter(role='declarant').order_by('first_name', 'username')
    declarant_data = _declarant_performance(_chart_ids_qs, declarants)

    # ── Redesigned dashboard: new context variables ──────────────────────

    # Shipment type KPI counts respect the active analytics filters.
    shipment_type_counts = _shipment_type_counts(filtered_shipments)

    # Urgency distribution (respects chart filters)
    _urg = _urgency_distribution(chart_qs)
    urgency_counts       = _urg['urgency_counts']
    urgency_total        = _urg['urgency_total']
    urgency_chart_labels = _urg['urgency_chart_labels']
    urgency_chart_data   = _urg['urgency_chart_data']
    urgency_chart_colors = _urg['urgency_chart_colors']
    selected_month = (date_from[:7] if date_from else timezone.now().strftime('%Y-%m'))

    # Monthly submission overview line chart
    _overview = _monthly_overview(chart_qs, overview_range)
    monthly_chart_labels = _overview['monthly_chart_labels']
    monthly_chart_data   = _overview['monthly_chart_data']
    monthly_chart_has_data = _overview['monthly_chart_has_data']
    monthly_chart_caption = _overview['monthly_chart_caption']

    # Pre-clearance SLA countdown buckets
    _due = _due_date_buckets(chart_qs)
    due_date_data         = _due['due_date_data']
    due_date_chart_data   = _due['due_date_chart_data']
    due_date_chart_labels = _due['due_date_chart_labels']
    due_date_chart_colors = _due['due_date_chart_colors']

    # MCDA vertical bar chart (fixed LCL / Air / FCL order)
    _bar = _wmcda_bar_chart(wmcda_scoreboard)
    wmcda_bar_labels = _bar['wmcda_bar_labels']
    wmcda_bar_data   = _bar['wmcda_bar_data']
    wmcda_bar_colors = _bar['wmcda_bar_colors']
    wmcda_bar_keys   = _bar['wmcda_bar_keys']

    # Top performing declarant
    top_declarant = _top_declarant(declarant_data)

    # ── Currency usage breakdown ───────────────────────────────────────────────
    _cur = _currency_breakdown(_chart_ids_qs)
    currency_breakdown    = _cur['currency_breakdown']
    currency_total        = _cur['currency_total']
    currency_chart_labels = _cur['currency_chart_labels']
    currency_chart_data   = _cur['currency_chart_data']
    currency_chart_colors = _cur['currency_chart_colors']

    # Cost comparison by shipment type — avg/total landed cost per mode
    _cost = _cost_by_type(date_from, date_to, declarant_filter)
    cost_by_type    = _cost['cost_by_type']
    cost_bar_labels = _cost['cost_bar_labels']
    cost_bar_data   = _cost['cost_bar_data']
    cost_bar_colors = _cost['cost_bar_colors']
    cost_bar_keys   = _cost['cost_bar_keys']

    # Feedback summary respects the active analytics filters.
    feedback_summary = _feedback_summary(_chart_ids_qs)

    return render(request, 'supervisor/analytics.html', {
        # KPI strip
        'total_all':                  total_all,
        'total_incoming':             filtered_shipments.filter(status='incoming').count(),
        'total_arrived':              filtered_shipments.filter(status='arrived').count(),
        'total_computed':             filtered_shipments.filter(status='computed').count(),
        'total_approved':             filtered_shipments.filter(status='approved').count(),
        'total_rejected':             filtered_shipments.filter(status='rejected').count(),
        'total_users':                User.objects.filter(role__in=['consignee', 'declarant'], is_active=True, is_pending_approval=False).count(),
        'total_consignees':           User.objects.filter(role='consignee', is_active=True, is_pending_approval=False).count(),
        'total_declarants':           User.objects.filter(role='declarant', is_active=True, is_pending_approval=False).count(),
        'consignee_approval_rate':    consignee_approval_rate,
        'total_computed_presented':   total_computed_presented,
        'total_consignee_approved':   total_consignee_approved,
        # chart data
        'chart_total':        chart_total,
        'status_rows':        status_rows_sorted,
        'wmcda_scoreboard':   wmcda_scoreboard,
        'wmcda_max':          wmcda_max,
        'wmcda_total':        wmcda_total,
        'wmcda_comparison_rows':      wmcda_comparison_rows,
        'wmcda_comparison_agreement': wmcda_comparison_agreement,
        'wmcda_comparison_total':     wmcda_comparison_total,
        'declarant_data':     declarant_data,
        # filters
        'date_from':          date_from,
        'date_to':            date_to,
        'declarant_filter':   declarant_filter,
        'overview_range':     overview_range,
        'declarants':         declarants,
        # chart data
        'pipeline_rows':      pipeline_rows,
        # shipment table
        'recent':    table_qs,
        'q':         q,
        'status_f':  status_f,
        # redesigned dashboard
        'shipment_type_counts':  shipment_type_counts,
        'urgency_counts':        urgency_counts,
        'urgency_total':         urgency_total,
        'urgency_chart_labels':  urgency_chart_labels,
        'urgency_chart_data':    urgency_chart_data,
        'urgency_chart_colors':  urgency_chart_colors,
        'due_date_data':         due_date_data,
        'due_date_chart_data':   due_date_chart_data,
        'due_date_chart_labels': due_date_chart_labels,
        'due_date_chart_colors': due_date_chart_colors,
        'monthly_chart_labels':  monthly_chart_labels,
        'monthly_chart_data':    monthly_chart_data,
        'monthly_chart_has_data': monthly_chart_has_data,
        'monthly_chart_caption': monthly_chart_caption,
        'top_declarant':         top_declarant,
        'feedback_summary':      feedback_summary,
        'selected_month':        selected_month,
        'wmcda_bar_labels':      wmcda_bar_labels,
        'wmcda_bar_data':        wmcda_bar_data,
        'wmcda_bar_colors':      wmcda_bar_colors,
        'wmcda_bar_keys':        wmcda_bar_keys,
        # cost comparison
        'cost_by_type':          cost_by_type,
        'cost_bar_labels':       cost_bar_labels,
        'cost_bar_data':         cost_bar_data,
        'cost_bar_colors':       cost_bar_colors,
        'cost_bar_keys':         cost_bar_keys,
        # currency analytics
        'currency_breakdown':      currency_breakdown,
        'currency_total':          currency_total,
        'currency_chart_labels':   currency_chart_labels,
        'currency_chart_data':     currency_chart_data,
        'currency_chart_colors':   currency_chart_colors,
    })


#  Live Status Counts (AJAX) 

@login_required
@supervisor_required
def analytics_status_counts(request):
    from django.http import JsonResponse
    qs = Shipment.objects.all()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    declarant_filter = request.GET.get('declarant', '').strip()
    if date_from:
        qs = qs.filter(submitted_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(submitted_at__date__lte=date_to)
    if declarant_filter:
        qs = qs.filter(declarant__username=declarant_filter)
    total = qs.count()
    counts = {}
    max_count = 0
    for key, label in Shipment.STATUS_CHOICES:
        c = qs.filter(status=key).count()
        counts[key] = {'count': c, 'label': label}
        if c > max_count:
            max_count = c
    return JsonResponse({'counts': counts, 'total': total, 'max_count': max_count})


#  Supervisor Shipment Detail (read-only) 
