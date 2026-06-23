import logging
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from apps.shipments.models import Shipment
from .common import consignee_required

logger = logging.getLogger('r3pcr.consignee')

@login_required
@consignee_required
def download_computation(request, shipment_id):
    """Download ECDT + MCDA results as PDF (default) or Excel (.xlsx).
    Use ?fmt=xlsx for Excel, ?fmt=pdf (or omit) for PDF."""
    shipment    = get_object_or_404(Shipment, id=shipment_id, consignee=request.user)
    computation = getattr(shipment, 'computation', None)
    advisory    = getattr(shipment, 'shipping_advisory', None)
    fmt = request.GET.get('fmt', 'pdf').lower()
    if fmt == 'xlsx':
        return _ecdt_xlsx(request, shipment, computation, advisory)
    return _ecdt_pdf(request, shipment, computation, advisory)


# ── helpers shared by both generators ────────────────────────────────────────

def _info_block(shipment, computation, request):
    """Return a list of (label, value) info rows for the document header."""
    date_str  = computation.computed_at.strftime('%B %d, %Y') if computation else '—'
    exrate    = f'PHP {float(computation.exchange_rate):,.4f}' if computation else '—'
    mode_str  = shipment.get_shipment_type_display() if shipment.shipment_type else '—'
    cname     = request.user.get_full_name() or request.user.username
    decl_name = '—'
    if computation and computation.computed_by:
        decl_name = (
            computation.computed_by.get_full_name()
            or computation.computed_by.username
        )
    return [
        ('HAWB / BOL No.',  shipment.hawb_number),
        ('Consignee',       cname),
        ('Declarant',       decl_name),
        ('Date Computed',   date_str),
        ('Shipment Mode',   mode_str),
        ('Container No.',   shipment.container_number or '—'),
        ('Job Number',     shipment.job_order_reference or '—'),
        ('Exchange Rate',   exrate),
    ]


def _ecdt_fee_rows(computation):
    """Ordered (label, amount, kind) duties/fees rows shared by the PDF + Excel
    ECDT generators. ``kind`` ∈ {'normal', 'cds', 'boc', 'total'} lets each
    format apply its own styling while the sequence/labels live in one place.
    """
    rows = [
        ('Customs Duty (CUD)',                   computation.customs_duty,   'normal'),
        ('Value Added Tax — 12% (VAT)',          computation.vat_amount,     'normal'),
        ('Import Processing Fee (IPF)',          computation.ipf,            'normal'),
        ('Documentary Stamp (CDS)',              130,                        'cds'),
        ('BOC Payable  (CUD + VAT + IPF + CDS)', computation.boc_payable,    'boc'),
        ('Brokerage Fee',                        computation.brokerage_fee,  'normal'),
    ]
    if computation.arrastre:
        rows.append(('Arrastre',              computation.arrastre,     'normal'))
    if computation.wharfage:
        rows.append(('Wharfage',              computation.wharfage,     'normal'))
    if computation.bank_charges:
        rows.append(('Bank Charges',          computation.bank_charges, 'normal'))
    if computation.csf_php:
        rows.append(('Container Service Fee', computation.csf_php,      'normal'))
    rows.append(('TOTAL LANDED COST', computation.total_landed_cost, 'total'))
    return rows


def _ecdt_mode_scores(advisory):
    """MCDA mode rows (label, key, score) in fixed Air/LCL/FCL order, shared by
    both generators (each sorts by score and applies its own styling)."""
    return [
        ('Air Freight',               'air',  advisory.air_score),
        ('LCL (Less Container Load)', 'lcl',  advisory.lcl_score),
        ('FCL (Full Container Load)', 'fcl',  advisory.fcl_score),
    ]


# ── Excel generator (light / navy theme) ─────────────────────────────────────

def _ecdt_xlsx(request, shipment, computation, advisory):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    NAVY    = '1E3A6E'
    WHITE   = 'FFFFFF'
    LGRAY   = 'F0F4F8'
    DGRAY   = '374151'
    MGRAY   = '9CA3AF'
    GREEN_D = '15803D'
    LGREEN  = 'DCFCE7'
    LBLUE   = 'EFF6FF'
    DBLUE   = '1E40AF'

    def xfill(c):
        return PatternFill('solid', fgColor=c)

    def xborder(color='CBD5E1', weight='thin'):
        s = Side(style=weight, color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    al_l = Alignment(horizontal='left',   vertical='center', indent=1)
    al_r = Alignment(horizontal='right',  vertical='center')
    al_c = Alignment(horizontal='center', vertical='center')

    def xcell(ws, row, col, value,
              bold=False, italic=False, color=DGRAY, bg=WHITE,
              align=None, brd=None, size=10, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, italic=italic, color=color, size=size, name='Calibri')
        c.fill = xfill(bg)
        if align:   c.alignment  = align
        if brd:     c.border     = brd
        if num_fmt: c.number_format = num_fmt
        return c

    thin = xborder()
    med  = xborder('94A3B8', 'medium')

    wb = openpyxl.Workbook()

    # ══ Sheet 1 — ECDT ════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'ECDT Summary'
    ws.sheet_view.showGridLines = False
    NC = 8
    for ci, w in enumerate([28, 8, 8, 14, 8, 16, 16, 18], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    r = 1
    ws.merge_cells(f'A{r}:H{r}')
    xcell(ws, r, 1, 'RTripleJ Customs Brokerage',
          bold=True, color=NAVY, bg=WHITE, align=al_c, size=16)
    ws.row_dimensions[r].height = 30
    r += 1
    ws.merge_cells(f'A{r}:H{r}')
    xcell(ws, r, 1, 'ECDT & MCDA Computation Sheet',
          bold=True, color=NAVY, bg=WHITE, align=al_c, size=11)
    ws.row_dimensions[r].height = 18
    r += 1
    # navy rule
    navy_top = Border(top=Side(style='medium', color=NAVY))
    for ci in range(1, NC + 1):
        ws.cell(row=r, column=ci).border = navy_top
    r += 1

    # disclaimer
    ws.merge_cells(f'A{r}:H{r}')
    xcell(ws, r, 1,
          'ESTIMATED COMPUTATION ONLY. FINAL ASSESSMENT WILL BE BASED ON CUSTOMS FINDINGS.',
          bold=True, italic=True, color='92400E', bg='FEF9C3', align=al_c, size=9)
    ws.row_dimensions[r].height = 16
    r += 1

    # info block
    for label, val in _info_block(shipment, computation, request):
        ws.merge_cells(f'A{r}:B{r}')
        xcell(ws, r, 1, label, bold=True, color=MGRAY, bg=LGRAY, align=al_l, brd=thin, size=9)
        ws.merge_cells(f'C{r}:H{r}')
        xcell(ws, r, 3, val,   color=DGRAY, bg=WHITE, align=al_l, brd=thin, size=10)
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1  # spacer

    if computation:
        items = computation.get_items()

        # items table
        currency_code = shipment.invoice_currency or 'USD'
        hdrs = ['DESCRIPTION', 'QTY', 'UNIT', 'HS CODE', 'DUTY %',
                'D/V (PHP)', 'CUD (PHP)', f'EXW ({currency_code})']
        for ci, h in enumerate(hdrs, start=1):
            xcell(ws, r, ci, h, bold=True, color=WHITE, bg=NAVY,
                  align=al_c, brd=thin, size=10)
        ws.row_dimensions[r].height = 20
        r += 1

        for idx, it in enumerate(items):
            row_bg = WHITE if idx % 2 == 0 else LGRAY
            vals = [
                it.get('description', ''),
                it.get('quantity', ''),
                it.get('unit', ''),
                it.get('hs_code', '') or '—',
                float(it.get('duty_rate', 0) or 0),
                float(it.get('dv_php',    0) or 0),
                float(it.get('cud',       0) or 0),
                float(it.get('exw',       0) or 0),
            ]
            for ci, v in enumerate(vals, start=1):
                al  = al_c if ci in (2, 3, 4, 5) else (al_r if ci > 4 else al_l)
                nfm = '0.00' if ci == 5 else ('#,##0.00' if ci in (6, 7, 8) else None)
                xcell(ws, r, ci, v, color=DGRAY, bg=row_bg,
                      align=al, brd=thin, size=10, num_fmt=nfm)
            r += 1

        r += 1  # spacer

        # fee summary header
        ws.merge_cells(f'A{r}:F{r}')
        xcell(ws, r, 1, 'CHARGE',       bold=True, color=WHITE, bg=NAVY,
              align=al_l, brd=thin, size=10)
        ws.merge_cells(f'G{r}:H{r}')
        xcell(ws, r, 7, 'AMOUNT (PHP)', bold=True, color=WHITE, bg=NAVY,
              align=al_r, brd=thin, size=10)
        ws.row_dimensions[r].height = 20
        r += 1

        def xfee(label, amount, bold=False, color=DGRAY, bg=WHITE):
            nonlocal r
            ws.merge_cells(f'A{r}:F{r}')
            xcell(ws, r, 1, label, bold=bold, color=color, bg=bg,
                  align=al_l, brd=thin, size=10)
            ws.merge_cells(f'G{r}:H{r}')
            xcell(ws, r, 7, float(amount or 0), bold=bold, color=color, bg=bg,
                  align=al_r, brd=thin, size=10, num_fmt='#,##0.00')
            r += 1

        for label, amount, kind in _ecdt_fee_rows(computation):
            if kind == 'boc':
                ws.merge_cells(f'A{r}:F{r}')
                xcell(ws, r, 1, label,
                      bold=True, color=DBLUE, bg=LBLUE, align=al_l, brd=thin, size=10)
                ws.merge_cells(f'G{r}:H{r}')
                xcell(ws, r, 7, float(amount or 0),
                      bold=True, color=DBLUE, bg=LBLUE, align=al_r, brd=thin,
                      size=10, num_fmt='#,##0.00')
                ws.row_dimensions[r].height = 18
                r += 1
            elif kind == 'total':
                ws.merge_cells(f'A{r}:F{r}')
                xcell(ws, r, 1, label,
                      bold=True, color=GREEN_D, bg=LGREEN, align=al_l, brd=med, size=11)
                ws.merge_cells(f'G{r}:H{r}')
                xcell(ws, r, 7, float(amount or 0),
                      bold=True, color=GREEN_D, bg=LGREEN, align=al_r, brd=med,
                      size=11, num_fmt='#,##0.00')
                ws.row_dimensions[r].height = 22
                r += 1
            else:
                xfee(label, amount, bg=(LGRAY if kind == 'cds' else WHITE))

    else:
        ws.merge_cells(f'A{r}:H{r}')
        xcell(ws, r, 1, 'No computation on file for this shipment.',
              color=MGRAY, bg=WHITE, align=al_c, size=10)

    # ══ Sheet 2 — MCDA ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet('MCDA Advisory')
    ws2.sheet_view.showGridLines = False
    for ci, w in enumerate([32, 14, 18], start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    r2 = 1
    ws2.merge_cells(f'A{r2}:C{r2}')
    xcell(ws2, r2, 1, 'RTripleJ Customs Brokerage',
          bold=True, color=NAVY, bg=WHITE, align=al_c, size=16)
    ws2.row_dimensions[r2].height = 30
    r2 += 1
    ws2.merge_cells(f'A{r2}:C{r2}')
    xcell(ws2, r2, 1, 'MCDA — Shipping Mode Advisory',
          bold=True, color=NAVY, bg=WHITE, align=al_c, size=11)
    ws2.row_dimensions[r2].height = 18
    r2 += 2

    if advisory:
        ws2.merge_cells(f'A{r2}:C{r2}')
        xcell(ws2, r2, 1,
              f'Declared Mode: '
              f'{shipment.get_shipment_type_display() if shipment.shipment_type else "—"}',
              bold=True, color=NAVY, bg=LGRAY, align=al_l, size=10)
        r2 += 1

        for ci, h in enumerate(['MODE', 'SCORE', 'RESULT'], start=1):
            xcell(ws2, r2, ci, h, bold=True, color=WHITE, bg=NAVY,
                  align=al_c, brd=thin, size=10)
        ws2.row_dimensions[r2].height = 20
        r2 += 1

        for mode_label, key, score in sorted(
            _ecdt_mode_scores(advisory), key=lambda x: (x[2] or 0), reverse=True
        ):
            is_rec = (key == advisory.recommended_type)
            bg     = LGREEN if is_rec else (WHITE if r2 % 2 == 0 else LGRAY)
            tag    = '★ Recommended' if is_rec else ''
            fc     = GREEN_D if is_rec else DGRAY
            xcell(ws2, r2, 1, mode_label,
                  bold=is_rec, color=fc, bg=bg, align=al_l, brd=thin, size=10)
            c2 = ws2.cell(row=r2, column=2, value=float(score or 0))
            c2.font = Font(bold=is_rec, color=fc, size=10, name='Calibri')
            c2.fill = xfill(bg)
            c2.number_format = '0.0000'
            c2.alignment = al_c
            c2.border = thin
            xcell(ws2, r2, 3, tag,
                  bold=is_rec, color=fc, bg=bg, align=al_c, brd=thin, size=10)
            r2 += 1

        r2 += 1
        if advisory.declarant_recommendation:
            ws2.merge_cells(f'A{r2}:C{r2}')
            xcell(ws2, r2, 1,
                  f"Declarant's Recommendation: "
                  f"{advisory.declarant_recommendation.upper()}",
                  bold=True, color=DBLUE, bg=LBLUE, align=al_l, size=11)
            r2 += 1
        if advisory.declarant_note:
            ws2.merge_cells(f'A{r2}:C{r2}')
            cn = ws2.cell(row=r2, column=1,
                          value=f'Note: "{advisory.declarant_note}"')
            cn.font = Font(italic=True, color=MGRAY, size=10, name='Calibri')
            cn.fill = xfill(LGRAY)
            cn.alignment = Alignment(
                horizontal='left', vertical='center', wrap_text=True)
            ws2.row_dimensions[r2].height = 36
    else:
        ws2.merge_cells('A5:C5')
        xcell(ws2, 5, 1, 'No MCDA advisory on file.',
              color=MGRAY, bg=WHITE, align=al_c, size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"R3PCR_{shipment.hawb_number}_ECDT_MCDA.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── PDF generator (light / navy theme) ───────────────────────────────────────

def _ecdt_pdf(request, shipment, computation, advisory):
    import io
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )

    NAVY   = colors.HexColor('#1E3A6E')
    WHITE  = colors.white
    LGRAY  = colors.HexColor('#F0F4F8')
    DGRAY  = colors.HexColor('#374151')
    MGRAY  = colors.HexColor('#9CA3AF')
    GREEN  = colors.HexColor('#15803D')
    LGREEN = colors.HexColor('#DCFCE7')
    LBLUE  = colors.HexColor('#EFF6FF')
    DBLUE  = colors.HexColor('#1E40AF')
    BORDER = colors.HexColor('#CBD5E1')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1*cm, bottomMargin=1.5*cm,
    )
    W = doc.width

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    p_company  = ps('co', fontSize=18, fontName='Helvetica-Bold',
                    textColor=NAVY, alignment=TA_CENTER, spaceAfter=2)
    p_subtitle = ps('su', fontSize=11, fontName='Helvetica-Bold',
                    textColor=NAVY, alignment=TA_CENTER, spaceAfter=6)
    p_section  = ps('se', fontSize=9,  fontName='Helvetica-Bold',
                    textColor=WHITE, backColor=NAVY, alignment=TA_LEFT,
                    leftIndent=4, leading=16, spaceAfter=2, spaceBefore=8)
    p_lbl      = ps('lb', fontSize=9, fontName='Helvetica-Bold',  textColor=MGRAY)
    p_val      = ps('vl', fontSize=9, fontName='Helvetica',       textColor=DGRAY)
    p_note     = ps('nt', fontSize=7, fontName='Helvetica-Oblique',
                    textColor=MGRAY, alignment=TA_CENTER)

    def hdr8(align=TA_LEFT):
        return ps(f'h8{align}', fontSize=8, fontName='Helvetica-Bold',
                  textColor=WHITE, alignment=align)

    def body8(align=TA_LEFT, bold=False, color=None):
        fn = 'Helvetica-Bold' if bold else 'Helvetica'
        return ps(f'bd8{align}{bold}', fontSize=8, fontName=fn,
                  textColor=color or DGRAY, alignment=align)

    def php(v):
        return f'₱{float(v or 0):,.2f}'

    story = []

    # header
    story.append(Paragraph('RTripleJ Customs Brokerage', p_company))
    story.append(Paragraph('ECDT &amp; MCDA Computation Sheet', p_subtitle))
    story.append(HRFlowable(width='100%', thickness=2, color=NAVY,
                             spaceAfter=6, spaceBefore=0))

    # disclaimer banner
    AMBER_BG = colors.HexColor('#FEF9C3')
    AMBER_TX = colors.HexColor('#92400E')
    disclaimer_tbl = Table(
        [[Paragraph(
            '<b><i>⚠ ESTIMATED COMPUTATION ONLY. '
            'FINAL ASSESSMENT WILL BE BASED ON CUSTOMS FINDINGS.</i></b>',
            ps('disc', fontSize=8, fontName='Helvetica-BoldOblique',
               textColor=AMBER_TX, alignment=TA_CENTER),
        )]],
        colWidths=[W],
    )
    disclaimer_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), AMBER_BG),
        ('BOX',           (0, 0), (-1, -1), 0.75, colors.HexColor('#FCD34D')),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
    ]))
    story.append(disclaimer_tbl)
    story.append(Spacer(1, 8))

    # info table
    info_rows = [
        [Paragraph(f'<b>{k}</b>', p_lbl), Paragraph(v, p_val)]
        for k, v in _info_block(shipment, computation, request)
    ]
    info_tbl = Table(info_rows, colWidths=[W * 0.28, W * 0.72])
    info_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (0, -1), LGRAY),
        ('BACKGROUND',    (1, 0), (1, -1), WHITE),
        ('GRID',          (0, 0), (-1, -1), 0.5, BORDER),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 10))

    if computation:
        items = computation.get_items()

        # ── line items ────────────────────────────────────────────────────────
        story.append(Paragraph('LINE ITEMS', p_section))
        story.append(Spacer(1, 1))

        cw = [W*0.27, W*0.11, W*0.10, W*0.08, W*0.14, W*0.15, W*0.15]
        item_rows = [[
            Paragraph('<b>DESCRIPTION</b>', hdr8(TA_LEFT)),
            Paragraph('<b>QTY / UNIT</b>',  hdr8(TA_CENTER)),
            Paragraph('<b>HS CODE</b>',      hdr8(TA_CENTER)),
            Paragraph('<b>DUTY %</b>',       hdr8(TA_RIGHT)),
            Paragraph('<b>D/V (PHP)</b>',    hdr8(TA_RIGHT)),
            Paragraph('<b>CUD (PHP)</b>',    hdr8(TA_RIGHT)),
            Paragraph(f'<b>EXW ({shipment.invoice_currency or "USD"})</b>', hdr8(TA_RIGHT)),
        ]]
        for it in items:
            qty_unit = str(it.get('quantity', ''))
            if it.get('unit'):
                qty_unit += f" {it.get('unit')}"
            item_rows.append([
                Paragraph(it.get('description', ''),              body8(TA_LEFT)),
                Paragraph(qty_unit.strip(),                        body8(TA_CENTER)),
                Paragraph(it.get('hs_code') or '—',               body8(TA_CENTER)),
                Paragraph(f"{float(it.get('duty_rate',0) or 0):.2f}%", body8(TA_RIGHT)),
                Paragraph(f"{float(it.get('dv_php', 0) or 0):,.2f}",   body8(TA_RIGHT)),
                Paragraph(f"{float(it.get('cud',    0) or 0):,.2f}",   body8(TA_RIGHT)),
                Paragraph(f"{float(it.get('exw',    0) or 0):,.2f}",   body8(TA_RIGHT)),
            ])

        item_tbl = Table(item_rows, colWidths=cw, repeatRows=1)
        item_style = [
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('GRID',          (0, 0), (-1, -1), 0.5, BORDER),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]
        for i in range(1, len(item_rows)):
            item_style.append((
                'BACKGROUND', (0, i), (-1, i),
                WHITE if i % 2 == 1 else LGRAY,
            ))
        item_tbl.setStyle(TableStyle(item_style))
        story.append(item_tbl)
        story.append(Spacer(1, 10))

        # ── fee summary ───────────────────────────────────────────────────────
        story.append(Paragraph('DUTIES &amp; FEES SUMMARY', p_section))
        story.append(Spacer(1, 1))

        fee_rows = [[
            Paragraph('<b>CHARGE</b>',       hdr8(TA_LEFT)),
            Paragraph('<b>AMOUNT (PHP)</b>', hdr8(TA_RIGHT)),
        ]]

        def pfee(label, amount, bold=False, lc=None, ac=None):
            fn  = 'Helvetica-Bold' if bold else 'Helvetica'
            fee_rows.append([
                Paragraph(label, ps(f'fl{len(fee_rows)}',
                          fontSize=9, fontName=fn, textColor=lc or DGRAY)),
                Paragraph(php(amount), ps(f'fa{len(fee_rows)}',
                          fontSize=9, fontName=fn,
                          textColor=ac or DGRAY, alignment=TA_RIGHT)),
            ])

        boc_idx = tlc_idx = None
        for label, amount, kind in _ecdt_fee_rows(computation):
            if kind == 'boc':
                boc_idx = len(fee_rows)
                pfee(label, amount, bold=True, lc=DBLUE, ac=DBLUE)
            elif kind == 'total':
                tlc_idx = len(fee_rows)
                fee_rows.append([
                    Paragraph('<b>TOTAL LANDED COST</b>',
                              ps('tll', fontSize=10, fontName='Helvetica-Bold', textColor=GREEN)),
                    Paragraph(php(amount),
                              ps('tlr', fontSize=10, fontName='Helvetica-Bold',
                                 textColor=GREEN, alignment=TA_RIGHT)),
                ])
            else:
                pfee(label, amount)

        fee_tbl = Table(fee_rows, colWidths=[W * 0.65, W * 0.35])
        fee_style = [
            ('BACKGROUND',    (0, 0),       (-1, 0),       NAVY),
            ('GRID',          (0, 0),       (-1, -1),      0.5, BORDER),
            ('VALIGN',        (0, 0),       (-1, -1),      'MIDDLE'),
            ('TOPPADDING',    (0, 0),       (-1, -1),      4),
            ('BOTTOMPADDING', (0, 0),       (-1, -1),      4),
            ('LEFTPADDING',   (0, 0),       (-1, -1),      6),
            ('RIGHTPADDING',  (0, 0),       (-1, -1),      6),
            ('BACKGROUND',    (0, boc_idx), (-1, boc_idx), LBLUE),
            ('LINEABOVE',     (0, boc_idx), (-1, boc_idx), 1.5, DBLUE),
            ('LINEBELOW',     (0, boc_idx), (-1, boc_idx), 1.5, DBLUE),
            ('BACKGROUND',    (0, tlc_idx), (-1, tlc_idx), LGREEN),
            ('LINEABOVE',     (0, tlc_idx), (-1, tlc_idx), 1.5, GREEN),
            ('LINEBELOW',     (0, tlc_idx), (-1, tlc_idx), 1.5, GREEN),
        ]
        for i in range(1, len(fee_rows)):
            if i not in (boc_idx, tlc_idx):
                fee_style.append((
                    'BACKGROUND', (0, i), (-1, i),
                    WHITE if i % 2 == 1 else LGRAY,
                ))
        fee_tbl.setStyle(TableStyle(fee_style))
        story.append(fee_tbl)
        story.append(Spacer(1, 12))

    # ── MCDA ─────────────────────────────────────────────────────────────────
    if advisory:
        story.append(Paragraph('MCDA — SHIPPING MODE ADVISORY', p_section))
        story.append(Spacer(1, 1))

        mode_scores = _ecdt_mode_scores(advisory)
        adv_rows = [[
            Paragraph('<b>MODE</b>',   hdr8(TA_LEFT)),
            Paragraph('<b>SCORE</b>',  hdr8(TA_CENTER)),
            Paragraph('<b>RESULT</b>', hdr8(TA_CENTER)),
        ]]
        adv_bgs = []
        for i, (label, key, score) in enumerate(
            sorted(mode_scores, key=lambda x: (x[2] or 0), reverse=True), start=1
        ):
            is_rec = (key == advisory.recommended_type)
            mc = GREEN if is_rec else DGRAY
            mf = 'Helvetica-Bold' if is_rec else 'Helvetica'
            tag = '★ Recommended' if is_rec else ''
            adv_rows.append([
                Paragraph(label, ps(f'am{i}', fontSize=8, fontName=mf, textColor=mc)),
                Paragraph(f'{float(score or 0):.4f}',
                          ps(f'as{i}', fontSize=8, fontName=mf,
                             textColor=mc, alignment=TA_CENTER)),
                Paragraph(tag,
                          ps(f'at{i}', fontSize=8, fontName=mf,
                             textColor=mc, alignment=TA_CENTER)),
            ])
            adv_bgs.append((i, LGREEN if is_rec else (WHITE if i % 2 == 1 else LGRAY)))

        adv_tbl = Table(adv_rows, colWidths=[W*0.50, W*0.20, W*0.30], repeatRows=1)
        adv_sty = [
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('GRID',          (0, 0), (-1, -1), 0.5, BORDER),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ]
        for row_i, bg in adv_bgs:
            adv_sty.append(('BACKGROUND', (0, row_i), (-1, row_i), bg))
        adv_tbl.setStyle(TableStyle(adv_sty))
        story.append(adv_tbl)
        story.append(Spacer(1, 6))

        if advisory.declarant_recommendation:
            story.append(Paragraph(
                f"<b>Declarant&#8217;s Recommendation:</b> "
                f"{advisory.declarant_recommendation.upper()}",
                ps('dr', fontSize=9, fontName='Helvetica-Bold', textColor=DBLUE),
            ))
        if advisory.declarant_note:
            story.append(Spacer(1, 2))
            story.append(Paragraph(
                f'<i>Note: &ldquo;{advisory.declarant_note}&rdquo;</i>',
                ps('dn', fontSize=8, fontName='Helvetica-Oblique', textColor=MGRAY),
            ))
        story.append(Spacer(1, 10))

    # footer
    story.append(HRFlowable(width='100%', thickness=1, color=NAVY,
                             spaceBefore=4, spaceAfter=4))
    story.append(Paragraph(
        f'Generated by R3-PCR · RTripleJ Customs Brokerage '
        f'· {shipment.hawb_number}',
        p_note,
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"R3PCR_{shipment.hawb_number}_ECDT_MCDA.pdf"
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Chart Data (AJAX) ───────────────────────────────────────────────────────
