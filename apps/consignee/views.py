from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from apps.shipments.models import Shipment, ShipmentDocument, StatusLog
from apps.shipments.status_progress import build_status_progress
from apps.notifications.utils import create_notification, notify_incoming_shipment
from .models import Feedback


# ─── Auto-generate HAWB ───────────────────────────────────────────────────────

def generate_hawb():
    year   = timezone.now().year
    prefix = f'R3PCR-{year}-'
    last   = (
        Shipment.objects
        .filter(hawb_number__startswith=prefix)
        .order_by('hawb_number')
        .last()
    )
    if last:
        try:
            seq      = int(last.hawb_number[len(prefix):])
            next_seq = seq + 1
        except (ValueError, IndexError):
            next_seq = 1
    else:
        next_seq = 1

    hawb = f'{prefix}{str(next_seq).zfill(6)}'
    while Shipment.objects.filter(hawb_number=hawb).exists():
        next_seq += 1
        hawb = f'{prefix}{str(next_seq).zfill(6)}'
    return hawb


# ─── Dashboard ────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    from django.db.models import Count
    shipments = Shipment.objects.filter(consignee=request.user)
    total = shipments.count()

    status_counts = {
        'incoming':     shipments.filter(status='incoming').count(),
        'arrived':      shipments.filter(status='arrived').count(),
        'computed':     shipments.filter(status='computed').count(),
        'approved':     shipments.filter(status='approved').count(),
        'rejected':     shipments.filter(status='rejected').count(),
        'for_revision': shipments.filter(status='for_revision').count(),
        'lodgement':    shipments.filter(status='lodgement').count(),
        'ongoing':      shipments.filter(status='ongoing').count(),
        'assessed':     shipments.filter(status='assessed').count(),
        'paid':         shipments.filter(status='paid').count(),
        'released':     shipments.filter(status='released').count(),
        'billed':       shipments.filter(status='billed').count(),
    }

    import_breakdown = list(
        shipments.values('import_type')
                 .annotate(count=Count('id'))
                 .order_by('-count')
    )
    import_labels = dict(Shipment.IMPORT_TYPE_CHOICES)
    for item in import_breakdown:
        item['label'] = import_labels.get(item['import_type'], item['import_type'])

    mode_breakdown = list(
        shipments.values('shipment_type')
                 .annotate(count=Count('id'))
                 .order_by('-count')
    )
    mode_labels = dict(Shipment.SHIPMENT_TYPE_CHOICES)
    for item in mode_breakdown:
        item['label'] = mode_labels.get(item['shipment_type'], item['shipment_type'] or 'Not specified')

    urgency_breakdown = list(
        shipments.values('urgency')
                 .annotate(count=Count('id'))
                 .order_by('-count')
    )
    urgency_labels = dict(Shipment.URGENCY_CHOICES)
    for item in urgency_breakdown:
        item['label'] = urgency_labels.get(item['urgency'], item['urgency'] or 'Unknown')

    context = {
        'total': total,
        **status_counts,
        'import_breakdown':   import_breakdown,
        'mode_breakdown':     mode_breakdown,
        'urgency_breakdown':  urgency_breakdown,
        'recent_shipments':   shipments.order_by('-submitted_at'),
    }

    from apps.supervisor.models import Announcement
    recent_announcements = Announcement.objects.filter(is_active=True).order_by('-created_at')[:3]
    context['recent_announcements'] = recent_announcements

    return render(request, 'consignee/dashboard.html', context)


# ─── Submit Shipment ──────────────────────────────────────────────────────────

@login_required
def submit_shipment(request):
    if request.method == 'POST':
        import_type   = request.POST.get('import_type')
        urgency       = request.POST.get('urgency')
        shipment_type = request.POST.get('shipment_type', '').strip()
        description   = request.POST.get('description', '').strip()

        hawb_number = generate_hawb()

        shipment = Shipment.objects.create(
            hawb_number=hawb_number,
            consignee=request.user,
            import_type=import_type,
            urgency=urgency,
            shipment_type=shipment_type or None,
            description=description,
            status='incoming',
        )

        for doc_type in ['invoice', 'packing_list', 'airway_bill']:
            file = request.FILES.get(doc_type)
            if file:
                ShipmentDocument.objects.create(
                    shipment=shipment,
                    document_type=doc_type,
                    file=file,
                )

        # Other supporting documents (multiple)
        for file in request.FILES.getlist('other_docs'):
            ShipmentDocument.objects.create(
                shipment=shipment,
                document_type='other',
                file=file,
            )

        for declarant in []:
            create_notification(
                recipient=declarant,
                shipment=shipment,
                notification_type='submission',
                title=f'New Shipment Ready to Claim — {hawb_number}',
                message=(
                    f'A new shipment ({hawb_number}) is in the incoming queue and '
                    f'available for any declarant to claim and process.'
                ),
            )
        notify_incoming_shipment(shipment)

        messages.success(
            request,
            f'Shipment submitted! Your Shipment Reference No. is '
            f'<strong>{hawb_number}</strong>.'
        )
        return redirect('consignee:my_submissions')

    return render(request, 'consignee/submit.html')


# ─── My Submissions ───────────────────────────────────────────────────────────

@login_required
def my_submissions(request):
    shipments = Shipment.objects.filter(consignee=request.user).order_by('-submitted_at')

    status_filter = request.GET.get('status', '').strip()
    q             = request.GET.get('q', '').strip()
    date_from     = request.GET.get('date_from', '').strip()
    date_to       = request.GET.get('date_to', '').strip()

    if status_filter:
        shipments = shipments.filter(status=status_filter)
    if q:
        shipments = shipments.filter(hawb_number__icontains=q)
    if date_from:
        shipments = shipments.filter(submitted_at__date__gte=date_from)
    if date_to:
        shipments = shipments.filter(submitted_at__date__lte=date_to)

    now            = timezone.now()
    shipments_list = list(shipments)
    for s in shipments_list:
        age_seconds  = (now - s.submitted_at).total_seconds()
        s.can_cancel     = s.status == 'incoming' and age_seconds <= 3600
        s.cancel_expired = s.status == 'incoming' and age_seconds > 3600

    return render(request, 'consignee/my_submissions.html', {
        'shipments':     shipments_list,
        'status_filter': status_filter,
        'q':             q,
        'date_from':     date_from,
        'date_to':       date_to,
    })


# ─── Shipment Detail ──────────────────────────────────────────────────────────

@login_required
def shipment_detail(request, shipment_id):
    """Consignee-facing detail page: status, advisory results, computation summary."""
    shipment    = get_object_or_404(Shipment, id=shipment_id, consignee=request.user)
    advisory    = getattr(shipment, 'shipping_advisory', None)
    computation = getattr(shipment, 'computation', None)
    status_logs = shipment.status_logs.order_by('-changed_at')

    # Rebuild full WMCDA data from saved advisory
    explanation       = None
    wmcda_scores      = None
    wmcda_breakdown   = None
    declared_score    = None
    declared_breakdown = None
    declared_rating   = None

    if advisory:
        try:
            from apps.computation.views import compute_wmcda
            wmcda_scores, _, wmcda_breakdown, explanation = compute_wmcda(
                float(advisory.gross_weight),
                float(advisory.cargo_volume),
                float(advisory.declared_value),
                advisory.urgency_level,
                float(advisory.distance_km),
            )
            if wmcda_scores and shipment.shipment_type:
                declared_score = wmcda_scores.get(shipment.shipment_type)
                if wmcda_breakdown:
                    declared_breakdown = wmcda_breakdown.get(shipment.shipment_type)
                if declared_score is not None:
                    if declared_score >= 0.80:
                        declared_rating = 'Excellent'
                    elif declared_score >= 0.65:
                        declared_rating = 'Good'
                    elif declared_score >= 0.50:
                        declared_rating = 'Fair'
                    else:
                        declared_rating = 'Poor'
        except Exception:
            pass

    sad_document = shipment.documents.filter(document_type='sad').first()

    # Current step sublabel for the status description box
    from apps.shipments.status_progress import CONSIGNEE_STATUS_SUBLABELS
    current_sublabel = CONSIGNEE_STATUS_SUBLABELS.get(shipment.status, '')

    context = {
        'shipment':          shipment,
        'advisory':          advisory,
        'computation':       computation,
        'status_logs':       status_logs,
        'explanation':       explanation,
        'wmcda_scores':      wmcda_scores,
        'wmcda_breakdown':   wmcda_breakdown,
        'declared_score':    declared_score,
        'declared_breakdown': declared_breakdown,
        'declared_rating':   declared_rating,
        'status_steps':      build_status_progress(shipment.status, 'consignee'),
        'sad_document':      sad_document,
        'current_sublabel':  current_sublabel,
    }
    return render(request, 'consignee/shipment_detail.html', context)


# ─── Upload Payment Receipt ──────────────────────────────────────────────────

@login_required
def upload_receipt(request, shipment_id):
    shipment = get_object_or_404(Shipment, id=shipment_id, consignee=request.user)

    if request.method == 'POST':
        file = request.FILES.get('payment_receipt')
        if not file:
            messages.error(request, 'Please select a file to upload.')
        elif shipment.status != 'paid':
            messages.error(request, 'Payment receipt can only be uploaded when shipment is marked Paid.')
        else:
            shipment.payment_receipt = file
            shipment.payment_receipt_uploaded_at = timezone.now()
            shipment.save()

            # Audit trail — record receipt upload in status log
            StatusLog.objects.create(
                shipment=shipment,
                changed_by=request.user,
                old_status=shipment.status,
                new_status=shipment.status,
                notes='Payment receipt uploaded by consignee.',
            )

            # Notify declarant
            if shipment.declarant:
                create_notification(
                    recipient=shipment.declarant,
                    shipment=shipment,
                    notification_type='status_update',
                    title=f'Payment Receipt Uploaded — {shipment.hawb_number}',
                    message=f'{request.user.get_full_name() or request.user.username} uploaded a payment receipt.',
                )
            messages.success(request, 'Payment receipt uploaded successfully.')

    return redirect('consignee:shipment_detail', shipment_id=shipment_id)


# ─── Feedback ─────────────────────────────────────────────────────────────────

@login_required
def submit_feedback(request, shipment_id):
    shipment = get_object_or_404(Shipment, id=shipment_id, consignee=request.user)

    # Only allow feedback on completed shipments
    if shipment.status not in ('approved', 'rejected'):
        messages.error(request, 'Feedback can only be submitted for completed shipments.')
        return redirect('consignee:shipment_detail', shipment_id=shipment_id)

    # One feedback per shipment
    if hasattr(shipment, 'feedback'):
        messages.info(request, 'You have already submitted feedback for this shipment.')
        return redirect('consignee:shipment_detail', shipment_id=shipment_id)

    if request.method == 'POST':
        rating  = request.POST.get('rating', '').strip()
        comment = request.POST.get('comment', '').strip()

        if not rating or not comment:
            messages.error(request, 'Please provide a rating and a comment.')
            return render(request, 'consignee/feedback.html', {'shipment': shipment})

        Feedback.objects.create(
            consignee=request.user,
            shipment=shipment,
            rating=int(rating),
            comment=comment,
        )
        messages.success(request, 'Thank you for your feedback! It will appear on our site once reviewed.')
        return redirect('consignee:shipment_detail', shipment_id=shipment_id)

    return render(request, 'consignee/feedback.html', {'shipment': shipment})


# ─── Approve Computation ─────────────────────────────────────────────────────

@login_required
def approve_computation(request, shipment_id):
    """Consignee approves the ECDT+WMCDA computation, advancing status to approved."""
    shipment = get_object_or_404(Shipment, id=shipment_id, consignee=request.user)

    if request.method == 'POST':
        if shipment.status != 'computed':
            messages.error(request, 'This shipment is not awaiting your approval.')
        else:
            old_status = shipment.status
            shipment.status = 'approved'
            shipment.save()
            StatusLog.objects.create(
                shipment=shipment,
                changed_by=request.user,
                old_status=old_status,
                new_status='approved',
                notes='Consignee approved the computation.',
            )
            messages.success(request, 'Computation approved. Your shipment will proceed to customs lodgement.')

    return redirect('consignee:shipment_detail', shipment_id=shipment_id)


# ─── Download Computation Results ────────────────────────────────────────────

@login_required
def download_computation(request, shipment_id):
    """Download ECDT + WMCDA results as PDF (default) or Excel (.xlsx).
    Use ?fmt=xlsx for Excel, ?fmt=pdf (or omit) for PDF."""
    shipment    = get_object_or_404(Shipment, id=shipment_id, consignee=request.user)
    computation = getattr(shipment, 'computation', None)
    advisory    = getattr(shipment, 'shipping_advisory', None)
    fmt = request.GET.get('fmt', 'pdf').lower()
    if fmt == 'xlsx':
        return _ecdt_xlsx(request, shipment, computation, advisory)
    return _ecdt_pdf(request, shipment, computation, advisory)


# ── helpers shared by both generators ────────────────────────────────────────

def _info_block(shipment, computation, request):
    """Return a list of (label, value) info rows for the document header."""
    date_str  = computation.computed_at.strftime('%B %d, %Y') if computation else '—'
    exrate    = f'PHP {float(computation.exchange_rate):,.4f}' if computation else '—'
    mode_str  = shipment.get_shipment_type_display() if shipment.shipment_type else '—'
    cname     = request.user.get_full_name() or request.user.username
    decl_name = '—'
    if computation and computation.computed_by:
        decl_name = (
            computation.computed_by.get_full_name()
            or computation.computed_by.username
        )
    return [
        ('HAWB / BOL No.',  shipment.hawb_number),
        ('Consignee',       cname),
        ('Declarant',       decl_name),
        ('Date Computed',   date_str),
        ('Shipment Mode',   mode_str),
        ('Exchange Rate',   exrate),
    ]


# ── Excel generator (light / navy theme) ─────────────────────────────────────

def _ecdt_xlsx(request, shipment, computation, advisory):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    NAVY    = '1E3A6E'
    WHITE   = 'FFFFFF'
    LGRAY   = 'F0F4F8'
    DGRAY   = '374151'
    MGRAY   = '9CA3AF'
    GREEN_D = '15803D'
    LGREEN  = 'DCFCE7'
    LBLUE   = 'EFF6FF'
    DBLUE   = '1E40AF'

    def xfill(c):
        return PatternFill('solid', fgColor=c)

    def xborder(color='CBD5E1', weight='thin'):
        s = Side(style=weight, color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    al_l = Alignment(horizontal='left',   vertical='center', indent=1)
    al_r = Alignment(horizontal='right',  vertical='center')
    al_c = Alignment(horizontal='center', vertical='center')

    def xcell(ws, row, col, value,
              bold=False, italic=False, color=DGRAY, bg=WHITE,
              align=None, brd=None, size=10, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, italic=italic, color=color, size=size, name='Calibri')
        c.fill = xfill(bg)
        if align:   c.alignment  = align
        if brd:     c.border     = brd
        if num_fmt: c.number_format = num_fmt
        return c

    thin = xborder()
    med  = xborder('94A3B8', 'medium')

    wb = openpyxl.Workbook()

    # ══ Sheet 1 — ECDT ════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'ECDT Summary'
    ws.sheet_view.showGridLines = False
    NC = 8
    for ci, w in enumerate([28, 8, 8, 14, 8, 16, 16, 18], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    r = 1
    ws.merge_cells(f'A{r}:H{r}')
    xcell(ws, r, 1, 'RTripleJ Customs Brokerage',
          bold=True, color=NAVY, bg=WHITE, align=al_c, size=16)
    ws.row_dimensions[r].height = 30
    r += 1
    ws.merge_cells(f'A{r}:H{r}')
    xcell(ws, r, 1, 'ECDT & WMCDA Computation Sheet',
          bold=True, color=NAVY, bg=WHITE, align=al_c, size=11)
    ws.row_dimensions[r].height = 18
    r += 1
    # navy rule
    navy_top = Border(top=Side(style='medium', color=NAVY))
    for ci in range(1, NC + 1):
        ws.cell(row=r, column=ci).border = navy_top
    r += 1

    # disclaimer
    ws.merge_cells(f'A{r}:H{r}')
    xcell(ws, r, 1,
          'ESTIMATED COMPUTATION ONLY. FINAL ASSESSMENT WILL BE BASED ON CUSTOMS FINDINGS.',
          bold=True, italic=True, color='92400E', bg='FEF9C3', align=al_c, size=9)
    ws.row_dimensions[r].height = 16
    r += 1

    # info block
    for label, val in _info_block(shipment, computation, request):
        ws.merge_cells(f'A{r}:B{r}')
        xcell(ws, r, 1, label, bold=True, color=MGRAY, bg=LGRAY, align=al_l, brd=thin, size=9)
        ws.merge_cells(f'C{r}:H{r}')
        xcell(ws, r, 3, val,   color=DGRAY, bg=WHITE, align=al_l, brd=thin, size=10)
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1  # spacer

    if computation:
        items = computation.get_items()

        # items table
        hdrs = ['DESCRIPTION', 'QTY', 'UNIT', 'HS CODE', 'DUTY %',
                'D/V (PHP)', 'CUD (PHP)', 'EXW (USD)']
        for ci, h in enumerate(hdrs, start=1):
            xcell(ws, r, ci, h, bold=True, color=WHITE, bg=NAVY,
                  align=al_c, brd=thin, size=10)
        ws.row_dimensions[r].height = 20
        r += 1

        for idx, it in enumerate(items):
            row_bg = WHITE if idx % 2 == 0 else LGRAY
            vals = [
                it.get('description', ''),
                it.get('quantity', ''),
                it.get('unit', ''),
                it.get('hs_code', '') or '—',
                float(it.get('duty_rate', 0) or 0),
                float(it.get('dv_php',    0) or 0),
                float(it.get('cud',       0) or 0),
                float(it.get('exw',       0) or 0),
            ]
            for ci, v in enumerate(vals, start=1):
                al  = al_c if ci in (2, 3, 4, 5) else (al_r if ci > 4 else al_l)
                nfm = '0.00' if ci == 5 else ('#,##0.00' if ci in (6, 7, 8) else None)
                xcell(ws, r, ci, v, color=DGRAY, bg=row_bg,
                      align=al, brd=thin, size=10, num_fmt=nfm)
            r += 1

        r += 1  # spacer

        # fee summary header
        ws.merge_cells(f'A{r}:F{r}')
        xcell(ws, r, 1, 'CHARGE',       bold=True, color=WHITE, bg=NAVY,
              align=al_l, brd=thin, size=10)
        ws.merge_cells(f'G{r}:H{r}')
        xcell(ws, r, 7, 'AMOUNT (PHP)', bold=True, color=WHITE, bg=NAVY,
              align=al_r, brd=thin, size=10)
        ws.row_dimensions[r].height = 20
        r += 1

        def xfee(label, amount, bold=False, color=DGRAY, bg=WHITE):
            nonlocal r
            ws.merge_cells(f'A{r}:F{r}')
            xcell(ws, r, 1, label, bold=bold, color=color, bg=bg,
                  align=al_l, brd=thin, size=10)
            ws.merge_cells(f'G{r}:H{r}')
            xcell(ws, r, 7, float(amount or 0), bold=bold, color=color, bg=bg,
                  align=al_r, brd=thin, size=10, num_fmt='#,##0.00')
            r += 1

        xfee('Customs Duty (CUD)',          computation.customs_duty)
        xfee('Value Added Tax — 12% (VAT)', computation.vat_amount)
        xfee('Import Processing Fee (IPF)', computation.ipf)
        xfee('Documentary Stamp (CDS)',     130, bg=LGRAY)

        # BOC Payable
        ws.merge_cells(f'A{r}:F{r}')
        xcell(ws, r, 1, 'BOC Payable  (CUD + VAT + IPF + CDS)',
              bold=True, color=DBLUE, bg=LBLUE, align=al_l, brd=thin, size=10)
        ws.merge_cells(f'G{r}:H{r}')
        xcell(ws, r, 7, float(computation.boc_payable or 0),
              bold=True, color=DBLUE, bg=LBLUE, align=al_r, brd=thin,
              size=10, num_fmt='#,##0.00')
        ws.row_dimensions[r].height = 18
        r += 1

        xfee('Brokerage Fee', computation.brokerage_fee)
        if computation.arrastre:      xfee('Arrastre',              computation.arrastre)
        if computation.wharfage:      xfee('Wharfage',              computation.wharfage)
        if computation.bank_charges:  xfee('Bank Charges',          computation.bank_charges)
        if computation.csf_php:       xfee('Container Service Fee', computation.csf_php)

        # Total Landed Cost
        ws.merge_cells(f'A{r}:F{r}')
        xcell(ws, r, 1, 'TOTAL LANDED COST',
              bold=True, color=GREEN_D, bg=LGREEN, align=al_l, brd=med, size=11)
        ws.merge_cells(f'G{r}:H{r}')
        xcell(ws, r, 7, float(computation.total_landed_cost or 0),
              bold=True, color=GREEN_D, bg=LGREEN, align=al_r, brd=med,
              size=11, num_fmt='#,##0.00')
        ws.row_dimensions[r].height = 22
        r += 1

    else:
        ws.merge_cells(f'A{r}:H{r}')
        xcell(ws, r, 1, 'No computation on file for this shipment.',
              color=MGRAY, bg=WHITE, align=al_c, size=10)

    # ══ Sheet 2 — WMCDA ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet('WMCDA Advisory')
    ws2.sheet_view.showGridLines = False
    for ci, w in enumerate([32, 14, 18], start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    r2 = 1
    ws2.merge_cells(f'A{r2}:C{r2}')
    xcell(ws2, r2, 1, 'RTripleJ Customs Brokerage',
          bold=True, color=NAVY, bg=WHITE, align=al_c, size=16)
    ws2.row_dimensions[r2].height = 30
    r2 += 1
    ws2.merge_cells(f'A{r2}:C{r2}')
    xcell(ws2, r2, 1, 'WMCDA — Shipping Mode Advisory',
          bold=True, color=NAVY, bg=WHITE, align=al_c, size=11)
    ws2.row_dimensions[r2].height = 18
    r2 += 2

    if advisory:
        ws2.merge_cells(f'A{r2}:C{r2}')
        xcell(ws2, r2, 1,
              f'Declared Mode: '
              f'{shipment.get_shipment_type_display() if shipment.shipment_type else "—"}',
              bold=True, color=NAVY, bg=LGRAY, align=al_l, size=10)
        r2 += 1

        for ci, h in enumerate(['MODE', 'SCORE', 'RESULT'], start=1):
            xcell(ws2, r2, ci, h, bold=True, color=WHITE, bg=NAVY,
                  align=al_c, brd=thin, size=10)
        ws2.row_dimensions[r2].height = 20
        r2 += 1

        mode_scores = [
            ('Air Freight',               'air',  advisory.air_score),
            ('LCL (Less Container Load)', 'lcl',  advisory.lcl_score),
            ('FCL (Full Container Load)', 'fcl',  advisory.fcl_score),
            ('Land Freight',              'land', advisory.land_score),
        ]
        for mode_label, key, score in sorted(
            mode_scores, key=lambda x: (x[2] or 0), reverse=True
        ):
            is_rec = (key == advisory.recommended_type)
            bg     = LGREEN if is_rec else (WHITE if r2 % 2 == 0 else LGRAY)
            tag    = '★ Recommended' if is_rec else ''
            fc     = GREEN_D if is_rec else DGRAY
            xcell(ws2, r2, 1, mode_label,
                  bold=is_rec, color=fc, bg=bg, align=al_l, brd=thin, size=10)
            c2 = ws2.cell(row=r2, column=2, value=float(score or 0))
            c2.font = Font(bold=is_rec, color=fc, size=10, name='Calibri')
            c2.fill = xfill(bg)
            c2.number_format = '0.0000'
            c2.alignment = al_c
            c2.border = thin
            xcell(ws2, r2, 3, tag,
                  bold=is_rec, color=fc, bg=bg, align=al_c, brd=thin, size=10)
            r2 += 1

        r2 += 1
        if advisory.declarant_recommendation:
            ws2.merge_cells(f'A{r2}:C{r2}')
            xcell(ws2, r2, 1,
                  f"Declarant's Recommendation: "
                  f"{advisory.declarant_recommendation.upper()}",
                  bold=True, color=DBLUE, bg=LBLUE, align=al_l, size=11)
            r2 += 1
        if advisory.declarant_note:
            ws2.merge_cells(f'A{r2}:C{r2}')
            cn = ws2.cell(row=r2, column=1,
                          value=f'Note: "{advisory.declarant_note}"')
            cn.font = Font(italic=True, color=MGRAY, size=10, name='Calibri')
            cn.fill = xfill(LGRAY)
            cn.alignment = Alignment(
                horizontal='left', vertical='center', wrap_text=True)
            ws2.row_dimensions[r2].height = 36
    else:
        ws2.merge_cells('A5:C5')
        xcell(ws2, 5, 1, 'No WMCDA advisory on file.',
              color=MGRAY, bg=WHITE, align=al_c, size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"R3PCR_{shipment.hawb_number}_ECDT_WMCDA.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── PDF generator (light / navy theme) ───────────────────────────────────────

def _ecdt_pdf(request, shipment, computation, advisory):
    import io
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )

    NAVY   = colors.HexColor('#1E3A6E')
    WHITE  = colors.white
    LGRAY  = colors.HexColor('#F0F4F8')
    DGRAY  = colors.HexColor('#374151')
    MGRAY  = colors.HexColor('#9CA3AF')
    GREEN  = colors.HexColor('#15803D')
    LGREEN = colors.HexColor('#DCFCE7')
    LBLUE  = colors.HexColor('#EFF6FF')
    DBLUE  = colors.HexColor('#1E40AF')
    BORDER = colors.HexColor('#CBD5E1')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1*cm, bottomMargin=1.5*cm,
    )
    W = doc.width

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    p_company  = ps('co', fontSize=18, fontName='Helvetica-Bold',
                    textColor=NAVY, alignment=TA_CENTER, spaceAfter=2)
    p_subtitle = ps('su', fontSize=11, fontName='Helvetica-Bold',
                    textColor=NAVY, alignment=TA_CENTER, spaceAfter=6)
    p_section  = ps('se', fontSize=9,  fontName='Helvetica-Bold',
                    textColor=WHITE, backColor=NAVY, alignment=TA_LEFT,
                    leftIndent=4, leading=16, spaceAfter=2, spaceBefore=8)
    p_lbl      = ps('lb', fontSize=9, fontName='Helvetica-Bold',  textColor=MGRAY)
    p_val      = ps('vl', fontSize=9, fontName='Helvetica',       textColor=DGRAY)
    p_note     = ps('nt', fontSize=7, fontName='Helvetica-Oblique',
                    textColor=MGRAY, alignment=TA_CENTER)

    def hdr8(align=TA_LEFT):
        return ps(f'h8{align}', fontSize=8, fontName='Helvetica-Bold',
                  textColor=WHITE, alignment=align)

    def body8(align=TA_LEFT, bold=False, color=None):
        fn = 'Helvetica-Bold' if bold else 'Helvetica'
        return ps(f'bd8{align}{bold}', fontSize=8, fontName=fn,
                  textColor=color or DGRAY, alignment=align)

    def php(v):
        return f'₱{float(v or 0):,.2f}'

    story = []

    # header
    story.append(Paragraph('RTripleJ Customs Brokerage', p_company))
    story.append(Paragraph('ECDT &amp; WMCDA Computation Sheet', p_subtitle))
    story.append(HRFlowable(width='100%', thickness=2, color=NAVY,
                             spaceAfter=6, spaceBefore=0))

    # disclaimer banner
    AMBER_BG = colors.HexColor('#FEF9C3')
    AMBER_TX = colors.HexColor('#92400E')
    disclaimer_tbl = Table(
        [[Paragraph(
            '<b><i>⚠ ESTIMATED COMPUTATION ONLY. '
            'FINAL ASSESSMENT WILL BE BASED ON CUSTOMS FINDINGS.</i></b>',
            ps('disc', fontSize=8, fontName='Helvetica-BoldOblique',
               textColor=AMBER_TX, alignment=TA_CENTER),
        )]],
        colWidths=[W],
    )
    disclaimer_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), AMBER_BG),
        ('BOX',           (0, 0), (-1, -1), 0.75, colors.HexColor('#FCD34D')),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
    ]))
    story.append(disclaimer_tbl)
    story.append(Spacer(1, 8))

    # info table
    info_rows = [
        [Paragraph(f'<b>{k}</b>', p_lbl), Paragraph(v, p_val)]
        for k, v in _info_block(shipment, computation, request)
    ]
    info_tbl = Table(info_rows, colWidths=[W * 0.28, W * 0.72])
    info_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (0, -1), LGRAY),
        ('BACKGROUND',    (1, 0), (1, -1), WHITE),
        ('GRID',          (0, 0), (-1, -1), 0.5, BORDER),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 10))

    if computation:
        items = computation.get_items()

        # ── line items ────────────────────────────────────────────────────────
        story.append(Paragraph('LINE ITEMS', p_section))
        story.append(Spacer(1, 1))

        cw = [W*0.27, W*0.11, W*0.10, W*0.08, W*0.14, W*0.15, W*0.15]
        item_rows = [[
            Paragraph('<b>DESCRIPTION</b>', hdr8(TA_LEFT)),
            Paragraph('<b>QTY / UNIT</b>',  hdr8(TA_CENTER)),
            Paragraph('<b>HS CODE</b>',      hdr8(TA_CENTER)),
            Paragraph('<b>DUTY %</b>',       hdr8(TA_RIGHT)),
            Paragraph('<b>D/V (PHP)</b>',    hdr8(TA_RIGHT)),
            Paragraph('<b>CUD (PHP)</b>',    hdr8(TA_RIGHT)),
            Paragraph('<b>EXW (USD)</b>',    hdr8(TA_RIGHT)),
        ]]
        for it in items:
            qty_unit = str(it.get('quantity', ''))
            if it.get('unit'):
                qty_unit += f" {it.get('unit')}"
            item_rows.append([
                Paragraph(it.get('description', ''),              body8(TA_LEFT)),
                Paragraph(qty_unit.strip(),                        body8(TA_CENTER)),
                Paragraph(it.get('hs_code') or '—',               body8(TA_CENTER)),
                Paragraph(f"{float(it.get('duty_rate',0) or 0):.2f}%", body8(TA_RIGHT)),
                Paragraph(f"{float(it.get('dv_php', 0) or 0):,.2f}",   body8(TA_RIGHT)),
                Paragraph(f"{float(it.get('cud',    0) or 0):,.2f}",   body8(TA_RIGHT)),
                Paragraph(f"{float(it.get('exw',    0) or 0):,.2f}",   body8(TA_RIGHT)),
            ])

        item_tbl = Table(item_rows, colWidths=cw, repeatRows=1)
        item_style = [
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('GRID',          (0, 0), (-1, -1), 0.5, BORDER),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]
        for i in range(1, len(item_rows)):
            item_style.append((
                'BACKGROUND', (0, i), (-1, i),
                WHITE if i % 2 == 1 else LGRAY,
            ))
        item_tbl.setStyle(TableStyle(item_style))
        story.append(item_tbl)
        story.append(Spacer(1, 10))

        # ── fee summary ───────────────────────────────────────────────────────
        story.append(Paragraph('DUTIES &amp; FEES SUMMARY', p_section))
        story.append(Spacer(1, 1))

        fee_rows = [[
            Paragraph('<b>CHARGE</b>',       hdr8(TA_LEFT)),
            Paragraph('<b>AMOUNT (PHP)</b>', hdr8(TA_RIGHT)),
        ]]

        def pfee(label, amount, bold=False, lc=None, ac=None):
            fn  = 'Helvetica-Bold' if bold else 'Helvetica'
            fee_rows.append([
                Paragraph(label, ps(f'fl{len(fee_rows)}',
                          fontSize=9, fontName=fn, textColor=lc or DGRAY)),
                Paragraph(php(amount), ps(f'fa{len(fee_rows)}',
                          fontSize=9, fontName=fn,
                          textColor=ac or DGRAY, alignment=TA_RIGHT)),
            ])

        pfee('Customs Duty (CUD)',             computation.customs_duty)
        pfee('Value Added Tax — 12% (VAT)',    computation.vat_amount)
        pfee('Import Processing Fee (IPF)',    computation.ipf)
        pfee('Documentary Stamp (CDS)',        130)
        boc_idx = len(fee_rows)
        pfee('BOC Payable  (CUD + VAT + IPF + CDS)',
             computation.boc_payable, bold=True, lc=DBLUE, ac=DBLUE)
        pfee('Brokerage Fee', computation.brokerage_fee)
        if computation.arrastre:      pfee('Arrastre',              computation.arrastre)
        if computation.wharfage:      pfee('Wharfage',              computation.wharfage)
        if computation.bank_charges:  pfee('Bank Charges',          computation.bank_charges)
        if computation.csf_php:       pfee('Container Service Fee', computation.csf_php)
        tlc_idx = len(fee_rows)
        fee_rows.append([
            Paragraph('<b>TOTAL LANDED COST</b>',
                      ps('tll', fontSize=10, fontName='Helvetica-Bold', textColor=GREEN)),
            Paragraph(php(computation.total_landed_cost),
                      ps('tlr', fontSize=10, fontName='Helvetica-Bold',
                         textColor=GREEN, alignment=TA_RIGHT)),
        ])

        fee_tbl = Table(fee_rows, colWidths=[W * 0.65, W * 0.35])
        fee_style = [
            ('BACKGROUND',    (0, 0),       (-1, 0),       NAVY),
            ('GRID',          (0, 0),       (-1, -1),      0.5, BORDER),
            ('VALIGN',        (0, 0),       (-1, -1),      'MIDDLE'),
            ('TOPPADDING',    (0, 0),       (-1, -1),      4),
            ('BOTTOMPADDING', (0, 0),       (-1, -1),      4),
            ('LEFTPADDING',   (0, 0),       (-1, -1),      6),
            ('RIGHTPADDING',  (0, 0),       (-1, -1),      6),
            ('BACKGROUND',    (0, boc_idx), (-1, boc_idx), LBLUE),
            ('LINEABOVE',     (0, boc_idx), (-1, boc_idx), 1.5, DBLUE),
            ('LINEBELOW',     (0, boc_idx), (-1, boc_idx), 1.5, DBLUE),
            ('BACKGROUND',    (0, tlc_idx), (-1, tlc_idx), LGREEN),
            ('LINEABOVE',     (0, tlc_idx), (-1, tlc_idx), 1.5, GREEN),
            ('LINEBELOW',     (0, tlc_idx), (-1, tlc_idx), 1.5, GREEN),
        ]
        for i in range(1, len(fee_rows)):
            if i not in (boc_idx, tlc_idx):
                fee_style.append((
                    'BACKGROUND', (0, i), (-1, i),
                    WHITE if i % 2 == 1 else LGRAY,
                ))
        fee_tbl.setStyle(TableStyle(fee_style))
        story.append(fee_tbl)
        story.append(Spacer(1, 12))

    # ── WMCDA ─────────────────────────────────────────────────────────────────
    if advisory:
        story.append(Paragraph('WMCDA — SHIPPING MODE ADVISORY', p_section))
        story.append(Spacer(1, 1))

        mode_scores = [
            ('Air Freight',               'air',  advisory.air_score),
            ('LCL (Less Container Load)', 'lcl',  advisory.lcl_score),
            ('FCL (Full Container Load)', 'fcl',  advisory.fcl_score),
            ('Land Freight',              'land', advisory.land_score),
        ]
        adv_rows = [[
            Paragraph('<b>MODE</b>',   hdr8(TA_LEFT)),
            Paragraph('<b>SCORE</b>',  hdr8(TA_CENTER)),
            Paragraph('<b>RESULT</b>', hdr8(TA_CENTER)),
        ]]
        adv_bgs = []
        for i, (label, key, score) in enumerate(
            sorted(mode_scores, key=lambda x: (x[2] or 0), reverse=True), start=1
        ):
            is_rec = (key == advisory.recommended_type)
            mc = GREEN if is_rec else DGRAY
            mf = 'Helvetica-Bold' if is_rec else 'Helvetica'
            tag = '★ Recommended' if is_rec else ''
            adv_rows.append([
                Paragraph(label, ps(f'am{i}', fontSize=8, fontName=mf, textColor=mc)),
                Paragraph(f'{float(score or 0):.4f}',
                          ps(f'as{i}', fontSize=8, fontName=mf,
                             textColor=mc, alignment=TA_CENTER)),
                Paragraph(tag,
                          ps(f'at{i}', fontSize=8, fontName=mf,
                             textColor=mc, alignment=TA_CENTER)),
            ])
            adv_bgs.append((i, LGREEN if is_rec else (WHITE if i % 2 == 1 else LGRAY)))

        adv_tbl = Table(adv_rows, colWidths=[W*0.50, W*0.20, W*0.30], repeatRows=1)
        adv_sty = [
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('GRID',          (0, 0), (-1, -1), 0.5, BORDER),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ]
        for row_i, bg in adv_bgs:
            adv_sty.append(('BACKGROUND', (0, row_i), (-1, row_i), bg))
        adv_tbl.setStyle(TableStyle(adv_sty))
        story.append(adv_tbl)
        story.append(Spacer(1, 6))

        if advisory.declarant_recommendation:
            story.append(Paragraph(
                f"<b>Declarant&#8217;s Recommendation:</b> "
                f"{advisory.declarant_recommendation.upper()}",
                ps('dr', fontSize=9, fontName='Helvetica-Bold', textColor=DBLUE),
            ))
        if advisory.declarant_note:
            story.append(Spacer(1, 2))
            story.append(Paragraph(
                f'<i>Note: &ldquo;{advisory.declarant_note}&rdquo;</i>',
                ps('dn', fontSize=8, fontName='Helvetica-Oblique', textColor=MGRAY),
            ))
        story.append(Spacer(1, 10))

    # footer
    story.append(HRFlowable(width='100%', thickness=1, color=NAVY,
                             spaceBefore=4, spaceAfter=4))
    story.append(Paragraph(
        f'Generated by R3-PCR · RTripleJ Customs Brokerage '
        f'· {shipment.hawb_number}',
        p_note,
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"R3PCR_{shipment.hawb_number}_ECDT_WMCDA.pdf"
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Cancel Submission ────────────────────────────────────────────────────────

@login_required
def cancel_submission(request, shipment_id):
    shipment = get_object_or_404(Shipment, id=shipment_id, consignee=request.user)

    if request.method == 'POST':
        age = timezone.now() - shipment.submitted_at
        if shipment.status != 'incoming':
            messages.error(request, 'Cannot cancel — this shipment is already being processed.')
        elif age.total_seconds() > 3600:
            messages.error(request, 'Cannot cancel — the 1-hour cancellation window has passed.')
        else:
            shipment.delete()
            messages.success(request, 'Shipment cancelled and removed.')
            return redirect('consignee:my_submissions')

    return redirect('consignee:my_submissions')
