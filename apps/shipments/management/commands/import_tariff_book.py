"""
Management command: import_tariff_book
=======================================
Reads the full Philippine AHTN 2022 tariff book from the user's local Excel
files (Sections I–XX, one .xlsx file per Chapter) and upserts every HS code
into the HSCode table.

Usage:
    python manage.py import_tariff_book --path "C:/path/to/TARIFF BOOK 2022"

Options:
    --path      Root folder containing the 20 Section sub-folders (required)
    --year      Which duty-rate column to use: 2024a | 2024b | 2025 | 2026 | 2027 | 2028
                Default: 2026
    --dry-run   Print records without writing to DB
    --clear     Delete all existing HSCode rows before importing (fresh seed)
"""

import os
import re
import glob

from django.core.management.base import BaseCommand, CommandError
from apps.shipments.models import HSCode

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ── Constants ─────────────────────────────────────────────────────────────────

# Matches standard AHTN codes (XXXX.XX.XX) and Philippine extended codes
# with a statistical suffix (XXXX.XX.XX.XXX or XXXX.XX.XX.XXXX).
HS_PATTERN = re.compile(r'^\d{4}\.\d{2}\.\d{2}(\.\d+)?$')

# Fallback column indices when dynamic detection fails.
# Standard header layout (has_hdg=True):
#   col: 0=Hdg  1=Code  2=Desc  3=2024a  4=2024b  5=2025  6=2026  7=2027  8=2028
# Headerless continuation (has_hdg=False):
#   col: 0=Code  1=Desc  2=2024a  3=2024b  4=2025  5=2026  6=2027  7=2028

RATE_COLS = {
    (True,  '2024a'): 3,
    (True,  '2024b'): 4,
    (True,  '2025'):  5,
    (True,  '2026'):  6,
    (True,  '2027'):  7,
    (True,  '2028'):  8,
    (False, '2024a'): 2,
    (False, '2024b'): 3,
    (False, '2025'):  4,
    (False, '2026'):  5,
    (False, '2027'):  6,
    (False, '2028'):  7,
}

CHAPTER_NAMES = {
    '01': 'Live Animals',
    '02': 'Meat & Offal',
    '03': 'Fish & Seafood',
    '04': 'Dairy & Eggs',
    '05': 'Other Animal Products',
    '06': 'Live Plants & Flowers',
    '07': 'Vegetables',
    '08': 'Fruits & Nuts',
    '09': 'Coffee, Tea & Spices',
    '10': 'Cereals',
    '11': 'Milling Products',
    '12': 'Oil Seeds & Plants',
    '13': 'Lac, Gums & Resins',
    '14': 'Vegetable Plaiting Materials',
    '15': 'Animal & Vegetable Fats',
    '16': 'Meat & Fish Preparations',
    '17': 'Sugars',
    '18': 'Cocoa & Chocolate',
    '19': 'Pastry & Food Preparations',
    '20': 'Preserved Vegetables & Fruits',
    '21': 'Miscellaneous Food',
    '22': 'Beverages & Spirits',
    '23': 'Residues & Animal Feed',
    '24': 'Tobacco',
    '25': 'Salt, Sulfur & Stone',
    '26': 'Ores & Slag',
    '27': 'Mineral Fuels & Oils',
    '28': 'Inorganic Chemicals',
    '29': 'Organic Chemicals',
    '30': 'Pharmaceutical Products',
    '31': 'Fertilizers',
    '32': 'Dyes, Pigments & Tanning',
    '33': 'Essential Oils & Cosmetics',
    '34': 'Soap & Waxes',
    '35': 'Albuminoids & Glues',
    '36': 'Explosives & Matches',
    '37': 'Photographic Goods',
    '38': 'Miscellaneous Chemicals',
    '39': 'Plastics',
    '40': 'Rubber',
    '41': 'Raw Hides & Skins',
    '42': 'Leather Articles',
    '43': 'Furskins',
    '44': 'Wood & Wood Products',
    '45': 'Cork',
    '46': 'Straw & Basketware',
    '47': 'Wood Pulp',
    '48': 'Paper & Paperboard',
    '49': 'Printed Books & Media',
    '50': 'Silk',
    '51': 'Wool & Fine Hair',
    '52': 'Cotton',
    '53': 'Other Vegetable Fibres',
    '54': 'Manmade Filaments',
    '55': 'Manmade Staple Fibres',
    '56': 'Wadding & Felt',
    '57': 'Carpets & Floor Coverings',
    '58': 'Special Woven Fabrics',
    '59': 'Coated Textiles',
    '60': 'Knitted Fabrics',
    '61': 'Knitted Apparel',
    '62': 'Woven Apparel',
    '63': 'Other Textile Articles',
    '64': 'Footwear',
    '65': 'Headgear',
    '66': 'Umbrellas',
    '67': 'Feathers & Artificial Flowers',
    '68': 'Stone & Cement Articles',
    '69': 'Ceramic Products',
    '70': 'Glass & Glassware',
    '71': 'Precious Stones & Metals',
    '72': 'Iron & Steel',
    '73': 'Iron & Steel Articles',
    '74': 'Copper & Articles',
    '75': 'Nickel',
    '76': 'Aluminum & Articles',
    '78': 'Lead',
    '79': 'Zinc',
    '80': 'Tin',
    '81': 'Other Base Metals',
    '82': 'Tools & Implements',
    '83': 'Misc Metal Articles',
    '84': 'Machinery & Mechanical Appliances',
    '85': 'Electrical Machinery & Equipment',
    '86': 'Railway & Tramway',
    '87': 'Vehicles & Parts',
    '88': 'Aircraft & Spacecraft',
    '89': 'Ships & Boats',
    '90': 'Optical & Medical Instruments',
    '91': 'Clocks & Watches',
    '92': 'Musical Instruments',
    '93': 'Arms & Ammunition',
    '94': 'Furniture & Bedding',
    '95': 'Toys, Games & Sports',
    '96': 'Miscellaneous Articles',
    '97': 'Works of Art',
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_hs_code(val):
    return val and isinstance(val, str) and HS_PATTERN.match(val.strip())


def _to_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(str(val).replace('%', '').strip())
    except (ValueError, TypeError):
        return default


def _clean_desc(val):
    if not val:
        return ''
    return ' '.join(str(val).replace('\n', ' ').split())


def _sheet_has_header(ws):
    """Return True if the sheet has the standard 4-row header block."""
    for row in ws.iter_rows(max_row=5, values_only=True):
        if row[0] and isinstance(row[0], str) and 'Hdg' in str(row[0]):
            return True
    return False


def _find_rate_col_dynamic(ws, year_key):
    """
    Scan the first 6 header rows to locate which column holds the target year
    label. Returns column index (0-based) or None if not found.

    For '2024a': finds the column labelled '2024' (first half-year).
    For '2024b': finds '2024' column + 1 (second half-year, immediately right).
    For all other years: finds the exact year string.
    """
    target = '2024' if year_key in ('2024a', '2024b') else year_key
    offset = 1 if year_key == '2024b' else 0
    for row in ws.iter_rows(max_row=6, values_only=True):
        for j, cell in enumerate(row):
            if cell and str(cell).strip() == target:
                return j + offset
    return None


def _detect_code_col(ws):
    """
    Scan up to 150 rows to find the first column that contains a valid HS code.
    Returns the column index (0-based), or None if no HS code found.
    """
    for row in ws.iter_rows(max_row=150, values_only=True):
        cells = list(row)
        for j, cell in enumerate(cells):
            if cell and isinstance(cell, str) and HS_PATTERN.match(cell.strip()):
                return j
    return None


def extract_from_sheet(ws, year_key='2026'):
    """
    Yield dicts {code, description, duty_rate, chapter} for every valid
    HS code row in the sheet.

    Uses dynamic column detection:
    - code_col  : first column in the sheet that contains a valid HS code
    - rate_col  : column whose header matches year_key (falls back to RATE_COLS)
    - desc_col  : 'Description' header column, or the first non-empty text cell
                  after the code within each row (per-row scan)

    Known limitation: some PDF-exported tables have the rate values on a
    *different row* from the HS code (PDF merged-cell artefact). Those tables
    return 0 codes; this cannot be fixed without manual data cleaning.
    """
    has_hdg  = _sheet_has_header(ws)
    code_col = _detect_code_col(ws)
    if code_col is None:
        return  # No valid HS codes in this sheet

    # Dynamic rate column — fall back to static map if header not found
    rate_col = _find_rate_col_dynamic(ws, year_key)
    if rate_col is None:
        rate_col = RATE_COLS.get((has_hdg, year_key), 6 if has_hdg else 5)

    # Find description column from 'Description' header label
    desc_col_header = None
    for row in ws.iter_rows(max_row=6, values_only=True):
        for j, cell in enumerate(row):
            if cell and isinstance(cell, str) and 'Description' in str(cell):
                desc_col_header = j
                break
        if desc_col_header is not None:
            break

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        raw   = str(cells[code_col]).strip() if len(cells) > code_col and cells[code_col] else ''

        if not _is_hs_code(raw):
            continue

        # Try header-derived desc column first; fall back to per-row scan
        desc = ''
        if desc_col_header is not None and len(cells) > desc_col_header:
            desc = _clean_desc(cells[desc_col_header])
        if not desc:
            # Scan up to 4 columns after code for the first text-like cell
            for offset in range(1, 5):
                j = code_col + offset
                if j >= len(cells):
                    break
                candidate = cells[j]
                if candidate and isinstance(candidate, str):
                    s = candidate.strip()
                    # Accept if it contains letters (not a pure rate number)
                    if s and re.search(r'[A-Za-z\-]', s):
                        desc = _clean_desc(s)
                        break

        if not desc:
            continue   # skip rows with no description (header artifacts)

        rate    = _to_float(cells[rate_col] if len(cells) > rate_col else None)
        chapter = raw[:2]

        yield {
            'code':        raw,
            'description': desc,
            'duty_rate':   rate,
            'chapter':     chapter,
        }


def collect_xlsx_files(root_path):
    """Walk section folders and return sorted list of .xlsx file paths."""
    pattern = os.path.join(root_path, '**', '*.xlsx')
    files   = sorted(glob.glob(pattern, recursive=True))
    return files


# ── Command ───────────────────────────────────────────────────────────────────

class Command(BaseCommand):
    help = 'Import full Philippine AHTN 2022 tariff book from Excel files into HSCode table'

    def add_arguments(self, parser):
        parser.add_argument(
            '--path', type=str, required=True,
            help='Root folder containing the 20 Section sub-folders'
        )
        parser.add_argument(
            '--year', type=str, default='2026',
            choices=['2024a', '2024b', '2025', '2026', '2027', '2028'],
            help='Which year duty rate to import (default: 2026)'
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Preview records without writing to DB'
        )
        parser.add_argument(
            '--clear', action='store_true',
            help='Delete ALL existing HSCode rows before importing (clean slate)'
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('openpyxl is required. Run: pip install openpyxl')

        root      = options['path']
        year_key  = options['year']
        dry_run   = options['dry_run']
        do_clear  = options['clear']

        if not os.path.isdir(root):
            raise CommandError(f'Path not found: {root}')

        xlsx_files = collect_xlsx_files(root)
        if not xlsx_files:
            raise CommandError(f'No .xlsx files found under: {root}')

        self.stdout.write(
            self.style.SUCCESS(f'Found {len(xlsx_files)} Excel file(s). Year: {year_key}. Dry-run: {dry_run}')
        )

        if do_clear and not dry_run:
            deleted, _ = HSCode.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'Cleared {deleted} existing HSCode rows.'))

        total_created  = 0
        total_updated  = 0
        total_skipped  = 0
        total_records  = 0
        current_chapter = None

        for xlsx_path in xlsx_files:
            filename = os.path.basename(xlsx_path)
            try:
                wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  ⚠ Cannot open {filename}: {e}'))
                continue

            file_count = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for record in extract_from_sheet(ws, year_key):
                    total_records += 1
                    file_count    += 1

                    chapter = record['chapter']
                    if chapter != current_chapter:
                        current_chapter = chapter
                        chap_name = CHAPTER_NAMES.get(chapter, '')
                        self.stdout.write(
                            f'  Chapter {chapter} — {chap_name}'
                        )

                    if dry_run:
                        self.stdout.write(
                            f'    [DRY] {record["code"]:14s} {record["duty_rate"]:6.1f}%  '
                            f'{record["description"][:60]}'
                        )
                        continue

                    # Upsert: update description & duty_rate if code already exists
                    obj, created = HSCode.objects.update_or_create(
                        code=record['code'],
                        defaults={
                            'description': record['description'],
                            'duty_rate':   record['duty_rate'],
                            'chapter':     record['chapter'],
                            'is_active':   True,
                        }
                    )
                    if created:
                        total_created += 1
                    else:
                        total_updated += 1

            wb.close()
            self.stdout.write(f'  ✓ {filename} — {file_count} codes')

        # ── Summary ───────────────────────────────────────────────────────────
        self.stdout.write('')
        if dry_run:
            self.stdout.write(
                self.style.SUCCESS(
                    f'DRY RUN complete. Would import {total_records} HS codes from {len(xlsx_files)} file(s).'
                )
            )
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    f'Import complete.\n'
                    f'  Created : {total_created}\n'
                    f'  Updated : {total_updated}\n'
                    f'  Total   : {total_records}\n'
                    f'  Files   : {len(xlsx_files)}'
                )
            )
