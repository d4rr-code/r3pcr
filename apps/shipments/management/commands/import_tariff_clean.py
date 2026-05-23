from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError

from apps.shipments.models import HSCode

try:
    import openpyxl
except ImportError:
    openpyxl = None


DEFAULT_PATH = (
    r'C:\Users\Francis\Documents\Codex\2026-05-23\hello\outputs'
    r'\tariff_book_clean.xlsx'
)
SHEET_NAME = 'Tariff Codes'


def clean_text(value):
    if value is None:
        return ''
    return ' '.join(str(value).strip().split())


def to_decimal(value):
    if value is None or value == '':
        return Decimal('0')
    try:
        return Decimal(str(value).replace('%', '').strip())
    except (InvalidOperation, ValueError):
        return Decimal('0')


class Command(BaseCommand):
    help = 'Import HS codes from the clean tariff workbook.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--path',
            default=DEFAULT_PATH,
            help='Path to tariff_book_clean.xlsx',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview the import without writing to the database.',
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('openpyxl is required. Run: pip install openpyxl')

        workbook_path = options['path']
        dry_run = options['dry_run']

        try:
            workbook = openpyxl.load_workbook(
                workbook_path,
                read_only=True,
                data_only=True,
            )
        except FileNotFoundError:
            raise CommandError(f'Workbook not found: {workbook_path}')
        except Exception as exc:
            raise CommandError(f'Could not open workbook: {exc}')

        if SHEET_NAME not in workbook.sheetnames:
            workbook.close()
            raise CommandError(f'Sheet not found: {SHEET_NAME}')

        worksheet = workbook[SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [clean_text(value) for value in next(rows)]
        except StopIteration:
            workbook.close()
            raise CommandError(f'Sheet is empty: {SHEET_NAME}')

        required = {
            'ahtn_code': 'code',
            'description': 'description',
            'mfn_2026': 'duty_rate',
            'chapter': 'chapter',
        }
        missing = [name for name in required if name not in headers]
        if missing:
            workbook.close()
            raise CommandError(f'Missing required column(s): {", ".join(missing)}')

        indexes = {name: headers.index(name) for name in required}
        skipped = 0
        records = {}

        for row in rows:
            code = clean_text(row[indexes['ahtn_code']])
            description = clean_text(row[indexes['description']])

            if not code or not description:
                skipped += 1
                continue

            records[code] = {
                'description': description,
                'duty_rate': to_decimal(row[indexes['mfn_2026']]),
                'chapter': clean_text(row[indexes['chapter']]),
                'is_active': True,
            }

        workbook.close()
        total = len(records)

        if dry_run:
            self.stdout.write(
                self.style.SUCCESS(
                    f'DRY RUN complete. Would import {total} HS codes; skipped {skipped}.'
                )
            )
            return

        existing_count = HSCode.objects.filter(code__in=records.keys()).count()
        upserts = [
            HSCode(code=code, **defaults)
            for code, defaults in records.items()
        ]
        HSCode.objects.bulk_create(
            upserts,
            batch_size=1000,
            update_conflicts=True,
            update_fields=['description', 'duty_rate', 'chapter', 'is_active'],
            unique_fields=['code'],
        )

        self.stdout.write(
            self.style.SUCCESS(
                f'Import complete. Created: {total - existing_count}. Updated: {existing_count}. '
                f'Skipped: {skipped}. Total processed: {total}.'
            )
        )
